import sys
import os
import re
import logging
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import traceback
# Excel 和圖片處理
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter  # 如果你有用 xlsxwriter 存檔可以留著
from matplotlib.figure import Figure
from PIL import Image

# 使 UI 相關依賴為可選，避免在 FastAPI/無介面環境下匯入失敗
try:
    from PyQt6 import QtWidgets, QtGui, QtCore
    from PyQt6.QtWidgets import QMessageBox, QLabel, QVBoxLayout, QScrollArea, QGridLayout, QPushButton
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas  # for UI charts
    from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
    from PIL.ImageQt import ImageQt
    UI_AVAILABLE = True
except Exception:
    # 定義最小替代，以便在無 UI 環境下仍可匯入本模組
    UI_AVAILABLE = False
    class _Dummy: pass
    class _QtWidgets:
        class QWidget: ...
        class QMainWindow: ...
        class QApplication: ...
        class QCheckBox: ...
        class QPushButton: ...
        class QProgressBar: ...
        class QScrollArea: ...
        class QGridLayout: ...
        class QVBoxLayout: ...
        class QHBoxLayout: ...
        class QTabWidget: ...
        class QStackedWidget: ...
        class QLabel: ...
        class QTableWidget: ...
        class QHeaderView: ...
        class QAbstractItemView: ...
        class QDialog: ...
        class QGroupBox: ...
        class QFormLayout: ...
        class QComboBox: ...
        class QStyle:
            class StandardPixmap: ...
    QtWidgets = _QtWidgets()
    class _QtGui:
        class QFont:
            class Weight:
                Bold = 0
    QtGui = _QtGui()
    class _QtCore:
        class Qt:
            class AlignmentFlag:
                AlignCenter = 0
                AlignLeft = 0
                AlignTop = 0
        class QAbstractItemView: ...
    QtCore = _QtCore()
    class FigureCanvas: ...
    class NavigationToolbar: ...
    class ImageQt: ...
plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'SimHei', 'Arial Unicode MS', 'Noto Sans CJK TC']
plt.rcParams['axes.unicode_minus'] = False  # 正確顯示負號

logger = logging.getLogger(__name__)
logger.setLevel(logging.WARNING)


def _is_finite_number(value) -> bool:
    try:
        return value is not None and pd.notna(value) and np.isfinite(float(value))
    except Exception:
        return False


def load_execution_time(raw_data_file):
    # 檢查檔案是否存在
    if not os.path.exists(raw_data_file):
        print(f" - load_execution_time: File does not exist: {raw_data_file}. Return None.")
        return None

    try:
        # 讀取 Excel 的 'Time' sheet
        # 加上 try/except 處理讀取 Sheet 可能發生的錯誤
        try:
            df = pd.read_excel(raw_data_file, sheet_name='Time', engine='openpyxl')
        except Exception as e:
            print(f" - load_execution_time: Unable to read 'Time' Sheet or file format error: {e}. Return None.")
            return None
      
        # === 修改點：檢查 DataFrame 是否為空 ===
        if df.empty:
            print(" - load_execution_time: 'Time' Sheet is empty. Return None.")
            return None
        # ====================================

        # 確保有 'execTime' 這個欄位
        if 'execTime' not in df.columns:
            print(" - load_execution_time: Cannot find 'execTime' column in 'Time' Sheet. Return None.")
            return None

        # 嘗試從第一列取得 'execTime' 的值
        # 由於上面已經檢查 df.empty，這裡 df.iloc[0] 應該不會再報錯 SINGLE POSITIONAL INDEXER IS OUT OF BOUND
        execution_time_str = df.iloc[0]['execTime']

        # === 修改點：檢查讀取到的值是否為空或無效 ===
        if pd.isna(execution_time_str):
             print(" - load_execution_time: 'execTime' cell is empty or invalid. Return None.")
             return None
        # ========================================

        # 嘗試將字串轉換為時間格式
        # 加上 try/except 處理轉換失敗的錯誤
        try:
            execution_time = pd.to_datetime(execution_time_str, format='%Y-%m-%d %H:%M:%S')
            print(f" - load_execution_time: Successfully read execution time: {execution_time}")
            return execution_time
        except ValueError as e:
            print(f" - load_execution_time: Cannot convert '{execution_time_str}' to datetime: {e}. Return None.")
            return None # 轉換失敗也返回 None

    except Exception as e:
        # 捕捉讀取或處理過程中的其他未知錯誤
        print(f" - load_execution_time: Unknown error occurred while reading execution time: {e}. Return None.")
        return None

def load_chart_information(raw_data_file):
    import pandas as pd
    import openpyxl
    print("Loading chart information...")

    # 自動尋找包含必要欄位的 sheet，不強制要求名稱為 'Chart'
    required_cols = {'GroupName', 'ChartName', 'USL', 'LSL'}
    wb = openpyxl.load_workbook(raw_data_file, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    all_charts_info = None
    for sname in sheet_names:
        df_try = pd.read_excel(raw_data_file, sheet_name=sname, engine='openpyxl')
        if required_cols.issubset(set(df_try.columns)):
            all_charts_info = df_try
            break

    if all_charts_info is None:
        raise ValueError(f"找不到含有 {required_cols} 欄位的 sheet，檔案內 sheets: {sheet_names}")

    expected_columns = ['GroupName', 'ChartName', 'Material_no', 'USL', 'LSL', 'UCL', 'LCL', 'Target', 'ChartID', 'Characteristics']
    for col in expected_columns:
        if col not in all_charts_info.columns:
            raise KeyError(f"Column '{col}' does not exist in chart information")

    return all_charts_info

def preprocess_raw_df(raw_df):
    import numpy as np
    import pandas as pd
    raw_df.replace([np.inf, -np.inf, 'na', 'NA', 'NaN', 'nan'], np.nan, inplace=True)
    required_columns = ['GroupName', 'ChartName', 'point_val', 'Batch_ID', 'point_time']
    missing_columns = [col for col in required_columns if col not in raw_df.columns]
    if missing_columns:
        raise ValueError(f"Missing columns in raw data: {missing_columns}")
    column_types = {
        'GroupName': 'str',
        'ChartName': 'str',
        'point_val': 'float',
        'Batch_ID': 'str',
        'point_time': 'str'
    }
    return raw_df.astype(column_types)

def format_datetime(dt):
    import pandas as pd
    try:
        return pd.to_datetime(dt, format='%Y/%m/%d %H:%M', errors='coerce')
    except Exception as e:
        print(f"Date formatting error: {e}")
        return pd.NaT

def format_and_clean_data(raw_df, chart_info):
    import pandas as pd
    # 性能優化：使用向量化操作代替 apply
    raw_df['point_time'] = pd.to_datetime(
        raw_df['point_time'], 
        format='%Y/%m/%d %H:%M', 
        errors='coerce'
    )
    
    raw_df.dropna(subset=['point_val', 'point_time'], inplace=True)
    if 'CHART_CREATE_TIME' in chart_info and pd.notna(chart_info['CHART_CREATE_TIME']):
        create_time = pd.to_datetime(chart_info['CHART_CREATE_TIME'], format="%m/%d/%Y %I:%M:%S %p", errors='coerce')
        if pd.notna(create_time):
            raw_df = raw_df[raw_df['point_time'] >= create_time]
    return raw_df
def update_chart_limits(raw_df, chart_info):
    import numpy as np
    # 排序並重設索引
    raw_df.sort_values(by='point_time', inplace=True)
    raw_df.reset_index(drop=True, inplace=True)
    
    # 確保必要的欄位存在，並初始化為 NaN
    required_columns = ['usl_val', 'lsl_val', 'ucl_val', 'lcl_val', 'target_val']
    for col in required_columns:
        if col not in raw_df.columns:
            raw_df[col] = np.nan  # 初始化欄位為 NaN
    
    # 使用向量化的方式更新欄位中的 NaN 值
    raw_df[required_columns] = raw_df[required_columns].fillna({
        'usl_val': chart_info['USL'],
        'lsl_val': chart_info['LSL'],
        'ucl_val': chart_info['UCL'],
        'lcl_val': chart_info['LCL'],
        'target_val': chart_info['Target']
    })
    
    # 四捨五入到 8 位小數
    raw_df = raw_df.round(8)
    
    return raw_df, chart_info

def exclude_oos_data(raw_df):
    import pandas as pd
    usl = raw_df['usl_val'].iat[0]
    lsl = raw_df['lsl_val'].iat[0]
    
    if pd.notna(usl) and pd.notna(lsl):
        return raw_df[(raw_df['point_val'] <= usl) & (raw_df['point_val'] >= lsl)]
    elif pd.isna(usl):
        return raw_df[raw_df['point_val'] >= lsl]
    elif pd.isna(lsl):
        return raw_df[raw_df['point_val'] <= usl]
    return raw_df  # 如果都沒有符合條件，則直接回傳原始資料

# 優化後的 preprocess_data 函數
def preprocess_data(chart_info, raw_df):
    try:
        raw_df = format_and_clean_data(raw_df, chart_info)  # 確保這個函數已經是最佳化的
        
        if raw_df.empty:
            return False, None, None, None
        
        raw_df, chart_info = update_chart_limits(raw_df, chart_info)  # 確保這個函數已經是最佳化的

        # OOS 過濾前保存完整資料供繪圖使用（含 OOS 點以便在圖上標記）
        tool_cols_pre = [c for c in ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id'] if c in raw_df.columns]
        extra_cols_pre = [c for c in ['Batch_ID', 'cpk'] if c in raw_df.columns]
        full_df = raw_df[['point_val', 'point_time'] + tool_cols_pre + extra_cols_pre].copy()

        raw_df = exclude_oos_data(raw_df)
        # 保留機台欄位以供後續 by-tool 繪圖使用
        tool_cols = [c for c in ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id'] if c in raw_df.columns]
        keep_cols = ['point_val', 'point_time'] + tool_cols
        raw_df = raw_df[keep_cols]
        
        
        chart_info = chart_info.rename({
            'Material_no': 'material_no', 
            'GroupName': 'group_name',
            'ChartName': 'chart_name'
        })
        
        return True, raw_df, chart_info, full_df
    except ValueError as ve:
        print(f'Skip chart processing due to missing columns: {ve}')
        return False, None, None, None
    except Exception as e:
        print(f'Error occurred during preprocessing: {e}')
        return False, None, None, None

# 優化後的 find_matching_file 函數
def find_matching_file(directory, group_name, chart_name):
    group_name = str(group_name)
    chart_name = str(chart_name)
    
    # 預編譯正則表達式
    pattern = re.compile(rf"{re.escape(group_name)}_{re.escape(chart_name)}(?:_\d+_\d+)?\.csv$")
    
    # 使用列表推導式來提高效率
    matching_files = [
        os.path.join(directory, filename)
        for filename in os.listdir(directory)
        if pattern.match(filename)
    ]
    
    return matching_files[0] if matching_files else None

# 優化後的 get_percentiles 函數
def get_percentiles(values):
    import numpy as np
    values = np.array(values)  # 確保數值是 NumPy 陣列，這樣計算會更快
    return {
        'P05': np.percentile(values, 5),
        'P50': np.percentile(values, 50),
        'P75': np.percentile(values, 75),
        'P25': np.percentile(values, 25),
        'P95': np.percentile(values, 95),
        'P99.865': np.percentile(values, 99.865),
        'P0.135': np.percentile(values, 0.135)
    }

# 優化後的 rolling_calculation 函數
def rolling_calculation(data_values, days_to_roll):
    # 確保數據是 NumPy 陣列，這樣切片會更高效
    data_values = np.array(data_values)
    
    # 滾動數據，取最後 'days_to_roll' 個元素
    return data_values[-days_to_roll:] if len(data_values) >= days_to_roll else data_values

def record_high_low_calculator(current_week_data, historical_data):
    """
    判斷當週數據是否創下歷史新高或新低
    
    Args:
        current_week_data: 當週數據的 point_val 值 (array-like)
        historical_data: 歷史數據的 point_val 值 (array-like)  
    
    Returns:
        dict: 包含 record_high, record_low, highlight_status 的字典
    """
    try:
        # 快速檢查：如果任一數據集為空，直接返回
        if len(current_week_data) == 0 or len(historical_data) == 0:
            return {
                'record_high': False,
                'record_low': False, 
                'highlight_status': 'NO_HIGHLIGHT'
            }
        
        # 性能優化：使用numpy操作，避免Python循環
        current_week_data = np.asarray(current_week_data)
        historical_data = np.asarray(historical_data)
        
        logger.debug(
            "record_high_low inputs | weekly_cnt=%s baseline_cnt=%s weekly_head=%s baseline_head=%s baseline_tail=%s",
            len(current_week_data),
            len(historical_data),
            current_week_data[:5] if len(current_week_data) >= 5 else current_week_data,
            historical_data[:5] if len(historical_data) >= 5 else historical_data,
            historical_data[-5:] if len(historical_data) >= 5 else historical_data,
        )
        
        # 計算當週最高值和最低值 - 使用numpy的快速操作
        current_max = np.max(current_week_data)
        current_min = np.min(current_week_data)
        
        # 計算歷史最高值和最低值 - 使用numpy的快速操作
        historical_max = np.max(historical_data)
        historical_min = np.min(historical_data)
        
        logger.debug(
            "record_high_low extrema | current_max=%.8f historical_max=%.8f current_min=%.8f historical_min=%.8f diff_max=%.8f diff_min=%.8f",
            current_max,
            historical_max,
            current_min,
            historical_min,
            current_max - historical_max,
            current_min - historical_min,
        )
        
        # 檢查當週數據是否包含歷史極值
        current_has_hist_max = np.any(current_week_data == historical_max)
        current_has_hist_min = np.any(current_week_data == historical_min)
        logger.debug(
            "record_high_low contains_hist | has_hist_max=%s has_hist_min=%s",
            current_has_hist_max,
            current_has_hist_min,
        )
        
        # 判斷是否創下新高或新低 - 簡單的數值比較，非常快速
        record_high = current_max > historical_max
        record_low = current_min < historical_min
        
        # 如果創下新高或新低，則需要高亮顯示
        highlight_status = 'HIGHLIGHT' if (record_high or record_low) else 'NO_HIGHLIGHT'
        
        logger.debug(
            "record_high_low result | current_max=%.4f historical_max=%.4f record_high=%s current_min=%.4f historical_min=%.4f record_low=%s highlight=%s",
            current_max,
            historical_max,
            record_high,
            current_min,
            historical_min,
            record_low,
            highlight_status,
        )
        
        return {
            'record_high': record_high,
            'record_low': record_low,
            'highlight_status': highlight_status
        }
        
    except Exception:
        logger.exception("record_high_low_calculator failed")
        return {
            'record_high': False,
            'record_low': False,
            'highlight_status': 'NO_HIGHLIGHT'
        }
def review_kshift_results(results, resolution, characteristic, data_percentiles, base_percentiles):
    # 設定 highlight 的初始值
    highlight_conditions = {key: 'NO_HIGHLIGHT' for key in ['P95_shift', 'P50_shift', 'P05_shift']}

    # 檢查各個百分位數的 k 值是否需要高亮 (簡化版: 絕對差值 > resolution OR K絕對值 > 2)
    for percentile in ['P95', 'P50', 'P05']:
        k_value = results.get(f'{percentile}_k', np.nan) # 使用 .get 安全獲取 K 值 (絕對值)

        # 獲取當前和基線的百分位數，並計算絕對差值 (K 值計算的分子)
        data_p = data_percentiles.get(percentile, np.nan)
        base_p = base_percentiles.get(percentile, np.nan)
        abs_diff = np.nan # 預設絕對差值為 NaN

        if not pd.isna(data_p) and not pd.isna(base_p):
            abs_diff = abs(data_p - base_p)

        # --- 修改點開始：簡化高亮條件 ---

        # 判斷 resolution 是否有效（可選填）
        has_valid_resolution = not pd.isna(resolution) and resolution is not None and resolution > 0

        # 判斷絕對差值是否顯著
        if has_valid_resolution:
            # 有填寫 resolution: 使用 resolution 作為閾值
            is_significant_diff = not pd.isna(abs_diff) and abs_diff >= resolution
        else:
            # 沒填寫 resolution: 只要有差異就算顯著
            is_significant_diff = not pd.isna(abs_diff)

        # 判斷 K 絕對值是否超過 1.67 (且非NaN)
        is_significant_k = not pd.isna(k_value) and abs(k_value) > 1.67

        # 新增判斷 k_value 是否為無限值
        is_infinite_k = not pd.isna(k_value) and np.isinf(abs(k_value))

        # 設定初始高亮: 絕對差值 > resolution 且 (K絕對值 > 1.67 或 K絕對值為無限)
        if is_significant_diff and (is_significant_k or is_infinite_k):  # 使用 AND 和 OR 結合邏輯
            highlight_conditions[f'{percentile}_shift'] = 'HIGHLIGHT'   

        # --- 修改點結束 ---


    # 根據 characteristic 來更新 highlight (這部分邏輯不變，用於取消高亮)
    # 確保訪問 results 和 percentiles 時使用 .get() 和檢查 None/NaN
    if characteristic == 'Bigger':
        # 檢查 data_percentiles 和 base_percentiles 的鍵是否存在且值非空
        if data_percentiles.get('P95') is not None and base_percentiles.get('P05') is not None and data_percentiles['P95'] >= base_percentiles['P05']:
            highlight_conditions['P95_shift'] = 'NO_HIGHLIGHT'
        if data_percentiles.get('P50') is not None and base_percentiles.get('P25') is not None and data_percentiles['P50'] >= base_percentiles['P25']:
            highlight_conditions['P50_shift'] = 'NO_HIGHLIGHT'
        # 檢查 results 的鍵是否存在且值非空
        if results.get('P95_k_ori') is not None and results['P95_k_ori'] >= 0:
            highlight_conditions['P95_shift'] = 'NO_HIGHLIGHT'
        if results.get('P50_k_ori') is not None and results['P50_k_ori'] >= 0:
            highlight_conditions['P50_shift'] = 'NO_HIGHLIGHT'
        if results.get('P05_k_ori') is not None and results['P05_k_ori'] >= 0:
            highlight_conditions['P05_shift'] = 'NO_HIGHLIGHT'

    elif characteristic in ['Smaller', 'Sigma']:  # Sigma 使用與 Smaller 相同的邏輯
        if data_percentiles.get('P05') is not None and base_percentiles.get('P95') is not None and data_percentiles['P05'] <= base_percentiles['P95']:
            highlight_conditions['P05_shift'] = 'NO_HIGHLIGHT'
        if data_percentiles.get('P50') is not None and base_percentiles.get('P75') is not None and data_percentiles['P50'] <= base_percentiles['P75']:
            highlight_conditions['P50_shift'] = 'NO_HIGHLIGHT'
        if results.get('P95_k_ori') is not None and results['P95_k_ori'] <= 0:
            highlight_conditions['P95_shift'] = 'NO_HIGHLIGHT'
        if results.get('P50_k_ori') is not None and results['P50_k_ori'] <= 0:
            highlight_conditions['P50_shift'] = 'NO_HIGHLIGHT'
        if results.get('P05_k_ori') is not None and results['P05_k_ori'] <= 0:
            highlight_conditions['P05_shift'] = 'NO_HIGHLIGHT'

    elif characteristic == 'Nominal':
        if data_percentiles.get('P95') is not None and base_percentiles.get('P95') is not None and data_percentiles['P95'] <= base_percentiles['P95']:
            highlight_conditions['P95_shift'] = 'NO_HIGHLIGHT'
        if data_percentiles.get('P05') is not None and base_percentiles.get('P05') is not None and data_percentiles['P05'] >= base_percentiles['P05']:
            highlight_conditions['P05_shift'] = 'NO_HIGHLIGHT'
        # 檢查 P25, P50, P75 的鍵是否存在且值非空
        if (base_percentiles.get('P25') is not None and
            data_percentiles.get('P50') is not None and
            base_percentiles.get('P75') is not None and
            base_percentiles['P25'] <= data_percentiles['P50'] <= base_percentiles['P75']):
            highlight_conditions['P50_shift'] = 'NO_HIGHLIGHT'
        if results.get('P95_k_ori') is not None and results['P95_k_ori'] <= 0:
            highlight_conditions['P95_shift'] = 'NO_HIGHLIGHT'
        if results.get('P05_k_ori') is not None and results['P05_k_ori'] >= 0:
            highlight_conditions['P05_shift'] = 'NO_HIGHLIGHT'

    return highlight_conditions



def safe_division(numerator, denominator, epsilon=1e-9):
    """
    執行安全除法，避免除以零或極小值。
    如果分母接近零，返回 np.nan。
    """
    if abs(denominator) < epsilon:
        # print(f"  kshift: 警告：嘗試除以零或極小值 ({denominator})") # 可以在需要時打開這行
        return np.nan # 或者返回 float('inf'), 根據您希望在結果中如何表示這種情況
    return np.round(numerator, 8) / denominator


def kshift_sigma_ratio_calculator(base, data, characteristic, resolution, ucl, lcl):
    """
    計算 K-shift 和 Sigma 比例相關指標，並判斷高亮狀態。
    處理週數據點數為 1 時的滾動計算和數據填充。
    加入安全除法避免標準差為零導致的問題。
    """
    print = logger.debug  # route verbose hot-path output through debug logging
    results = {
        'P95_k': np.nan,
        'P50_k': np.nan,
        'P05_k': np.nan,
        # 確保所有 results 的鍵都有初始值，包括 review_kshift_results 返回的
        'P95_k_ori': np.nan,
        'P50_k_ori': np.nan,
        'P05_k_ori': np.nan,
        'P95_shift': 'NO_HIGHLIGHT',
        'P50_shift': 'NO_HIGHLIGHT',
        'P05_shift': 'NO_HIGHLIGHT'
    }

    print("--- 進入 kshift_sigma_ratio_calculator 函數 ---")
    # 這裡 base 和 data 應該是字典，包含 'values' 和 'percentiles' (如果已計算)
    # 根據您之前的調試，analyze_chart 傳入的 base 和 data 是包含 'values', 'cnt', 'mean', 'sigma' 的字典
    # 但 kshift_sigma_ratio_calculator 內部使用了 base['values'] 和 data['values']
    # 這裡假設 base 和 data 是包含 'values' 鍵的字典
    if 'values' not in base or 'values' not in data:
         print("  kshift: 錯誤：輸入數據字典缺少 'values' 鍵。")
         return pd.Series(results)

    data_values = data['values']
    base_values = base['values']

    data_cnt = len(data_values)
    base_cnt = len(base_values) # 也獲取基線數據長度

    print(f"  kshift: 接收到的 data_values shape: {data_values.shape}, base_values shape: {base_values.shape}")
    print(f"  kshift: data_cnt: {data_cnt}, base_cnt: {base_cnt}")

    # 如果週數據少於 1 個點，直接返回預設結果
    if data_cnt < 1:
        print("  kshift: data_cnt < 1, 返回預設結果。")
        return pd.Series(results)

    # 計算基線百分位數。請確保 get_percentiles 能處理 base_cnt = 3 的情況
    try:
        base_percentiles = get_percentiles(base_values)
        print(f"  kshift: 計算出的 base_percentiles (部分): P05={base_percentiles.get('P05')}, P50={base_percentiles.get('P50')}, P95={base_percentiles.get('P95')}")
        # 檢查計算分母所需的關鍵百分位數是否存在且不是 NaN
        if np.isnan(base_percentiles.get('P99.865', np.nan)) or np.isnan(base_percentiles.get('P0.135', np.nan)) or np.isnan(base_percentiles.get('P50', np.nan)):
             print("  kshift: 警告：基線百分位數計算結果無效 (包含 NaN)，可能基線數據不足。無法計算 K 值。")
             return pd.Series(results) # 無法計算分母，返回預設結果

    except Exception as e:
         print(f"  kshift: 計算基線百分位數時發生錯誤: {e}")
         traceback.print_exc()
         return pd.Series(results)


    rolled_data = None # 預設沒有滾動數據
    data_percentiles = None # 預設沒有當前週數據的百分位數

    if data_cnt == 1:
        print("  kshift: 處理 data_cnt == 1 分支 (週數據只有 1 點)")

        # 只借不扣：計算需要從 base_values 借入的點數，不修改原始 base_values
        needed = max(0, 5 - len(data_values))
        if needed > 0 and len(base_values) > 0:
            borrowed_data = base_values[-needed:]  # 取 base_values 最後 needed 個點
            rolled_data_values = np.concatenate((borrowed_data, data_values))
            print(f"  kshift: 簡單切片：從 base_values 借用末尾 {len(borrowed_data)} 點，"
                  f"assembled rolled_data_values shape: {rolled_data_values.shape}")
        else:
            rolled_data_values = np.copy(data_values)
            print(f"  kshift: 無需借用或 base_values 為空，rolled_data_values shape: {rolled_data_values.shape}")

        # 點數不足 5 時無法穩定計算，返回預設值
        if len(rolled_data_values) < 5:
            print(f"  kshift: 警告：base_values 點數不足，無法湊滿 5 點 "
                  f"(實際 {len(rolled_data_values)} 點)。返回預設值。")
            return pd.Series(results)

        # 計算百分位數 (原始單點週數據 & 滾動/填充後數據)
        try:
            data_percentiles = get_percentiles(data_values)  # 原始單點週數據
            print(f"  kshift: 原始週數據 percentiles (data_cnt=1): {data_percentiles}")

            rolled_data_percentiles = get_percentiles(rolled_data_values)  # 滾動後數據
            print(f"  kshift: 滾動數據 percentiles (shape={rolled_data_values.shape}): {rolled_data_percentiles}")

            rolled_data = {'values': rolled_data_values, 'percentiles': rolled_data_percentiles}

            for p in ['P95', 'P50', 'P05']:
                if np.isnan(data_percentiles.get(p, np.nan)):
                    print(f"  kshift: 警告：原始週數據 {p} 百分位數為 NaN。無法計算 K 值。")
                    return pd.Series(results)
                if np.isnan(rolled_data_percentiles.get(p, np.nan)):
                    print(f"  kshift: 警告：滾動數據 {p} 百分位數為 NaN。影響滾動 K 值計算。")

        except Exception as e:
            print(f"  kshift: 計算百分位數時發生錯誤 (data_cnt=1 分支): {e}")
            traceback.print_exc()
            return pd.Series(results)


    elif data_cnt >= 2:
        print(f"  kshift: 處理 data_cnt >= 2 分支, data_cnt: {data_cnt}")
        try:
             data_percentiles = get_percentiles(data_values)
             print(f"  kshift: 當前週數據 percentiles (data_cnt>1): {data_percentiles}")
             # 檢查計算K值所需的當前百分位數是否存在且非NaN
             for p in ['P95', 'P50', 'P05']:
                 if np.isnan(data_percentiles.get(p, np.nan)):
                      print(f"  kshift: 警告：當前週數據 {p} 百分位數為 NaN。無法計算 K 值。")
                      return pd.Series(results)

        except Exception as e:
            print(f"  kshift: 計算百分位數時發生錯誤 (data_cnt>=2 分支): {e}")
            traceback.print_exc()
            return pd.Series(results)


        rolled_data = None # data_cnt >= 2 時，沒有滾動數據的概念用於 highlight 判斷


    else: # 這個分支理論上不會走到，因為開頭已經處理 data_cnt < 1
        print(f"  kshift: Warning: 未預期的 data_cnt 情況: {data_cnt}")
        return pd.Series(results)

    # --- 計算分母 ---
    try:
        # 檢查 UCL/LCL 是否有效（支援單邊規格）
        ucl_valid = not pd.isna(ucl) and ucl is not None
        lcl_valid = not pd.isna(lcl) and lcl is not None

        # 計算百分位數基礎分母值
        p95k_percentile = safe_division(base_percentiles.get('P99.865', np.nan) - base_percentiles.get('P50', np.nan), 3)
        p50k_percentile = safe_division(base_percentiles.get('P99.865', np.nan) - base_percentiles.get('P0.135', np.nan), 6)
        p05k_percentile = safe_division(base_percentiles.get('P50', np.nan) - base_percentiles.get('P0.135', np.nan), 3)

        # P95 分母計算：需要 UCL
        if ucl_valid:
            p95k_ucl = safe_division(ucl - base_percentiles.get('P50', np.nan), 6)
            p95k_deno = np.round(np.max([p95k_percentile, p95k_ucl]), 8)
            print(f"  kshift: P95 分母使用 max(百分位數={p95k_percentile}, UCL計算={p95k_ucl}) = {p95k_deno}")
        else:
            p95k_deno = np.round(p95k_percentile, 8)
            print(f"  kshift: UCL 無效，P95 分母直接使用百分位數 = {p95k_deno}")

        # P50 分母計算：需要 UCL 和 LCL
        if ucl_valid and lcl_valid:
            p50k_ucl_lcl = safe_division(ucl - lcl, 12)
            p50k_deno = np.round(np.max([p50k_percentile, p50k_ucl_lcl]), 8)
            print(f"  kshift: P50 分母使用 max(百分位數={p50k_percentile}, UCL-LCL計算={p50k_ucl_lcl}) = {p50k_deno}")
        else:
            p50k_deno = np.round(p50k_percentile, 8)
            print(f"  kshift: UCL/LCL 無效，P50 分母直接使用百分位數 = {p50k_deno}")

        # P05 分母計算：需要 LCL
        if lcl_valid:
            p05k_lcl = safe_division(base_percentiles.get('P50', np.nan) - lcl, 6)
            p05k_deno = np.round(np.max([p05k_percentile, p05k_lcl]), 8)
            print(f"  kshift: P05 分母使用 max(百分位數={p05k_percentile}, LCL計算={p05k_lcl}) = {p05k_deno}")
        else:
            p05k_deno = np.round(p05k_percentile, 8)
            print(f"  kshift: LCL 無效，P05 分母直接使用百分位數 = {p05k_deno}")

        # YC edit：分母為 0 時的處理邏輯
        if p95k_deno == 0:
            if p05k_deno == 0:
                p95k_deno = p50k_deno
            elif p50k_deno == 0:
                p95k_deno = p05k_deno
            else:
                p95k_deno = min(p50k_deno, p05k_deno)
        if p05k_deno == 0:
            if p95k_deno == 0:
                p05k_deno = p50k_deno
            elif p50k_deno == 0:
                p05k_deno = p95k_deno
            else:
                p05k_deno = min(p50k_deno, p95k_deno)
        if p50k_deno == 0:
            if p95k_deno == 0:
                p50k_deno = p05k_deno
            elif p05k_deno == 0:
                p50k_deno = p95k_deno
            else:
                p50k_deno = min(p05k_deno, p95k_deno)

        denominators = {
            'p95k_deno': p95k_deno,
            'p50k_deno': p50k_deno,
            'p05k_deno': p05k_deno
        }
        print(f"  kshift: 計算出的分母: {denominators}")

        # 檢查分母是否有效 (非 NaN, 非 Inf)
        if np.isnan(p95k_deno) or np.isnan(p50k_deno) or np.isnan(p05k_deno) or np.isinf(p95k_deno) or np.isinf(p50k_deno) or np.isinf(p05k_deno):
            print("  kshift: 警告：計算出的分母無效 (包含 NaN 或 Inf)。無法計算 K 值。")
            return pd.Series(results)

    except Exception as e:
        print(f"  kshift: 計算分母時發生錯誤: {e}")
        traceback.print_exc()
        return pd.Series(results)


    # --- 計算 K 值 ---
    try:
        # 計算 K 值 (原始) - 使用安全除法
        results['P95_k_ori'] = safe_division(np.round(data_percentiles.get('P95', np.nan) - base_percentiles.get('P95', np.nan), 8), p95k_deno)
        results['P50_k_ori'] = safe_division(np.round(data_percentiles.get('P50', np.nan) - base_percentiles.get('P50', np.nan), 8), p50k_deno)
        results['P05_k_ori'] = safe_division(np.round(data_percentiles.get('P05', np.nan) - base_percentiles.get('P05', np.nan), 8), p05k_deno)

        # 計算 K 值 (絕對值) - 使用安全除法
        results['P95_k'] = safe_division(np.round(abs(data_percentiles.get('P95', np.nan) - base_percentiles.get('P95', np.nan)), 8), p95k_deno)
        results['P50_k'] = safe_division(np.round(abs(data_percentiles.get('P50', np.nan) - base_percentiles.get('P50', np.nan)), 8), p50k_deno)
        results['P05_k'] = safe_division(np.round(abs(data_percentiles.get('P05', np.nan) - base_percentiles.get('P05', np.nan)), 8), p05k_deno)

        print(f"  kshift: 計算出的 K 值結果: {results}")

    except Exception as e:
        print(f"  kshift: 計算 K 值時發生錯誤: {e}")
        traceback.print_exc()
        return pd.Series(results)


    # --- 判斷當前高亮條件 ---
    try:
        # 確保傳給 review_kshift_results 的 percentiles 字典是完整的
        current_highlight_conditions = review_kshift_results(results, resolution, characteristic, data_percentiles, base_percentiles)
        print(f"  kshift: current_highlight_conditions: {current_highlight_conditions}")
    except Exception as e:
        print(f"  kshift: 判斷當前高亮條件時發生錯誤: {e}")
        traceback.print_exc()
        # 如果判斷高亮失敗，相關結果可能不準確，但可以返回計算出的 K 值
        current_highlight_conditions = {key: 'ERROR' for key in ['P95_shift', 'P50_shift', 'P05_shift']} # 用 ERROR 標記


    # --- 計算滾動結果高亮條件 (如果存在滾動數據) ---
    rolling_highlight_conditions = {key: 'NO_HIGHLIGHT' for key in ['P95_shift', 'P50_shift', 'P05_shift']}

    if rolled_data is not None:
        print(f"  kshift: 處理 rolled_data != None 分支，rolled_data shape: {rolled_data['values'].shape}")
        print(f"  kshift: 滾動後 base_percentiles: {base_percentiles}")
        try:
            # 計算滾動結果 (K 值) - 使用安全除法
            rolling_results = {
                'P95_k': safe_division(np.round(abs(rolled_data['percentiles'].get('P95', np.nan) - base_percentiles.get('P95', np.nan)), 8), p95k_deno),
                'P50_k': safe_division(np.round(abs(rolled_data['percentiles'].get('P50', np.nan) - base_percentiles.get('P50', np.nan)), 8), p50k_deno),
                'P05_k': safe_division(np.round(abs(rolled_data['percentiles'].get('P05', np.nan) - base_percentiles.get('P05', np.nan)), 8), p05k_deno),
                'P95_k_ori': safe_division(np.round((rolled_data['percentiles'].get('P95', np.nan) - base_percentiles.get('P95', np.nan)), 8), p95k_deno),
                'P50_k_ori': safe_division(np.round((rolled_data['percentiles'].get('P50', np.nan) - base_percentiles.get('P50', np.nan)), 8), p50k_deno),
                'P05_k_ori': safe_division(np.round((rolled_data['percentiles'].get('P05', np.nan) - base_percentiles.get('P05', np.nan)), 8), p05k_deno),
            }
            print(f"  kshift: 計算出的 rolling_results: {rolling_results}")

            # 判斷滾動結果高亮條件
            # 確保傳給 review_kshift_results 的 percentiles 字典是完整的
            rolling_highlight_conditions = review_kshift_results(rolling_results, resolution, characteristic, rolled_data['percentiles'], base_percentiles)
            print(f"  kshift: rolling_highlight_conditions: {rolling_highlight_conditions}")

        except Exception as e:
            print(f"  kshift: 判斷滾動高亮條件時發生錯誤: {e}")
            traceback.print_exc()
            # 如果判斷滾動高亮失敗，用 ERROR 標記
            rolling_highlight_conditions = {key: 'ERROR' for key in ['P95_shift', 'P50_shift', 'P05_shift']}


    # --- 最終的高亮條件 ---
    # 結合當前和滾動的高亮結果
    # 檢查 current_highlight_conditions 和 rolling_highlight_conditions 中的值是否是預期的 'HIGHLIGHT'/'NO_HIGHLIGHT'/'ERROR'
    results['P95_shift'] = 'HIGHLIGHT' if current_highlight_conditions.get('P95_shift') == 'HIGHLIGHT' and (rolled_data is None or rolling_highlight_conditions.get('P95_shift') == 'HIGHLIGHT') else 'NO_HIGHLIGHT'
    results['P50_shift'] = 'HIGHLIGHT' if current_highlight_conditions.get('P50_shift') == 'HIGHLIGHT' and (rolled_data is None or rolling_highlight_conditions.get('P50_shift') == 'HIGHLIGHT') else 'NO_HIGHLIGHT'
    results['P05_shift'] = 'HIGHLIGHT' if current_highlight_conditions.get('P05_shift') == 'HIGHLIGHT' and (rolled_data is None or rolling_highlight_conditions.get('P05_shift') == 'HIGHLIGHT') else 'NO_HIGHLIGHT'

    print(f"  kshift: 最終 shift 結果: P95={results['P95_shift']}, P50={results['P50_shift']}, P05={results['P05_shift']}")
    print("--- 退出 kshift_sigma_ratio_calculator 函數 ---")

    return pd.Series(results)

# 數據類型判斷
def determine_data_type(data_values):
    """
    判斷數據是離散型還是連續型
    
    判斷標準：
    1. (unique數值種類/總樣本數N < 1/3 且 unique數值種類 < 5) OR
    2. (總樣本數N >= 30 且 unique數值種類 <= 10)
    滿足以上任一條件即認定為離散型
    
    Parameters:
    - data_values: 數據值的 numpy array 或 pandas Series
    
    Returns:
    - 'discrete' 或 'continuous'
    """
    import numpy as np
    print = logger.debug  # route verbose hot-path output through debug logging

    # 移除 NaN 值
    clean_values = data_values.dropna() if hasattr(data_values, 'dropna') else data_values[~np.isnan(data_values)]
    
    if len(clean_values) == 0:
        return 'continuous'  # 預設為連續型
    
    unique_values = np.unique(clean_values)
    unique_count = len(unique_values)
    total_count = len(clean_values)
    unique_ratio = unique_count / total_count
    
    print(f"  數據類型判斷: 唯一值數量={unique_count}, 總數量={total_count}, 比例={unique_ratio:.3f}")
    
    # 判斷邏輯：
    # 條件1: unique數值種類/總樣本數N < 1/3 且 unique數值種類 < 5
    condition1 = (unique_ratio <= 1/3) and (unique_count <= 5)
    
    # 條件2: 總樣本數N >= 30 且 unique數值種類 <= 10
    condition2 = (total_count >= 30) and (unique_count <= 10)
    
    if condition1 or condition2:
        print(f"    判定為離散型 - 條件1滿足: {condition1}, 條件2滿足: {condition2}")
        return 'discrete'
    else:
        print(f"    判定為連續型 - 條件1滿足: {condition1}, 條件2滿足: {condition2}")
        return 'continuous'

# OOC計算
def ooc_calculator(data, ucl, lcl):
    data_cnt = len(data)
    ooc_cnt = ((data['point_val'] > ucl) | (data['point_val'] < lcl)).sum()
    ooc_ratio = ooc_cnt / data_cnt if data_cnt != 0 else 0
    return data_cnt, ooc_cnt, ooc_ratio

# OOC結果檢查
def review_ooc_results(ooc_cnt, ooc_ratio, threshold=0.05):
    return 'HIGHLIGHT' if ooc_ratio > threshold and ooc_cnt > 1 else 'NO_HIGHLIGHT'

# 計算Sticking Rate
def sticking_rate_calculator(baseline_data, weekly_data):
    def get_mode(data):
        return data.mode()[0]

    def get_percentage(data, value):
        return (data == value).sum() / len(data)

    # 如果週資料少於10筆，與基線資料進行合併
    if len(weekly_data) < 10:
        rolling_window_size = 20 if len(baseline_data) > 1000 else 10
        weekly_data = pd.concat([baseline_data.tail(rolling_window_size), weekly_data])

    threshold = 0.7
    baseline_mode = get_mode(baseline_data)
    weekly_mode = get_mode(weekly_data)

    baseline_mode_percentage_in_baseline = get_percentage(baseline_data, baseline_mode)
    baseline_mode_percentage_in_weekly = get_percentage(weekly_data, baseline_mode)
    weekly_mode_percentage_in_baseline = get_percentage(baseline_data, weekly_mode)
    weekly_mode_percentage_in_weekly = get_percentage(weekly_data, weekly_mode)

    baseline_mode_diff = abs(baseline_mode_percentage_in_baseline - baseline_mode_percentage_in_weekly)
    weekly_mode_diff = abs(weekly_mode_percentage_in_baseline - weekly_mode_percentage_in_weekly)

    highlight_needed = (baseline_mode_diff >= threshold) or (weekly_mode_diff >= threshold)
    highlight_status = 'HIGHLIGHT' if highlight_needed else 'NO_HIGHLIGHT'

    return {
        'baseline_mode': baseline_mode,
        'weekly_mode': weekly_mode,
        'baseline_mode_percentage_in_baseline': baseline_mode_percentage_in_baseline,
        'baseline_mode_percentage_in_weekly': baseline_mode_percentage_in_weekly,   
        'weekly_mode_percentage_in_baseline': weekly_mode_percentage_in_baseline,
        'weekly_mode_percentage_in_weekly': weekly_mode_percentage_in_weekly,
        'highlight_status': highlight_status
    }

# 趨勢檢查
def trending(raw_df, weekly_start_date, weekly_end_date, baseline_start_date, baseline_end_date):
    # 時間欄位轉換
    raw_df['point_time'] = pd.to_datetime(raw_df['point_time'])
    weekly_end_date = pd.to_datetime(weekly_end_date)
    baseline_start_date = pd.to_datetime(baseline_start_date)
    baseline_end_date = pd.to_datetime(baseline_end_date)

    # [優化] 預過濾 49 天窗口（7週 × 7天）一次性獲取所有需要的數據
    window_start = weekly_end_date - timedelta(days=48)  # 包含最後一天，所以是 48
    weekly_window_df = raw_df[
        (raw_df['point_time'] >= window_start) &
        (raw_df['point_time'] <= weekly_end_date)
    ].copy()

    if weekly_window_df.empty:
        return 'NO_HIGHLIGHT'

    # [優化] 使用 floor division 計算每個數據點屬於哪一週（相對於 weekly_end_date）
    # week_id = 0 表示最近一週，week_id = 6 表示第7週（最早）
    days_from_end = (weekly_end_date - weekly_window_df['point_time']).dt.days
    weekly_window_df['week_id'] = (days_from_end // 7).clip(upper=6)

    # [優化] 一次性分組計算所有週的統計數據
    weekly_grouped = weekly_window_df.groupby('week_id')['point_val'].agg(['median', 'count'])

    # [重要] 使用 reindex 確保所有 7 週都存在，即使沒有數據（填充 NaN 和 0）
    weekly_grouped = weekly_grouped.reindex(range(7))
    weekly_grouped['count'] = weekly_grouped['count'].fillna(0)

    # 提取列表（保持原始順序：week_id 0 = 最新週）
    weekly_medians = weekly_grouped['median'].tolist()
    weekly_counts = weekly_grouped['count'].fillna(0).astype(int).tolist()

    # 檢查最近幾週的資料點數條件
    def check_weeks_condition(weeks_counts):
        if len(weeks_counts) >= 4 and sum(x >= 10 for x in weeks_counts[:4]) >= 3 and weeks_counts[0] >= 10:
            return 4
        elif len(weeks_counts) >= 5 and sum(x >= 6 for x in weeks_counts[:5]) >= 4 and weeks_counts[0] >= 6:
            return 5
        elif len(weeks_counts) >= 6 and sum(x >= 3 for x in weeks_counts[:6]) >= 5 and weeks_counts[0] >= 3:
            return 6
        elif len(weeks_counts) >= 7 and sum(x >= 1 for x in weeks_counts[:7]) >= 6 and weeks_counts[0] >= 1:
            return 7
        return 0

    num_weeks_to_check = check_weeks_condition(weekly_counts)

    if num_weeks_to_check == 0:
        return 'NO_HIGHLIGHT'

    # 趨勢檢查函式
    def is_trending_up(medians):
        return all(earlier > later for earlier, later in zip(medians, medians[1:]))

    def is_trending_down(medians):
        return all(earlier < later for earlier, later in zip(medians, medians[1:]))

    # 基準區間百分位
    baseline_df = raw_df[
        (raw_df['point_time'] >= baseline_start_date) &
        (raw_df['point_time'] <= baseline_end_date)
    ]
    baseline_values = baseline_df['point_val']

    if baseline_values.empty:
        return 'NO_HIGHLIGHT'

    p95 = np.percentile(baseline_values, 95)
    p05 = np.percentile(baseline_values, 5)

    # 檢查是否上升或下降
    check_medians = [m for m in weekly_medians[:num_weeks_to_check] if not np.isnan(m)]

    if len(check_medians) < 2:
        return 'NO_HIGHLIGHT'  # 資料不夠比趨勢

    if is_trending_up(check_medians) and check_medians[0] > p95:
        return 'HIGHLIGHT'
    elif is_trending_down(check_medians) and check_medians[0] < p05:
        return 'HIGHLIGHT'
    return 'NO_HIGHLIGHT'

# 離散型 OOB 處理函數
def discrete_oob_calculator(base_data, weekly_data, chart_info, raw_df=None, 
                            weekly_start_date=None, weekly_end_date=None, 
                            baseline_start_date=None, baseline_end_date=None):
    """
    離散型數據的 OOB 計算方法
    包含修改後的 k-shift、新增的 category_LT_Shift 和 trending
    
    Parameters:
    - base_data: 基線數據字典 (包含 'values', 'cnt', 'mean', 'sigma')
    - weekly_data: 週數據字典 (包含 'values', 'cnt', 'mean', 'sigma')
    - chart_info: 圖表信息
    - raw_df: 原始數據 DataFrame (用於 trending 計算)
    - weekly_start_date: 週開始日期
    - weekly_end_date: 週結束日期
    - baseline_start_date: 基線開始日期
    - baseline_end_date: 基線結束日期

    
    Returns:
    - dict: 包含 OOB 結果的字典
    """
    import numpy as np
    print = logger.debug  # route verbose hot-path output through debug logging
    
    print(f"  離散型 OOB 計算: 基線數據點數={base_data['cnt']}, 週數據點數={weekly_data['cnt']}")
    
    results = {
        'HL_P95_shift': 'NO_HIGHLIGHT',
        'HL_P50_shift': 'NO_HIGHLIGHT', 
        'HL_P05_shift': 'NO_HIGHLIGHT',
        'HL_sticking_shift': 'NO_HIGHLIGHT',
        'HL_trending': 'NO_HIGHLIGHT',
        'HL_high_OOC': 'NO_HIGHLIGHT',
        'HL_category_LT_shift': 'NO_HIGHLIGHT',  # 新增的離散型專用項目
        'discrete_method': True
    }
    
    try:
        # 1. 使用與連續型相同的 sticking_rate_calculator
        print("  離散型 OOB: 計算 sticking rate...")
        sticking_rate_results = sticking_rate_calculator(
            pd.Series(base_data['values']), 
            pd.Series(weekly_data['values'])
        )
        results['HL_sticking_shift'] = sticking_rate_results.get('highlight_status', 'NO_HIGHLIGHT')
        
        # 2. Trending 計算
        print("  離散型 OOB: 計算 trending...")
        if (raw_df is not None and weekly_start_date is not None and 
            weekly_end_date is not None and baseline_start_date is not None and 
            baseline_end_date is not None):
            trending_result = discrete_trending_calculator(
                raw_df, weekly_start_date, weekly_end_date, 
                baseline_start_date, baseline_end_date
            )
            results['HL_trending'] = trending_result
        else:
            results['HL_trending'] = 'NO_HIGHLIGHT'  # 缺少必要參數時不高亮
        
        # 3. 使用與連續型相同的 high_OOC 檢查
        print("  離散型 OOB: 計算 OOC...")
        weekly_df = pd.DataFrame({'point_val': weekly_data['values']})
        ooc_results = ooc_calculator(weekly_df, chart_info.get('UCL'), chart_info.get('LCL'))
        ooc_highlight = review_ooc_results(ooc_results[1], ooc_results[2])
        results['HL_high_OOC'] = ooc_highlight
        
        # 4. 修改後的 k-shift 計算（加入 capping rule）
        print("  離散型 OOB: 計算修改後的 K-shift...")
        kshift_results = discrete_kshift_calculator(
            base_data, weekly_data, 
            chart_info.get('Characteristics'), 
            chart_info.get('Resolution'), 
            chart_info.get('UCL'), 
            chart_info.get('LCL')
        )
        results['HL_P95_shift'] = kshift_results.get('P95_shift', 'NO_HIGHLIGHT')
        results['HL_P50_shift'] = kshift_results.get('P50_shift', 'NO_HIGHLIGHT')
        results['HL_P05_shift'] = kshift_results.get('P05_shift', 'NO_HIGHLIGHT')
        
        # 5. 新增的 category_LT_Shift 計算
        print("  離散型 OOB: 計算 category_LT_Shift...")
        category_lt_results = category_lt_shift_calculator(base_data, weekly_data)
        results['HL_category_LT_shift'] = category_lt_results.get('highlight_status', 'NO_HIGHLIGHT')
        
        print(f"  離散型 OOB 計算完成: {results}")
        
    except Exception:
        logger.exception("discrete_oob_calculator failed")
    
    return results

def discrete_trending_calculator(raw_df, weekly_start_date, weekly_end_date, baseline_start_date, baseline_end_date):
    """
    離散型數據的 trending 計算（移植自原 trending 函數）
    """
    from datetime import timedelta
    import numpy as np
    import pandas as pd
    
    # 時間欄位轉換
    raw_df['point_time'] = pd.to_datetime(raw_df['point_time'])
    weekly_end_date = pd.to_datetime(weekly_end_date)
    baseline_start_date = pd.to_datetime(baseline_start_date)
    baseline_end_date = pd.to_datetime(baseline_end_date)

    # 每週資料的摘要
    weekly_summary = []
    current_end = weekly_end_date
    week_count = 0

    while week_count < 7:
        current_start = current_end - timedelta(days=6)
        week_data = raw_df[
            (raw_df['point_time'] >= current_start) &
            (raw_df['point_time'] <= current_end)
        ]['point_val']

        weekly_summary.append({
            'week_start': current_start,
            'week_end': current_end,
            'median': week_data.median() if not week_data.empty else np.nan,
            'count': len(week_data)
        })

        current_end = current_start - timedelta(days=1)
        week_count += 1

    weekly_data = pd.DataFrame(weekly_summary)

    if weekly_data.empty:
        return 'NO_HIGHLIGHT'

    weekly_medians = weekly_data['median'].tolist()
    weekly_counts = weekly_data['count'].tolist()

    # 檢查最近幾週的資料點數條件
    def check_weeks_condition(weeks_counts):
        if len(weeks_counts) >= 4 and sum(x >= 10 for x in weeks_counts[:4]) >= 3 and weeks_counts[0] >= 10:
            return 4
        elif len(weeks_counts) >= 5 and sum(x >= 6 for x in weeks_counts[:5]) >= 4 and weeks_counts[0] >= 6:
            return 5
        elif len(weeks_counts) >= 6 and sum(x >= 3 for x in weeks_counts[:6]) >= 5 and weeks_counts[0] >= 3:
            return 6
        elif len(weeks_counts) >= 7 and sum(x >= 1 for x in weeks_counts[:7]) >= 6 and weeks_counts[0] >= 1:
            return 7
        return 0

    num_weeks_to_check = check_weeks_condition(weekly_counts)

    if num_weeks_to_check == 0:
        return 'NO_HIGHLIGHT'

    # 趨勢檢查函式
    def is_trending_up(medians):
        return all(earlier > later for earlier, later in zip(medians, medians[1:]))

    def is_trending_down(medians):
        return all(earlier < later for earlier, later in zip(medians, medians[1:]))

    # 基準區間百分位
    baseline_df = raw_df[
        (raw_df['point_time'] >= baseline_start_date) &
        (raw_df['point_time'] <= baseline_end_date)
    ]
    baseline_values = baseline_df['point_val']

    if baseline_values.empty:
        return 'NO_HIGHLIGHT'

    p95 = np.percentile(baseline_values, 95)
    p05 = np.percentile(baseline_values, 5)

    # 檢查是否上升或下降
    check_medians = [m for m in weekly_medians[:num_weeks_to_check] if not np.isnan(m)]

    if len(check_medians) < 2:
        return 'NO_HIGHLIGHT'  # 資料不夠比趨勢

    if is_trending_up(check_medians) and check_medians[0] > p95:
        return 'HIGHLIGHT'
    elif is_trending_down(check_medians) and check_medians[0] < p05:
        return 'HIGHLIGHT'
    return 'NO_HIGHLIGHT'

# 修改後的 K-shift 函數（加入 capping rule）
def discrete_kshift_calculator(base_data, weekly_data, characteristic, resolution, ucl, lcl):
    """
    離散型數據的 K-shift 計算，加入 capping rule
    
    Capping rule: 如果當周點數<=10 且 當周P95/P50/P05沒有超過 baseline的P05和P95範圍外，就不HL
    """
    import numpy as np
    print = logger.debug  # route verbose hot-path output through debug logging

    print("  discrete_kshift: 開始計算離散型 K-shift")
    
    # 先使用原本的 kshift_sigma_ratio_calculator 獲取結果
    kshift_results = kshift_sigma_ratio_calculator(
        base_data, weekly_data, characteristic, resolution, ucl, lcl
    )
    
    weekly_cnt = weekly_data['cnt']
    print(f"  discrete_kshift: 當周點數 = {weekly_cnt}")
    
    # 應用 capping rule
    if weekly_cnt <= 10:
        print("  discrete_kshift: 當周點數 <= 10，檢查 capping rule")
        
        try:
            # 計算當周和基線的百分位數
            weekly_percentiles = get_percentiles(weekly_data['values'])
            base_percentiles = get_percentiles(base_data['values'])
            
            weekly_p95 = weekly_percentiles.get('P95')
            weekly_p50 = weekly_percentiles.get('P50') 
            weekly_p05 = weekly_percentiles.get('P05')
            base_p95 = base_percentiles.get('P95')
            base_p05 = base_percentiles.get('P05')
            
            print(f"  discrete_kshift: 當周百分位數 - P95:{weekly_p95}, P50:{weekly_p50}, P05:{weekly_p05}")
            print(f"  discrete_kshift: 基線範圍 - P05:{base_p05}, P95:{base_p95}")
            
            # 檢查當周百分位數是否都在基線 P05-P95 範圍內
            if (not pd.isna(weekly_p95) and not pd.isna(base_p05) and not pd.isna(base_p95) and
                not pd.isna(weekly_p50) and not pd.isna(weekly_p05)):
                
                p95_in_range = base_p05 <= weekly_p95 <= base_p95
                p50_in_range = base_p05 <= weekly_p50 <= base_p95  
                p05_in_range = base_p05 <= weekly_p05 <= base_p95
                
                print(f"  discrete_kshift: 範圍檢查 - P95 in range:{p95_in_range}, P50 in range:{p50_in_range}, P05 in range:{p05_in_range}")
                
                if p95_in_range and p50_in_range and p05_in_range:
                    print("  discrete_kshift: Capping rule 觸發 - 所有百分位數都在基線範圍內，設為 NO_HIGHLIGHT")
                    kshift_results['P95_shift'] = 'NO_HIGHLIGHT'
                    kshift_results['P50_shift'] = 'NO_HIGHLIGHT' 
                    kshift_results['P05_shift'] = 'NO_HIGHLIGHT'
                else:
                    print("  discrete_kshift: 有百分位數超出基線範圍，維持原始 K-shift 結果")
            else:
                print("  discrete_kshift: 百分位數計算有 NaN 值，維持原始 K-shift 結果")
                
        except Exception as e:
            print(f"  discrete_kshift: Capping rule 檢查時發生錯誤: {e}")
            # 發生錯誤時維持原始結果
    else:
        print("  discrete_kshift: 當周點數 > 10，不適用 capping rule")
    
    print(f"  discrete_kshift: 最終結果 - P95:{kshift_results.get('P95_shift')}, P50:{kshift_results.get('P50_shift')}, P05:{kshift_results.get('P05_shift')}")
    
    return kshift_results

# 新的 category_LT_Shift 函數
def category_lt_shift_calculator(base_data, weekly_data, threshold=0.7):
    """
    計算 category_LT_Shift
    
    邏輯：
    1. 當周<20則rolling to 20筆
    2. 拿當周data範圍去對應baseline同樣data範圍
    3. 檢查data所佔比例是否超過70%
    
    Parameters:
    - base_data: 基線數據字典
    - weekly_data: 週數據字典  
    - threshold: 佔比差異閾值，預設0.7 (70%)
    
    Returns:
    - dict: 包含highlight_status的結果字典
    """
    import numpy as np
    import pandas as pd
    print = logger.debug  # route verbose hot-path output through debug logging

    print("  category_LT_shift: 開始計算")
    
    result = {
        'highlight_status': 'NO_HIGHLIGHT',
        'weekly_range': None,
        'baseline_ratio_in_range': None,
        'weekly_ratio_in_range': None, 
        'ratio_diff': None
    }
    
    try:
        weekly_values = weekly_data['values'].copy()
        base_values = base_data['values'].copy()
        
        print(f"  category_LT_shift: 原始當周點數 = {len(weekly_values)}")
        
        # 1. 如果當周 < 20 則 rolling to 20筆
        if len(weekly_values) < 20:
            print(f"  category_LT_shift: 當周點數 < 20，rolling 到 20 筆")
            
            # 需要從基線數據中補充
            needed_points = 20 - len(weekly_values)
            if len(base_values) >= needed_points:
                # 取基線最後的點來補充
                additional_points = base_values[-needed_points:]
                weekly_values = np.concatenate([additional_points, weekly_values])
                print(f"  category_LT_shift: 補充後當周點數 = {len(weekly_values)}")
            else:
                print(f"  category_LT_shift: 基線數據不足以補充到20筆，使用現有數據")
                weekly_values = np.concatenate([base_values, weekly_values])
        
        # 2. 計算當周數據範圍
        weekly_min = np.min(weekly_values)
        weekly_max = np.max(weekly_values)
        result['weekly_range'] = (weekly_min, weekly_max)
        
        print(f"  category_LT_shift: 當周數據範圍 = [{weekly_min:.3f}, {weekly_max:.3f}]")
        
        # 3. 計算基線數據在此範圍內的比例
        baseline_in_range = base_values[(base_values >= weekly_min) & (base_values <= weekly_max)]
        baseline_ratio = len(baseline_in_range) / len(base_values) if len(base_values) > 0 else 0
        result['baseline_ratio_in_range'] = baseline_ratio
        
        # 4. 計算當周數據在此範圍內的比例（應該是100%，因為就是用當周數據定義的範圍）
        weekly_ratio = 1
        result['weekly_ratio_in_range'] = weekly_ratio
        
        # 5. 計算比例差異
        ratio_diff = abs(weekly_ratio - baseline_ratio)
        result['ratio_diff'] = ratio_diff
        
        print(f"  category_LT_shift: 基線在範圍內比例 = {baseline_ratio:.3f}")
        print(f"  category_LT_shift: 當周在範圍內比例 = {weekly_ratio:.3f}")
        print(f"  category_LT_shift: 比例差異 = {ratio_diff:.3f}")
        
        # 6. 判斷是否需要高亮
        if ratio_diff > threshold:
            result['highlight_status'] = 'HIGHLIGHT'
            print(f"  category_LT_shift: 比例差異 {ratio_diff:.3f} > {threshold}，需要 HIGHLIGHT")
        else:
            result['highlight_status'] = 'NO_HIGHLIGHT' 
            print(f"  category_LT_shift: 比例差異 {ratio_diff:.3f} <= {threshold}，NO_HIGHLIGHT")
            
    except Exception:
        logger.exception("category_lt_shift_calculator failed")
        result['highlight_status'] = 'NO_HIGHLIGHT'
    
    return result


def process_single_chart(chart_info, raw_df, initial_baseline_start_date, baseline_end_date, weekly_start_date, weekly_end_date):
    print = logger.debug  # route verbose hot-path output through debug logging
    print("--- 進入外部 process_single_chart 函數 ---")
    print(f"  接收到的 raw_df shape: {raw_df.shape}")
    print(f"  週數據範圍: {weekly_start_date} 至 {weekly_end_date}")
    # 注意：這裡接收的是 initial_baseline_start_date (通常是往前一年)
    print(f"  初始基線數據範圍 (往前一年): {initial_baseline_start_date} 至 {baseline_end_date}")

    if raw_df is None or raw_df.empty:
        print("  raw_df 是空的或 None, 返回 None")
        return None

    try:
        print("  正在篩選週數據...")
        weekly_data = raw_df[(raw_df['point_time'] >= weekly_start_date) & (raw_df['point_time'] <= weekly_end_date)].copy() # Use copy()
        print(f"  篩選後 weekly_data shape: {weekly_data.shape}")

        if weekly_data.empty:
             print(f'未找到週數據, GroupName: {chart_info.get("group_name", "N/A")}, ChartName: {chart_info.get("chart_name", "N/A")}, 返回 None')
             return None

        # --- 基線數據範圍選擇邏輯開始 ---

        # 步驟 1: 使用初始的一年基線範圍過濾數據並計數
        print("  正在篩選初始一年基線數據...")
        baseline_data_one_year = raw_df[(raw_df['point_time'] >= initial_baseline_start_date) & (raw_df['point_time'] <= baseline_end_date)].copy() # Use copy()
        baseline_count_one_year = len(baseline_data_one_year)
        print(f"  初始一年基線數據點數量: {baseline_count_one_year}")

        # 步驟 2: 根據計數決定最終使用的基線開始日期
        # === 新增：基線數據不足標記 ===
        baseline_insufficient = False
        
        if baseline_count_one_year < 10:
            # 如果少於 10 點，將基線期擴展到兩年
            actual_baseline_start_date = baseline_end_date - pd.Timedelta(days=365 * 2)
            print(f"  基線數據點數量 ({baseline_count_one_year}) < 10，將基線期擴展至兩年: {actual_baseline_start_date} 至 {baseline_end_date}")
                
            # 檢查擴展後的數量
            baseline_data_two_year = raw_df[(raw_df['point_time'] >= actual_baseline_start_date) & (raw_df['point_time'] <= baseline_end_date)].copy()
            baseline_count_two_year = len(baseline_data_two_year)
            print(f"  擴展至兩年後基線數據點數量: {baseline_count_two_year}")
            
            if baseline_count_two_year < 10:
                print(f"  ⚠️  擴展至兩年後仍少於10點 ({baseline_count_two_year})，將跳過 OOB 分析但繼續處理其他功能")
                baseline_insufficient = True
            print(f"  基線數據點數量 ({baseline_count_one_year}) < 10，將基線期擴展至兩年: {actual_baseline_start_date} 至 {baseline_end_date}")
        else:
            # 如果大於等於 10 點，使用一年的基線期
            actual_baseline_start_date = initial_baseline_start_date
            print(f"  基線數據點數量 ({baseline_count_one_year}) >= 10，使用一年基線期: {actual_baseline_start_date} 至 {baseline_end_date}")

        # 步驟 3: 使用最終確定的基線範圍過濾數據
        print("  正在篩選最終基線數據...")
        baseline_data = raw_df[(raw_df['point_time'] >= actual_baseline_start_date) & (raw_df['point_time'] <= baseline_end_date)].copy() # Use copy()
        print(f"  篩選後 baseline_data shape (使用 {len(baseline_data)} 點從 {actual_baseline_start_date} 至 {baseline_end_date}): {baseline_data.shape}")


        # === 修改：基線為空時仍然繼續處理 ===
        baseline_empty = baseline_data.empty
        if baseline_empty:
             print(f'基線數據為空，但仍繼續處理 WE Rule 和圖表生成, GroupName: {chart_info.get("group_name", "N/A")}, ChartName: {chart_info.get("chart_name", "N/A")}')
             baseline_insufficient = True

        # --- 基線數據範圍選擇邏輯結束 ---


        # 計算統計數據（週數據與基線數據）
        def calculate_statistics(data):
             # 新增檢查，避免對只有一個點的數據計算標準差產生 NaN (ddof=1 時)
             if data.shape[0] <= 1:
                  sigma = 0.0 if data.shape[0] == 1 else 0.0 # 單點或零點標準差視為 0
             else:
                  sigma = data['point_val'].std() # ddof=1 是 pandas 預設，計算樣本標準差

             # 如果 sigma 是 NaN (例如，所有值都相同，但數據點多於 1 且少於某個閾值，或計算出問題)
             if np.isnan(sigma):
                 print(f"  calculate_statistics 警告: 計算 sigma 得到 NaN. Data shape: {data.shape}")
                 sigma = 0.0 # 將無效的標準差視為 0

             return {
                 'values': data['point_val'].values,
                 'cnt': data.shape[0],
                 'mean': data['point_val'].mean(),
                 'sigma': sigma # 使用處理過的 sigma
                 }

        print("  正在計算週數據統計...")
        weekly_data_dict = calculate_statistics(weekly_data)
        print(f"  週數據統計結果 (部分): cnt={weekly_data_dict['cnt']}, mean={weekly_data_dict['mean']}, sigma={weekly_data_dict['sigma']}")


        # IMPORTANT: 這裡的 baseline_data_dict 現在是使用 *實際確定* 的基線範圍數據計算的
        print("  正在計算基線數據統計...")
        baseline_data_dict = calculate_statistics(baseline_data) if not baseline_empty else None
        if baseline_data_dict is not None:
            print(f"  基線數據統計結果 (部分): cnt={baseline_data_dict['cnt']}, mean={baseline_data_dict['mean']}, sigma={baseline_data_dict['sigma']}")
        else:
            print("  基線數據為空，跳過基線統計輸出")

        # 確保基線統計數據的標準差不會導致後續計算問題
        if baseline_data_dict and 'sigma' in baseline_data_dict and (baseline_data_dict['sigma'] == 0 or np.isnan(baseline_data_dict['sigma'])):
             print("  警告: 基線標準差為零或無效，可能影響 K 值計算和需要標準差的其他指標。")
             # 您可以選擇在這裡返回 None，或讓後續函數自行處理 NaN/inf

        print("  正在呼叫 kshift_sigma_ratio_calculator...")
        # 傳入使用實際基線範圍計算出的 baseline_data_dict
        # kshift_sigma_ratio_calculator 需要處理 sigma=0 或其他分母為 0 的情況 (已在 safe_division 中處理)
        if not baseline_insufficient and not baseline_empty:
            kshift_results = kshift_sigma_ratio_calculator(baseline_data_dict, weekly_data_dict, chart_info.get('Characteristics'), chart_info.get('Resolution'), chart_info.get('UCL'), chart_info.get('LCL')) # 使用 .get 防止 key 錯誤
        else:
            kshift_results = {'P95_shift': 'NO_HIGHLIGHT', 'P50_shift': 'NO_HIGHLIGHT', 'P05_shift': 'NO_HIGHLIGHT'}

        print(f"  kshift_sigma_ratio_calculator 返回: {kshift_results}")

        print("  正在呼叫 ooc_calculator...")
        # ooc_calculator 使用週數據計算 OOC 點數
        ooc_results = ooc_calculator(weekly_data, chart_info.get('UCL'), chart_info.get('LCL')) # 使用 .get 防止 key 錯誤
        print(f"  ooc_calculator 返回: {ooc_results}")

        print("  正在呼叫 review_ooc_results...")
        ooc_highlight = review_ooc_results(ooc_results[1], ooc_results[2]) # 注意 ooc_results[1] 是 ooc_cnt, ooc_results[2] 是 ooc_points
        print(f"  review_ooc_results 返回: {ooc_highlight}")

        print("  正在呼叫 sticking_rate_calculator...")
        # sticking_rate_calculator 需要週數據和基線數據的 Series
        # IMPORTANT: 這裡傳入的 baseline_data['point_val'] 是使用 *實際確定* 的基線範圍數據
        sticking_rate_results = sticking_rate_calculator(baseline_data['point_val'], weekly_data['point_val']) if not baseline_insufficient and not baseline_empty else {'highlight_status': 'NO_HIGHLIGHT'}
        print(f"  sticking_rate_calculator 返回: {sticking_rate_results}")

        print("  正在呼叫 trending...")
        # trending 也需要使用實際確定後的基線範圍
        trending_results = trending(raw_df, weekly_start_date, weekly_end_date, actual_baseline_start_date, baseline_end_date) if not baseline_insufficient and not baseline_empty else 'NO_HIGHLIGHT'
        print(f"  trending 返回: {trending_results}")

        print("  正在呼叫 record_high_low_calculator...")
        # DEBUG: 輸出時間範圍信息
        print(f"  DEBUG: 基線時間範圍 - 從 {actual_baseline_start_date} 到 {baseline_end_date}")
        print(f"  DEBUG: 當週時間範圍 - 從 {weekly_start_date} 到 {weekly_end_date}")
        print(f"  DEBUG: 基線結束與當週開始間隔 = {weekly_start_date - baseline_end_date}")
        # 計算當週數據是否創下歷史新高或新低（基線先過濾 OOC 點）
        if not baseline_insufficient and not baseline_empty:
            _ucl = chart_info.get('UCL')
            _lcl = chart_info.get('LCL')
            _baseline_for_record = baseline_data.copy()
            if _ucl is not None and not pd.isna(_ucl):
                _baseline_for_record = _baseline_for_record[_baseline_for_record['point_val'] <= _ucl]
            if _lcl is not None and not pd.isna(_lcl):
                _baseline_for_record = _baseline_for_record[_baseline_for_record['point_val'] >= _lcl]
            if _baseline_for_record.empty:
                _baseline_for_record = baseline_data  # fallback：若全部被過濾則還原
            record_results = record_high_low_calculator(weekly_data['point_val'].values, _baseline_for_record['point_val'].values)
        else:
            record_results = {'highlight_status': 'NO_HIGHLIGHT', 'record_high': False, 'record_low': False}
        print(f"  record_high_low_calculator 返回: {record_results}")

        # 判斷是否需要 highlight (任何一個子指標需要高亮，則總體高亮)
        highlight_status = 'HIGHLIGHT' if (
             kshift_results.get('P95_shift') == 'HIGHLIGHT' or
             kshift_results.get('P50_shift') == 'HIGHLIGHT' or
             kshift_results.get('P05_shift') == 'HIGHLIGHT' or
             sticking_rate_results.get('highlight_status') == 'HIGHLIGHT' or
             trending_results == 'HIGHLIGHT' or
             ooc_highlight == 'HIGHLIGHT' or # 應該也要考慮 ooc_highlight
             record_results.get('highlight_status') == 'HIGHLIGHT' # 新增 record high/low 判斷
        ) else 'NO_HIGHLIGHT'
        print(f"  計算出的 highlight_status: {highlight_status}")


        # 組織結果
        # 注意使用 .get(key, default_value) 來安全存取字典鍵，防止 KeyError
        result = {
            'data_cnt': ooc_results[0], # 週數據點數
            'ooc_cnt': ooc_results[1], # 週數據 OOC 點數
            'WE_Rule': '', # 這個欄位在 GUI 類的 build_result 中填充
            'OOB_Rule': '' if not baseline_empty else 'N/A - No Baseline', # 基線為空時標記
            'HL_P95_shift': kshift_results.get('P95_shift', 'N/A'), # 使用 get 並提供預設值，避免 key 錯誤
            'HL_P50_shift': kshift_results.get('P50_shift', 'N/A'),
            'HL_P05_shift': kshift_results.get('P05_shift', 'N/A'),
            'HL_sticking_shift': sticking_rate_results.get('highlight_status', 'N/A'),
            'HL_trending': trending_results, # trending_results 本身就是 HIGHLIGHT/NO_HIGHLIGHT
            'HL_high_OOC': ooc_highlight, # ooc_highlight 本身就是 HIGHLIGHT/NO_HIGHLIGHT
            'HL_record_high_low': record_results.get('highlight_status', 'N/A'), # 新增 record high/low 欄位
            'record_high': record_results.get('record_high', False), # 是否創新高
            'record_low': record_results.get('record_low', False), # 是否創新低
            'Material_no': chart_info.get('material_no', 'N/A'),
            'group_name': chart_info.get('group_name', 'N/A'),
            'chart_name': chart_info.get('chart_name', 'N/A'),
            'chart_ID': chart_info.get('ChartID', 'N/A'),
            'Characteristics': chart_info.get('Characteristics', 'N/A'),
            'USL': chart_info.get('USL', 'N/A'),
            'LSL': chart_info.get('LSL', 'N/A'),
            'UCL': chart_info.get('UCL', 'N/A'),
            'LCL': chart_info.get('LCL', 'N/A'),
            'Target': chart_info.get('Target', 'N/A'),
            'Resolution': chart_info.get('Resolution', 'N/A'),
            'baseline_insufficient': baseline_insufficient,  # 新增標記，供後續使用
            'baseline_empty': baseline_empty  # 新增標記，記錄基線是否為空
            # 可以考慮添加 actual_baseline_start_date 到結果中，用於記錄實際使用的基線範圍
            # 'Actual_Baseline_Start': actual_baseline_start_date
        }
        print("--- 外部 process_single_chart 函數成功退出 ---")
        return result

    except Exception:
        logger.exception(
            "process_single_chart failed | group=%s chart=%s",
            chart_info.get("group_name", "N/A"),
            chart_info.get("chart_name", "N/A"),
        )
        return None

def calculate_sigma(UCL, LCL, mean):
    sigma_upper = (UCL - mean) / 3
    sigma_lower = (mean - LCL) / 3
    return sigma_upper, sigma_lower    
def check_rules(raw_df, chart_info):
    import pandas as pd
    import numpy as np

    mean = chart_info.get('Target')
    UCL = chart_info.get('UCL')
    LCL = chart_info.get('LCL')
    characteristics = chart_info['Characteristics']

    # 計算 sigma（可能返回 NaN）
    sigma_upper, sigma_lower = calculate_sigma(UCL, LCL, mean)

    # 檢查 sigma 是否有效
    sigma_valid = not pd.isna(sigma_upper) and not pd.isna(sigma_lower) and not pd.isna(mean)

    if sigma_valid:
        UWL = mean + 2 * sigma_upper
        LWL = mean - 2 * sigma_lower
    else:
        UWL = np.nan
        LWL = np.nan
        print("  [Warning] check_rules: Sigma 無效，WE2-WE10 將設為 False")

    rules = {
        "WE2": False,
        "WE3": False,
        "WE4": False,
        "WE6": False,
        "WE7": False,
        "WE8": False,
        "WE9": False,
        "WE10": False,
        "CU1": False,
        "CU2": False
    }

    # WE1/WE5: 只需要 UCL/LCL，不依賴 sigma
    if not pd.isna(UCL) and UCL is not None:
        rules["WE1"] = raw_df['point_val'].iloc[-1] > UCL
    if not pd.isna(LCL) and LCL is not None:
        rules["WE5"] = raw_df['point_val'].iloc[-1] < LCL

    # CU1/CU2: 趨勢規則，不依賴 sigma
    if chart_info.get('CU1', 'N') == 'Y' and len(raw_df) >= 7:
        tail_7 = raw_df['point_val'].tail(7)
        diffs = tail_7.diff().dropna()
        rules["CU1"] = bool((diffs > 0).all())

    if chart_info.get('CU2', 'N') == 'Y' and len(raw_df) >= 7:
        tail_7 = raw_df['point_val'].tail(7)
        diffs = tail_7.diff().dropna()
        rules["CU2"] = bool((diffs < 0).all())

    # WE2-WE10 需要 sigma 有效才能判斷
    if not sigma_valid:
        return rules  # Sigma 無效，直接返回（WE1/WE5/CU1/CU2 已判斷）

    if chart_info.get('WE2', 'N') == 'Y' and len(raw_df) >= 3:
        rules["WE2"] = (raw_df['point_val'].tail(3) > UWL).sum() >= 2 if characteristics not in ['Bigger', 'Smaller', 'Sigma'] else False
    if chart_info.get('WE3', 'N') == 'Y' and len(raw_df) >= 5:
        threshold = mean + sigma_upper  # 修正：使用標準的 1σ 線
        rules["WE3"] = (raw_df['point_val'].tail(5) > threshold).sum() >= 4
    if chart_info.get('WE4', 'N') == 'Y' and len(raw_df) >= 8:
        rules["WE4"] = (raw_df['point_val'].tail(8) > mean).all()
    if chart_info.get('WE6', 'N') == 'Y' and len(raw_df) >= 3:
        rules["WE6"] = (raw_df['point_val'].tail(3) < LWL).sum() >= 2 if characteristics not in ['Bigger', 'Smaller', 'Sigma'] else False
    if chart_info.get('WE7', 'N') == 'Y' and len(raw_df) >= 5:
        threshold = mean - sigma_lower  # 修正：使用標準的 1σ 線
        rules["WE7"] = (raw_df['point_val'].tail(5) < threshold).sum() >= 4
    if chart_info.get('WE8', 'N') == 'Y' and len(raw_df) >= 8:
        rules["WE8"] = (raw_df['point_val'].tail(8) < mean).all()
    if chart_info.get('WE9', 'N') == 'Y' and len(raw_df) >= 15:
        # 取得最後 15 筆資料
        tail_points = raw_df['point_val'].tail(15)
        
        # 如果所有資料點報定值（唯一值數量為 1），則直接返回 False
        if tail_points.nunique() == 1:  # 檢查唯一值數量是否為 1
            rules["WE9"] = False
        else:
            # 修正：使用 >= 和 <= 包含邊界值
            condition_result = (tail_points >= (mean - sigma_lower)) & \
                            (tail_points <= (mean + sigma_upper))
            rules["WE9"] = condition_result.all()
    if chart_info.get('WE10', 'N') == 'Y' and len(raw_df) >= 8:                   
        rules["WE10"] = ((raw_df['point_val'].tail(8) < (mean - sigma_lower) ) | 
                        (raw_df['point_val'].tail(8) > (mean + sigma_upper) )).all() if characteristics not in ['Bigger', 'Smaller', 'Sigma'] else False
    return rules
def calculate_cpk(raw_df, chart_info):
    mean = raw_df['point_val'].mean()
    std = raw_df['point_val'].std()
    characteristic = chart_info['Characteristics']
    usl = chart_info.get('USL', None)
    lsl = chart_info.get('LSL', None)

    cpk = None

    if std > 0:
        if characteristic == 'Nominal':
            if usl is not None and lsl is not None:
                cpu = (usl - mean) / (3 * std)
                cpl = (mean - lsl) / (3 * std)
                cpk = min(cpu, cpl)
        elif characteristic in ['Smaller', 'Sigma']:  # Sigma 使用與 Smaller 相同的邏輯
            if usl is not None:
                cpk = (usl - mean) / (3 * std)
        elif characteristic == 'Bigger':
            if lsl is not None:
                cpk = (mean - lsl) / (3 * std)

    if cpk is not None:
        cpk = round(cpk, 3)  # 統一四捨五入到小數第三位

    return {'Cpk': cpk}
def plot_spc_chart(raw_df, chart_info, weekly_start_date, weekly_end_date, debug=False, output_dir: str = 'output'):
    import os
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd

    plt.figure(figsize=(14, 6))

    group_name = chart_info['group_name']
    display_group_name = "" if group_name == "Default" else f"Group: [{group_name}]"
    title = (f"{display_group_name}[{chart_info['chart_name']}][{chart_info['Characteristics']}]\n"
             f"UCL: [{chart_info['UCL']}] | Target: [{chart_info['Target']}] | LCL: [{chart_info['LCL']}]")
    plt.title(title, loc='left', fontsize=12)

    # === 先排序 + reset_index，確保 index = 0..N-1 ===
    raw_df = raw_df.copy()
    raw_df['point_time'] = pd.to_datetime(raw_df['point_time'])
    raw_df = raw_df.sort_values('point_time').reset_index(drop=True)

    # === 使用全部數據，無點數限制 ===
    logger.debug("plot_spc_chart points=%s", len(raw_df))
    original_length = len(raw_df)

    points_num = len(raw_df)
    x_values = np.arange(points_num)

    # === 控制線 ===
    for y_val, text, color in [
        (chart_info.get('UCL'), 'UCL', '#E83F6F'),
        (chart_info.get('Target'), 'Target', '#087E8B'),
        (chart_info.get('LCL'), 'LCL', '#E83F6F'),
    ]:
        if _is_finite_number(y_val):
            y_num = float(y_val)
            plt.hlines(y_num, -0.8, points_num + 2, colors=color, linestyles='--', linewidth=1)
            plt.text(x=points_num + 2, y=y_num, s=text, va='center', ha='left', fontsize=10, color=color)

    # === 畫數據線 ===
    plt.plot(x_values, raw_df['point_val'], color='#5863F8', marker='o', linestyle='-')

    # === 標記 OOS 點（超出 USL/LSL 的點用橘色 × 標示）===
    usl_val = chart_info.get('USL', chart_info.get('UCL'))
    lsl_val = chart_info.get('LSL', chart_info.get('LCL'))
    _oos_mask = pd.Series([False] * len(raw_df))
    if pd.notna(usl_val):
        _oos_mask = _oos_mask | (raw_df['point_val'] > usl_val)
    if pd.notna(lsl_val):
        _oos_mask = _oos_mask | (raw_df['point_val'] < lsl_val)
    if _oos_mask.any():
        plt.scatter(x_values[_oos_mask.values], raw_df['point_val'][_oos_mask.values],
                    marker='x', color='#FF6600', s=120, linewidths=2, zorder=5, label='OOS')

    # === 找當週的 index ===
    ws = pd.to_datetime(weekly_start_date)
    we = pd.to_datetime(weekly_end_date)

    start_index = raw_df[raw_df['point_time'] >= ws].index.min()
    end_index   = raw_df[raw_df['point_time'] <= we].index.max()

    if debug:
        logger.debug("plot_spc_chart window | weekly_start=%s weekly_end=%s", ws, we)
        logger.debug("plot_spc_chart indices | start_index=%s time=%s", start_index, raw_df.loc[start_index, 'point_time'])
        logger.debug("plot_spc_chart indices | end_index=%s time=%s", end_index, raw_df.loc[end_index, 'point_time'])

    # === 檢查 rule，標紅點 ===
    violated_rules = {rule: False for rule in chart_info.get('rule_list', [])}

    for i in range(start_index, end_index + 1):
        data_subset = raw_df.iloc[:i+1].tail(15)
        if not data_subset.empty:
            rules = check_rules(data_subset.copy(), chart_info)
            for rule, violated in rules.items():
                if violated:
                    violated_rules[rule] = True
                    plt.plot(i, raw_df['point_val'].iloc[i], 'ro', markersize=10)

    # === X 軸 ===
    interval = max(1, len(raw_df) // 30)
    plt.xticks(x_values[::interval], raw_df['point_time'].dt.strftime("%Y-%m-%d")[::interval], rotation=90)

    # === 區間上色 ===
    plt.axvspan(start_index, end_index, color='#E83F6F', alpha=0.1, label='Weekly Data')
    if start_index > 0:
        plt.axvspan(0, start_index - 1, color='#3772FF', alpha=0.1, label='Baseline Data')

    plt.xlim([x_values[0] - 1, None])
    plt.legend()

    # === 美化 ===
    ax = plt.gca()
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    plt.tight_layout()

    output_path = output_dir
    os.makedirs(output_path, exist_ok=True)
    safe_group_name = "" if group_name == "Default" else group_name
    image_path = f"{output_path}/SPC_{safe_group_name}_{chart_info['chart_name']}.png"

    try:
        plt.savefig(image_path, bbox_inches='tight')
    finally:
        plt.close()

    return image_path, violated_rules


def plot_weekly_spc_chart(raw_df, chart_info, weekly_start_date, weekly_end_date, debug=False, output_dir: str = 'output'):
    import os
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd

    # work on a local copy，確保 point_time 是 datetime，並依時間排序、重設 index（得到連續的 global index）
    df = raw_df.copy()
    df['point_time'] = pd.to_datetime(df['point_time'])
    df = df.sort_values('point_time').reset_index(drop=True)

    ws = pd.to_datetime(weekly_start_date)
    we = pd.to_datetime(weekly_end_date)

    # 取當週資料（index 保留為 global index）
    df_weekly = df[(df['point_time'] >= ws) & (df['point_time'] <= we)].copy()

    if df_weekly.empty:
        # 若當週沒有資料，仍畫一張空圖以避免例外
        plt.figure(figsize=(14, 6))
        plt.title("No weekly data", loc='left')
        plt.tight_layout()
        output_path = output_dir
        os.makedirs(output_path, exist_ok=True)
        image_path = f'{output_path}/Weekly_SPC_empty.png'
        try:
            plt.savefig(image_path, bbox_inches='tight')
        finally:
            plt.close()
        return image_path

    # 若需要限制點數（可選）
    max_weekly_points = 300
    if len(df_weekly) > max_weekly_points:
        df_weekly = df_weekly.tail(max_weekly_points).copy()

    points_num = len(df_weekly)
    x_values = np.arange(points_num)

    plt.figure(figsize=(14, 6))

    group_name = chart_info.get('group_name', '')
    display_group_name = "" if group_name == "Default" else f"Group: [{group_name}]"
    title = (f"{display_group_name}[{chart_info['chart_name']}][{chart_info['Characteristics']}]\n"
             f"UCL: [{chart_info['UCL']}] | Target: [{chart_info['Target']}] | LCL: [{chart_info['LCL']}]")
    plt.title(title, loc='left', fontsize=12)

    # 繪製控制線（使用 weekly 範圍作為長度參考）
    for y_val, text, color in [
        (chart_info.get('UCL'), 'UCL', '#E83F6F'),
        (chart_info.get('Target'), 'Target', '#087E8B'),
        (chart_info.get('LCL'), 'LCL', '#E83F6F'),
    ]:
        if _is_finite_number(y_val):
            y_num = float(y_val)
            plt.hlines(y_num, -0.8, points_num + 2, colors=color, linestyles='--', linewidth=1)
            plt.text(x=points_num + 2, y=y_num, s=text, va='center', ha='left', fontsize=10, color=color)

    # 畫 weekly 的折線（x 軸使用 0..N-1）
    plt.plot(x_values, df_weekly['point_val'].values, color='#5863F8', marker='o', linestyle='-')

    # === 標記 OOS 點（超出 USL/LSL 的點用橘色 × 標示）===
    _w_usl = chart_info.get('USL', chart_info.get('UCL'))
    _w_lsl = chart_info.get('LSL', chart_info.get('LCL'))
    _w_oos = pd.Series([False] * len(df_weekly))
    if pd.notna(_w_usl):
        _w_oos = _w_oos | (df_weekly['point_val'].values > _w_usl)
    if pd.notna(_w_lsl):
        _w_oos = _w_oos | (df_weekly['point_val'].values < _w_lsl)
    if _w_oos.any():
        plt.scatter(x_values[_w_oos.values], df_weekly['point_val'].values[_w_oos.values],
                    marker='x', color='#FF6600', s=120, linewidths=2, zorder=5, label='OOS')

    # 檢查每一個 weekly 點：用 global index (df_weekly.index) 去取 global 的前 idx+1 筆資料來檢查 rules
    violated_points = []  # 收集觸發的點 (pos_in_weekly, global_index, time, value, rules)
    for pos_in_weekly, (global_idx, row) in enumerate(df_weekly.iterrows()):
        # full_data_subset = 全部原始資料從頭到這個 global index 的最後 15 筆
        full_data_subset = df.iloc[:global_idx + 1].tail(15)
        if full_data_subset.empty:
            continue
        rules = check_rules(full_data_subset.copy(), chart_info)
        if any(rules.values()):
            # 在 weekly plot 的位置畫紅點（x = pos_in_weekly）
            plt.plot(pos_in_weekly, row['point_val'], 'ro', markersize=10)
            violated_points.append((pos_in_weekly, global_idx, row['point_time'], row['point_val'], rules))
            if debug:
                logger.debug(
                    "plot_weekly_spc violation | weekly_pos=%s global_idx=%s time=%s value=%s rules=%s",
                    pos_in_weekly,
                    global_idx,
                    row["point_time"],
                    row["point_val"],
                    {k: v for k, v in rules.items() if v},
                )

    # X axis labels
    interval = max(1, points_num // 30)
    plt.xticks(x_values[::interval], df_weekly['point_time'].dt.strftime("%Y-%m-%d")[::interval], rotation=90)

    plt.axvspan(0, points_num - 1, color='#E83F6F', alpha=0.1, label='Weekly Data')
    plt.xlim([x_values[0] - 1, None])
    plt.legend()

    ax = plt.gca()
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    plt.tight_layout()

    output_path = output_dir
    os.makedirs(output_path, exist_ok=True)
    safe_group_name = "" if group_name == "Default" else group_name
    image_path = f'{output_path}/Weekly_SPC_{safe_group_name}_{chart_info["chart_name"]}.png'
    try:
        plt.savefig(image_path, bbox_inches='tight')
    finally:
        plt.close()

    # 回傳圖片路徑（如需，也可以回傳 violated_points 供 debug 使用）
    return image_path


def plot_qq_plot(raw_df, chart_info, output_dir: str = 'output'):
    """
    繪製 QQ Plot（常態分佈檢定圖）。
    支援多機台分色繪製（偵測 ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id']）。
    回傳輸出的 PNG 檔案路徑。
    """
    import os
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd
    from scipy import stats

    df = raw_df.copy()
    df = df.dropna(subset=['point_val'])

    group_name = chart_info.get('group_name', '')
    chart_name = chart_info.get('chart_name', '')
    display_group_name = "" if group_name == "Default" else f"Group: [{group_name}] "

    possible_tool_cols = ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id']
    tool_col = next((c for c in possible_tool_cols if c in df.columns), None)

    # 判斷是否有多機台
    tools = None
    if tool_col:
        valid_series = df[tool_col].dropna().astype(str).str.strip()
        valid_series = valid_series[valid_series != ""]
        if valid_series.nunique() > 1:
            tools = sorted(valid_series.unique())

    fig, ax = plt.subplots(figsize=(14, 6))
    title = (f"{display_group_name}[{chart_name}][{chart_info.get('Characteristics', '')}]\n"
             f"Q-Q Plot (Normality Check)")
    ax.set_title(title, loc='left', fontsize=12)
    ax.set_xlabel("Theoretical Quantiles")
    ax.set_ylabel("Sample Quantiles")

    if tools is not None:
        colors = plt.cm.tab10(np.linspace(0, 1, len(tools)))
        for i, tool in enumerate(tools):
            mask = df[tool_col].astype(str).str.strip() == tool
            vals = df.loc[mask, 'point_val'].dropna().values
            if len(vals) < 2:
                continue
            (osm, osr), (slope, intercept, _) = stats.probplot(vals, dist='norm')
            ax.scatter(osm, osr, s=15, color=colors[i], label=str(tool), alpha=0.7)
            ax.plot(osm, slope * np.array(osm) + intercept, color=colors[i], linewidth=1)
        ax.legend(title=tool_col, fontsize=8, markerscale=1.5)
    else:
        vals = df['point_val'].values
        if len(vals) >= 2:
            (osm, osr), (slope, intercept, _) = stats.probplot(vals, dist='norm')
            ax.scatter(osm, osr, s=15, color='#5863F8', alpha=0.8)
            ax.plot(osm, slope * np.array(osm) + intercept, color='#E83F6F', linewidth=1.5, label='Normal fit')
            ax.legend(fontsize=8)

    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    fig.tight_layout()

    os.makedirs(output_dir, exist_ok=True)
    safe_group_name = "" if group_name == "Default" else group_name
    image_path = f'{output_dir}/QQ_{safe_group_name}_{chart_name}_qq_plot.png'
    try:
        fig.savefig(image_path, bbox_inches='tight')
    finally:
        plt.close(fig)

    return image_path


# ============================================================================
# 統一輔助函數
# ============================================================================

def get_unified_title(chart_info):
    """
    統一標題生成邏輯
    標題範例：[GroupName][ChartName][Nominal]
    """
    group_name = chart_info.get('group_name', '')
    display_group = f"[{group_name}]" if group_name and group_name != "Default" else ""
    return f"{display_group}[{chart_info['chart_name']}][{chart_info['Characteristics']}]"


def add_right_cl_labels(ax, chart_info, x_pos=None):
    """
    在圖表右側添加 UCL/Target/LCL 標籤（使用 axes transform）
    """
    import pandas as pd
    if x_pos is None:
        x_pos = 1.002
    labels = [
        (chart_info.get('UCL'), 'UCL', '#E83F6F'),
        (chart_info.get('Target'), 'Target', '#087E8B'),
        (chart_info.get('LCL'), 'LCL', '#E83F6F'),
    ]
    for y_val, text, color in labels:
        if _is_finite_number(y_val):
            ax.text(x_pos, float(y_val), text, va='center', ha='left',
                    fontsize=10, color=color, fontweight='bold',
                    transform=ax.get_yaxis_transform())


def add_spc_background_zones(ax, df, weekly_start_date, weekly_end_date):
    """
    在 ax 上添加週區間（紅色）與基線區間（藍色）的背景底色
    """
    import pandas as pd
    try:
        ws = pd.to_datetime(weekly_start_date)
        we = pd.to_datetime(weekly_end_date)
        weekly_mask = (df['point_time'] >= ws) & (df['point_time'] <= we)
        if weekly_mask.any():
            start_idx = df.index[weekly_mask].min()
            end_idx = df.index[weekly_mask].max()
            ax.axvspan(start_idx - 0.5, end_idx + 0.5, color='#E83F6F', alpha=0.15, zorder=-1, label='Weekly')
            if start_idx > 0:
                ax.axvspan(-0.5, start_idx - 0.5, color='#3772FF', alpha=0.15, zorder=-1, label='Baseline')
    except Exception:
        pass


# ============================================================================
# 新增畫圖函數（PNG 輸出版）
# ============================================================================

def plot_spc_chart_interactive(raw_df, chart_info, weekly_start_date, weekly_end_date,
                               record_results=None, debug=False, use_batch_id_labels=False,
                               oob_info="N/A", output_dir: str = 'output'):
    """
    Total SPC 完整趨勢圖（PNG 版）
    與 plot_spc_chart 功能相同，但使用統一標題格式與輔助函數，
    子標題顯示 "Total Trend Analysis"。
    回傳 (image_path, violated_rules)
    """
    import os
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd

    raw_df = raw_df.copy()
    raw_df['point_time'] = pd.to_datetime(raw_df['point_time'])
    raw_df = raw_df.sort_values('point_time').reset_index(drop=True)

    print(f"  plot_spc_chart_interactive: 數據點數 {len(raw_df)}")

    points_num = len(raw_df)
    x_values = np.arange(points_num)

    fig, ax = plt.subplots(figsize=(14, 6))

    unified_title = get_unified_title(chart_info)
    ax.set_title(f"{unified_title}\nTotal Trend Analysis", loc='left', fontsize=10)

    # 控制線
    for y_val, color in [(chart_info.get('UCL'), '#E83F6F'),
                          (chart_info.get('Target'), '#087E8B'),
                          (chart_info.get('LCL'), '#E83F6F')]:
        if _is_finite_number(y_val):
            ax.hlines(float(y_val), -0.8, points_num + 2, colors=color, linestyles='--', linewidth=1)
    add_right_cl_labels(ax, chart_info)

    # 背景底色
    add_spc_background_zones(ax, raw_df, weekly_start_date, weekly_end_date)

    # 主折線
    ax.plot(x_values, raw_df['point_val'], color='#5863F8', marker='o', linestyle='-', markersize=4)

    # 找當週 index
    ws = pd.to_datetime(weekly_start_date)
    we = pd.to_datetime(weekly_end_date)
    start_index = raw_df[raw_df['point_time'] >= ws].index.min()
    end_index = raw_df[raw_df['point_time'] <= we].index.max()

    if debug:
        print(f"[DEBUG] start_index={start_index}, end_index={end_index}")

    # 檢查 rule，標紅點
    violated_rules = {rule: False for rule in chart_info.get('rule_list', [])}
    if pd.notna(start_index) and pd.notna(end_index):
        for i in range(int(start_index), int(end_index) + 1):
            data_subset = raw_df.iloc[:i + 1].tail(15)
            if not data_subset.empty:
                rules = check_rules(data_subset.copy(), chart_info)
                for rule, violated in rules.items():
                    if violated:
                        violated_rules[rule] = True
                        ax.plot(i, raw_df['point_val'].iloc[i], 'ro', markersize=6)

    # X 軸
    interval = max(1, len(raw_df) // 30)
    ax.set_xticks(x_values[::interval])
    if use_batch_id_labels and 'Batch_ID' in raw_df.columns:
        ax.set_xticklabels(raw_df['Batch_ID'].astype(str)[::interval], rotation=90, fontsize=7)
    else:
        ax.set_xticklabels(raw_df['point_time'].dt.strftime("%Y-%m-%d")[::interval], rotation=90, fontsize=7)

    ax.set_xlim([x_values[0] - 1, None])
    ax.legend(loc='upper left', fontsize=7)
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    plt.tight_layout()

    output_path = output_dir
    os.makedirs(output_path, exist_ok=True)
    group_name = chart_info.get('group_name', '')
    safe_group_name = "" if group_name == "Default" else group_name
    image_path = f"{output_path}/SPC_Interactive_{safe_group_name}_{chart_info['chart_name']}.png"
    try:
        plt.savefig(image_path, bbox_inches='tight')
    finally:
        plt.close(fig)

    return image_path, violated_rules


def plot_spc_by_tool_color(raw_df, chart_info, weekly_start_date, weekly_end_date, oob_info="N/A", output_dir: str = 'output'):
    """
    Total SPC By Tool（顏色區分），PNG 版本。
    各機台以不同顏色的點標示，背景底色同 plot_spc_chart_interactive。
    自動偵測機台欄位 ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id']。
    回傳 image_path
    """
    import os
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd

    df = raw_df.copy()
    df['point_time'] = pd.to_datetime(df['point_time'])
    df = df.sort_values('point_time').reset_index(drop=True)

    # 自動偵測機台欄位
    possible_tool_cols = ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id']
    tool_col = next((c for c in possible_tool_cols if c in df.columns), None)
    if tool_col is None:
        df['ByTool'] = 'Unknown'
        tool_col = 'ByTool'
    df[tool_col] = df[tool_col].fillna('Unknown').astype(str)

    unique_tools = sorted(df[tool_col].unique())
    cmap = plt.cm.tab10
    tool_color_map = {tool: cmap(i % 10) for i, tool in enumerate(unique_tools)}

    fig, ax = plt.subplots(figsize=(14, 6))

    group_name = chart_info.get('group_name', '')
    display_group_name = "" if group_name == "Default" else f"Group: [{group_name}]"
    title = (f"{display_group_name}[{chart_info['chart_name']}][{chart_info['Characteristics']}]\n"
             f"UCL: [{chart_info.get('UCL')}] | Target: [{chart_info.get('Target')}] | LCL: [{chart_info.get('LCL')}]")
    ax.set_title(title, loc='left', fontsize=12)

    # 背景底色（weekly / baseline）
    add_spc_background_zones(ax, df, weekly_start_date, weekly_end_date)

    # 灰色連接線（時間順序）
    ax.plot(df.index, df['point_val'], color='#696969', alpha=0.4, zorder=1)

    # 各機台彩色點
    for tool in unique_tools:
        subset = df[df[tool_col] == tool]
        ax.plot(subset.index, subset['point_val'], marker='o', linestyle='',
                color=tool_color_map[tool], label=tool, markersize=5, zorder=3)

    # 控制線
    for y_val, color in [(chart_info.get('UCL'), '#E83F6F'),
                          (chart_info.get('Target'), '#087E8B'),
                          (chart_info.get('LCL'), '#E83F6F')]:
        if _is_finite_number(y_val):
            ax.hlines(float(y_val), -0.5, len(df), colors=color, linestyles='--', linewidth=1, zorder=2)
    add_right_cl_labels(ax, chart_info)

    # X 軸：顯示 point_time 日期，採樣間隔避免標籤重疊
    n = len(df)
    interval = max(1, n // 30)
    tick_pos = df.index[::interval]
    tick_labels = df['point_time'].dt.strftime('%Y-%m-%d').iloc[::interval]
    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_labels, rotation=90)

    ax.legend(loc='upper left', fontsize=8, ncol=4)
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    fig.tight_layout()

    os.makedirs(output_dir, exist_ok=True)
    safe_group_name = "" if group_name == "Default" else group_name
    image_path = f"{output_dir}/SPC_ByToolColor_{safe_group_name}_{chart_info['chart_name']}.png"
    try:
        fig.savefig(image_path, bbox_inches='tight')
    finally:
        plt.close(fig)

    return image_path


def plot_spc_by_tool_group(raw_df, chart_info, oob_info="N/A", output_dir: str = 'output'):
    """
    Total SPC By Tool（水平分組），PNG 版本。
    各機台資料按 [機台, point_time] 排列在 X 軸，機台邊界繪製虛線。
    自動偵測機台欄位 ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id']。
    回傳 image_path
    """
    import os
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd

    df = raw_df.copy()
    df['point_time'] = pd.to_datetime(df['point_time'])

    # 自動偵測機台欄位
    possible_tool_cols = ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id']
    tool_col = next((c for c in possible_tool_cols if c in df.columns), None)
    if tool_col is None:
        df['ByTool'] = 'Unknown'
        tool_col = 'ByTool'
    df[tool_col] = df[tool_col].fillna('Unknown').astype(str)

    # 依 [機台欄位, point_time] 排序並 reset_index
    df = df.sort_values([tool_col, 'point_time']).reset_index(drop=True)

    unique_tools = sorted(df[tool_col].unique())
    cmap = plt.cm.tab10
    tool_color_map = {tool: cmap(i % 10) for i, tool in enumerate(unique_tools)}

    fig, ax = plt.subplots(figsize=(14, 6))

    group_name = chart_info.get('group_name', '')
    display_group_name = "" if group_name == "Default" else f"Group: [{group_name}]"
    title = (f"{display_group_name}[{chart_info['chart_name']}][{chart_info['Characteristics']}]\n"
             f"UCL: [{chart_info.get('UCL')}] | Target: [{chart_info.get('Target')}] | LCL: [{chart_info.get('LCL')}]")
    ax.set_title(title, loc='left', fontsize=12)

    for i, tool in enumerate(unique_tools):
        subset = df[df[tool_col] == tool]
        ax.plot(subset.index, subset['point_val'], marker='o', markersize=4,
                color=tool_color_map[tool], label=tool, alpha=0.8, zorder=3)
        # 機台邊界虛線（分群感）
        if i > 0:
            boundary = subset.index.min() - 0.5
            ax.axvline(x=boundary, color='gray', linestyle='--', linewidth=1, alpha=0.5, zorder=1)

    # 控制線
    for y_val, color in [(chart_info.get('UCL'), '#E83F6F'),
                          (chart_info.get('Target'), '#087E8B'),
                          (chart_info.get('LCL'), '#E83F6F')]:
        if _is_finite_number(y_val):
            ax.hlines(float(y_val), -0.5, len(df), colors=color, linestyles='--', linewidth=1, zorder=2)
    add_right_cl_labels(ax, chart_info)

    # X 軸：顯示 point_time 日期，採樣間隔避免標籤重疊
    n = len(df)
    interval = max(1, n // 30)
    tick_pos = df.index[::interval]
    tick_labels = df['point_time'].dt.strftime('%Y-%m-%d').iloc[::interval]
    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_labels, rotation=90)

    ax.legend(loc='upper left', fontsize=8, ncol=4)
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    fig.tight_layout()

    os.makedirs(output_dir, exist_ok=True)
    safe_group_name = "" if group_name == "Default" else group_name
    image_path = f"{output_dir}/SPC_ByToolGroup_{safe_group_name}_{chart_info['chart_name']}.png"
    try:
        fig.savefig(image_path, bbox_inches='tight')
    finally:
        plt.close(fig)

    return image_path


def plot_weekly_spc_chart_interactive(raw_df, chart_info, weekly_start_date, weekly_end_date,
                                      record_results=None, debug=False, use_batch_id_labels=False,
                                      oob_info="N/A", output_dir: str = 'output'):
    """
    Weekly SPC 圖表（PNG 版）
    與 plot_weekly_spc_chart 功能相同，但使用統一標題格式，
    子標題顯示 "Weekly Trend Analysis"。
    回傳 image_path
    """
    import os
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd

    df = raw_df.copy()
    df['point_time'] = pd.to_datetime(df['point_time'])
    df = df.sort_values('point_time').reset_index(drop=True)

    ws = pd.to_datetime(weekly_start_date)
    we = pd.to_datetime(weekly_end_date)
    df_weekly = df[(df['point_time'] >= ws) & (df['point_time'] <= we)].copy()

    output_path = output_dir
    os.makedirs(output_path, exist_ok=True)
    group_name = chart_info.get('group_name', '')
    safe_group_name = "" if group_name == "Default" else group_name
    image_path = f"{output_path}/Weekly_SPC_Interactive_{safe_group_name}_{chart_info['chart_name']}.png"

    if df_weekly.empty:
        fig, ax = plt.subplots(figsize=(14, 6))
        ax.set_title("No weekly data", loc='left')
        plt.tight_layout()
        try:
            plt.savefig(image_path, bbox_inches='tight')
        finally:
            plt.close(fig)
        return image_path

    print(f"  plot_weekly_spc_chart_interactive: 週數據點數 {len(df_weekly)}")

    points_num = len(df_weekly)
    x_values = np.arange(points_num)

    fig, ax = plt.subplots(figsize=(14, 6))

    unified_title = get_unified_title(chart_info)
    ax.set_title(f"{unified_title}\nWeekly Trend Analysis", loc='left', fontsize=10)

    # 控制線
    for y_val, color in [(chart_info.get('UCL'), '#E83F6F'),
                          (chart_info.get('Target'), '#087E8B'),
                          (chart_info.get('LCL'), '#E83F6F')]:
        if _is_finite_number(y_val):
            ax.hlines(float(y_val), -0.8, points_num + 2, colors=color, linestyles='--', linewidth=1)
    add_right_cl_labels(ax, chart_info)

    # 週資料底色（全紅）
    ax.axvspan(0, max(points_num - 1, 0.5), color='#E83F6F', alpha=0.08, label='Weekly')

    # 主折線
    ax.plot(x_values, df_weekly['point_val'].values, color='#5863F8', marker='o', linestyle='-', markersize=4)

    # 檢查 rule，標紅點
    violated_points = []
    for pos_in_weekly, (global_idx, row) in enumerate(df_weekly.iterrows()):
        full_data_subset = df.iloc[:global_idx + 1].tail(15)
        if full_data_subset.empty:
            continue
        rules = check_rules(full_data_subset.copy(), chart_info)
        if any(rules.values()):
            ax.plot(pos_in_weekly, row['point_val'], 'ro', markersize=6)
            violated_points.append((pos_in_weekly, global_idx, row['point_time'], row['point_val'], rules))
            if debug:
                print(f'[VIOL] weekly_pos={pos_in_weekly} global_idx={global_idx} '
                      f'time={row["point_time"]} value={row["point_val"]} '
                      f'rules={ {k: v for k, v in rules.items() if v} }')

    # X 軸
    interval = max(1, points_num // 30)
    ax.set_xticks(x_values[::interval])
    if use_batch_id_labels and 'Batch_ID' in df_weekly.columns:
        ax.set_xticklabels(df_weekly['Batch_ID'].astype(str)[::interval], rotation=90, fontsize=7)
    else:
        ax.set_xticklabels(df_weekly['point_time'].dt.strftime("%Y-%m-%d")[::interval], rotation=90, fontsize=7)

    ax.set_xlim([x_values[0] - 1, None])
    ax.legend(loc='upper left', fontsize=7)
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    plt.tight_layout()
    try:
        plt.savefig(image_path, bbox_inches='tight')
    finally:
        plt.close(fig)

    return image_path


def save_results_to_excel(results_df, scale_factor=0.3, output_path: str = 'result_with_images.xlsx'):
    def _valid_image_path(path_value):
        if path_value is None:
            return None
        path_str = str(path_value).strip()
        if not path_str or path_str.upper() in {"N/A", "NAN", "NONE", "<NA>"}:
            return None
        return path_str if os.path.isfile(path_str) else None

    results_df['group_name'] = results_df['group_name'].replace("Default", "")  # 替換 Default 為空白

    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet()

    cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_name': 'Arial', 'font_size': 10})
    header_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_name': 'Arial', 'font_size': 12, 'bold': True})

    col_widths = {}

    for col_idx, header in enumerate(results_df.columns):
        worksheet.write(0, col_idx + 2, header, header_format)
        col_widths[col_idx + 2] = max(len(header), col_widths.get(col_idx + 2, 0))

    max_image_height = 0
    image_column_width = 0

    for row_idx, row in enumerate(results_df.itertuples(index=False), start=1):
        img_path = _valid_image_path(row.chart_path)
        weekly_spc_chart_path = _valid_image_path(row.weekly_chart_path)

        x_offset = 0
        y_offset = 10
        options = {
            'x_scale': scale_factor,
            'y_scale': scale_factor,
            'x_offset': x_offset,
            'y_offset': y_offset,
            'object_position': 1
        }

        for image_col, valid_path in ((0, img_path), (1, weekly_spc_chart_path)):
            if not valid_path:
                continue
            worksheet.insert_image(row_idx, image_col, valid_path, options)
            with Image.open(valid_path) as img:
                image_width, image_height = img.size
            scaled_width = image_width * scale_factor
            scaled_height = image_height * scale_factor

            if scaled_height > max_image_height:
                max_image_height = scaled_height
            if scaled_width > image_column_width:
                image_column_width = scaled_width

        for col_idx, value in enumerate(row, start=1):
            worksheet.write(row_idx, col_idx + 1, value, cell_format)
            col_widths[col_idx + 1] = max(col_widths.get(col_idx + 1, 0), len(str(value)))

    worksheet.set_column(0, 0, image_column_width / 7)
    worksheet.set_column(1, 1, image_column_width / 7)

    for col_idx, width in col_widths.items():
        worksheet.set_column(col_idx, col_idx, width + 5)

    for row_idx in range(1, len(results_df) + 1):
        worksheet.set_row(row_idx, max_image_height)

    workbook.close()


# 🔧 封裝路徑處理函式
def resource_path(relative_path):
    if getattr(sys, 'frozen', False):  # 如果是打包環境
        base_path = os.path.dirname(sys.executable)
    else:  # 開發環境
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# 常數定義
HEADERS = ["Total Chart", "Weekly Chart", "Chart Info."]
OOB_KEYS = ['HL_P95_shift', 'HL_P50_shift', 'HL_P05_shift', 'HL_sticking_shift', 'HL_trending', 'HL_high_OOC', 'HL_record_high_low', 'HL_category_LT_shift']


class SPCApp(QtWidgets.QMainWindow): # 將 QTabWidget 改為 QMainWindow
    def __init__(self):
        super().__init__()

        self.filepath = resource_path('input/All_Chart_Information.xlsx')
        self.raw_data_directory = resource_path('input/raw_charts/')
        self.image_path = resource_path('image.png')
        self.results = []

        # 性能優化：添加快取
        self.csv_cache = {}  # CSV 文件快取
        self.chart_types_cache = {}  # 數據類型快取
        
        self.filter_type_combo = None
        self.filter_value_combo = None
        self.header_container = None
        # 新增用於兩個圓餅圖和一個長條圖的 Canvas 屬性
        self.status_pie_canvas = None
        self.processed_violation_pie_canvas = None
        self.anomaly_bar_canvas = None
        # Summary Tab 中圖表相關的屬性
        self.charts_main_layout = None
        self.charts_horizontal_layout = None

        # Summary Tab 中 TableWidget 屬性
        self.violation_table_label = None
        self.violation_table = None

        # QMainWindow 需要一個中央小部件
        self.central_widget = QtWidgets.QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_horizontal_layout = QtWidgets.QHBoxLayout(self.central_widget) # 主要的水平佈局

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("SPC Chart Processor")
        screen = QtWidgets.QApplication.primaryScreen()
        size = screen.size()
        w = int(size.width() * 1)
        h = int(size.height() * 0.9)
        self.setGeometry(0, 0, w, h)
        self.setStyleSheet("""
            * {
                color: #000957;
                font-weight: bold;
            }
            /* TabWidget related styles (for oob_system_tabs) */
            QTabWidget::pane {
                border: 1px solid #c4c4c3;
                top: -1px;
                background: #f4f6f9;
            }

            QTabWidget::tab-bar {
                left: 5px;
            }

            QTabBar::tab {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #E1E1E1, stop: 0.4 #DDDDDD,
                                            stop: 0.5 #D8D8D8, stop: 1.0 #D3D3D3);
                border: 1px solid #c4ccff;
                border-bottom-color: #c2c7cb;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                min-width: 8ex;
                padding: 8px;
                font-weight: bold;
                color: #000957;
            }

            QTabBar::tab:selected, QTabBar::tab:hover {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #fafafa, stop: 0.4 #f4f4f4,
                                            stop: 0.5 #e7e7e7, stop: 1.0 #fafafa);
            }

            QTabBar::tab:selected {
                border-color: #c2c7cb;
                border-bottom-color: #f4f6f9;
            }

            QTabBar::tab:!selected {
                margin-top: 2px;
            }

            QWidget {
                font-family: 'Segoe UI';
                background-color: #f4f6f9;
                color: #000957;
            }
            QLabel {
                font-size: 14px;
                color: #000957;
            }
            QProgressBar {
                border-radius: 12px;
                height: 25px;
                background: #e0e0e0;
                text-align: center;
                color: #000957;
                margin: 5px;
            }
            QProgressBar::chunk {
                background: linear-gradient(45deg, #344CB7, #577BC1);
                border-radius: 12px;
            }
            QPushButton {
                background-color: #344CB7;
                color: white;
                border-radius: 12px;
                padding: 12px 25px;
                font-size: 16px;
                border: none;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #577BC1;
            }
            QComboBox {
                background-color: white;
                border: 1px solid #344CB7;
                border-radius: 8px;
                padding: 8px;
                color: #000957;
                font-weight: bold;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                selection-background-color: #FFEB00;
                selection-color: #000957;
            }
            QScrollArea {
                border: none;
            }
            QVBoxLayout {
                spacing: 10px;
            }
            /* Styles for the left menu buttons */
            QPushButton.menu_button { /* Added a class for menu buttons */
                background-color: #344CB7;
                color: white;
                border-radius: 8px; /* Slightly smaller radius for menu buttons */
                padding: 10px 15px;
                font-size: 14px;
                text-align: left; /* Align text to left */
                border: none;
                font-weight: bold;
            }
            QPushButton.menu_button:hover {
                background-color: #577BC1;
            }
            QPushButton.menu_button:checked { /* Style for selected button */
                background-color: #000957; /* Darker blue when selected */
                border-left: 5px solid #FFEB00; /* Yellow accent on left */
                padding-left: 10px; /* Adjust padding due to border */
            }
        """)

        # --- 左側選單區域 ---
        self.left_menu_widget = QtWidgets.QWidget()
        self.left_menu_widget.setFixedWidth(180) # 設定選單寬度
        self.left_menu_widget.setStyleSheet("background-color: #344CB7;") # 選單背景色
        self.left_menu_layout = QtWidgets.QVBoxLayout(self.left_menu_widget)
        self.left_menu_layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignTop) # 按鈕靠頂部對齊
        self.left_menu_layout.setContentsMargins(10, 20, 10, 10)
        self.left_menu_layout.setSpacing(15) # 選單項目間距

        # 選單按鈕
        self.home_button = self._create_menu_button("Home")
        self.split_data_button = self._create_menu_button("Split Data")
        self.oob_system_button = self._create_menu_button("OOB && SPC System")
        self.cpk_calculation_button = self._create_menu_button("Cpk Calculator")
        # --- 新增 Tool Matching 按鈕 ---
        self.tool_matching_button = self._create_menu_button("Tool Matching")

        self.left_menu_layout.addWidget(self.home_button)
        self.left_menu_layout.addWidget(self.split_data_button)
        self.left_menu_layout.addWidget(self.oob_system_button)
        self.left_menu_layout.addWidget(self.cpk_calculation_button)
        self.left_menu_layout.addWidget(self.tool_matching_button)
        self.left_menu_layout.addStretch()

        # 將左側選單添加到主水平佈局
        self.main_horizontal_layout.addWidget(self.left_menu_widget)

        # --- 右側內容區域 (QStackedWidget) ---
        self.content_stacked_widget = QtWidgets.QStackedWidget()
        self.main_horizontal_layout.addWidget(self.content_stacked_widget)

        # 1. 首頁內容
        self.home_page = self._create_home_page()
        self.content_stacked_widget.addWidget(self.home_page)

        # 2. 拆分資料頁面內容
        self.split_data_page = self._create_split_data_page()
        self.content_stacked_widget.addWidget(self.split_data_page)

        # 3. OOB System 頁面內容 (包含 Chart Processing 和 Summary Dashboard)
        self.oob_system_tabs = QtWidgets.QTabWidget()

        # 4. Tool Matching 頁面內容
        self.tool_matching_page = self._create_tool_matching_page()
        self.content_stacked_widget.addWidget(self.tool_matching_page)
        self.oob_system_tabs.setObjectName("OOBSystemTabs") # 給予一個名稱以便於樣式控制
        
        # --- 建立第一個分頁 (圖表處理與顯示) ---
        self.processing_tab_widget = QtWidgets.QWidget()
        processing_layout = QtWidgets.QVBoxLayout(self.processing_tab_widget)

        self.display_gui_checkbox = QtWidgets.QCheckBox("Show Charts in GUI")
        self.display_gui_checkbox.setChecked(True)
        processing_layout.addWidget(self.display_gui_checkbox)

        # Assuming create_start_button and create_progress_bar are defined elsewhere
        # Placeholder for start_button
        self.start_button = QtWidgets.QPushButton("Start Process")
        processing_layout.addWidget(self.start_button)
        # 連接 Start 按鈕到 process_charts
        self.start_button.clicked.connect(self.process_charts)

        # Placeholder for progress_bar
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setValue(0)
        processing_layout.addWidget(self.progress_bar)

        self.image_container = QtWidgets.QScrollArea(self)
        self.image_container.setWidgetResizable(True)
        processing_layout.addWidget(self.image_container)

        self.image_grid_widget = QtWidgets.QWidget()
        self.image_grid_layout = QtWidgets.QGridLayout(self.image_grid_widget)
        self.image_grid_layout.setSpacing(20)
        self.image_container.setWidget(self.image_grid_widget)

        processing_layout.setContentsMargins(10, 10, 10, 10)
        processing_layout.setSpacing(10)
        
        # 將 Chart Processing Tab 添加到 oob_system_tabs
        self.oob_system_tabs.addTab(self.processing_tab_widget, "Chart Processing")


        # --- 建立第二個分頁 (Summary Dashboard) ---
        # 正確建立 summary dashboard 與其屬性
        self.setup_summary_dashboard_tab()
        self.oob_system_tabs.addTab(self.summary_tab_widget, "Summary Dashboard")

        # 將 OOB System 整個 QTabWidget 添加到 content_stacked_widget
        self.content_stacked_widget.addWidget(self.oob_system_tabs)

        # --- 新增 Cpk Calculator 頁面內容到 stacked widget ---
        self.cpk_calculation_page = self._create_cpk_calculation_page()
        self.content_stacked_widget.addWidget(self.cpk_calculation_page)


        # 連接選單按鈕到 QStackedWidget 的頁面切換（只切換頁面，不手動 setChecked）
        self.home_button.clicked.connect(lambda: self.content_stacked_widget.setCurrentWidget(self.home_page))
        self.split_data_button.clicked.connect(lambda: self.content_stacked_widget.setCurrentWidget(self.split_data_page))
        self.oob_system_button.clicked.connect(lambda: self.content_stacked_widget.setCurrentWidget(self.oob_system_tabs))
        self.cpk_calculation_button.clicked.connect(lambda: self.content_stacked_widget.setCurrentWidget(self.cpk_calculation_page))
        self.tool_matching_button.clicked.connect(lambda: self.content_stacked_widget.setCurrentWidget(self.tool_matching_page))

        # === QButtonGroup 互斥設定與預設選中 ===
        self.menu_button_group = QtWidgets.QButtonGroup(self)
        self.menu_button_group.addButton(self.home_button)
        self.menu_button_group.addButton(self.split_data_button)
        self.menu_button_group.addButton(self.oob_system_button)
        self.menu_button_group.addButton(self.cpk_calculation_button)
        self.menu_button_group.addButton(self.tool_matching_button)
        self.menu_button_group.setExclusive(True)
        self.home_button.setChecked(True)

    def _create_tool_matching_page(self):
        """
        建立 Tool Matching 頁面 (Widget)。
        """
        from tool_matching_widget_osat import ToolMatchingWidget
        widget = ToolMatchingWidget(self)
        return widget

    # --- 新增輔助方法用於建立選單按鈕 ---
    def _create_menu_button(self, text):
        button = QtWidgets.QPushButton(text)
        button.setCheckable(True) # Make button checkable for selection feedback
        button.setFont(QtGui.QFont("Segoe UI", 14, QtGui.QFont.Weight.Bold))
        button.setStyleSheet("QPushButton.menu_button { background-color: #344CB7; color: white; border-radius: 8px; padding: 10px 15px; font-size: 14px; text-align: left; border: none; font-weight: bold; }"
                             "QPushButton.menu_button:hover { background-color: #577BC1; }"
                             "QPushButton.menu_button:checked { background-color: #000957; border-left: 5px solid #FFEB00; padding-left: 10px; }")
        # Apply the class for styling
        button.setProperty("class", "menu_button")
        return button

    # --- 新增首頁和拆分資料頁面的佔位符方法 ---
    def _create_home_page(self):
            widget = QtWidgets.QWidget()
            layout = QtWidgets.QVBoxLayout(widget)
            # Update the label content for "Welcome to Supplier SPC!"
            label = QtWidgets.QLabel("<h1>Welcome to Supplier SPC!</h1><p>Please select an option from the left menu.</p>")
            label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(label)
            return widget
    def _create_split_data_page(self):
        """
        這個方法現在會創建並返回你的 SplitDataWidget 實例。
        """
        widget = SplitDataWidget(self) # 創建 SplitDataWidget 的實例，並將 MainWindow 作為其父物件
        return widget
    
    def _create_cpk_calculation_page(self):
        from spc_cpk_dashboard_osat import SPCCpkDashboard
        widget = SPCCpkDashboard(self)
        return widget
    
    def setup_summary_dashboard_tab(self):
        self.summary_tab_widget = QtWidgets.QWidget()
        self.summary_tab_widget.setObjectName("SummaryTabWidget")
        summary_layout = QtWidgets.QVBoxLayout(self.summary_tab_widget)

        # === Summary Dashboard UI 元素 ===
        self.summary_title_label = QtWidgets.QLabel("<b>Summary Dashboard</b>")
        self.summary_title_label.setFont(QtGui.QFont("Segoe UI", 16, QtGui.QFont.Weight.Bold))
        self.summary_title_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        summary_layout.addWidget(self.summary_title_label)

        # 統計數字 Label 的網格佈局
        summary_stats_layout = QtWidgets.QGridLayout()
        summary_stats_layout.setObjectName("SummaryStatsGrid")

        # Label 定義保持不變
        self.total_charts_label_summary = QtWidgets.QLabel("Total Charts: N/A")
        self.processed_charts_label_summary = QtWidgets.QLabel("Processed Successfully: N/A")
        self.skipped_charts_label_summary = QtWidgets.QLabel("No Data: N/A")
        self.ooc_charts_label_summary = QtWidgets.QLabel("Charts with OOC: N/A")
        self.we_count_charts_label_summary = QtWidgets.QLabel("Charts with WE Rule: N/A")
        self.oob_charts_label_summary = QtWidgets.QLabel("Charts with OOB: N/A")

        summary_stats_layout.addWidget(self.total_charts_label_summary, 0, 0)
        summary_stats_layout.addWidget(self.processed_charts_label_summary, 0, 1)
        summary_stats_layout.addWidget(self.skipped_charts_label_summary, 0, 2)

        summary_stats_layout.addWidget(self.ooc_charts_label_summary, 1, 0)
        summary_stats_layout.addWidget(self.we_count_charts_label_summary, 1, 1)
        summary_stats_layout.addWidget(self.oob_charts_label_summary, 1, 2)

        summary_stats_layout.setSpacing(15)

        summary_layout.addLayout(summary_stats_layout)

        # --- 圖表顯示區域 ---
        self.charts_main_layout = QtWidgets.QVBoxLayout()
        self.charts_horizontal_layout = QtWidgets.QHBoxLayout()
        self.charts_main_layout.addLayout(self.charts_horizontal_layout)
        summary_layout.addLayout(self.charts_main_layout)

        # --- 違規圖表詳細列表區域 (保持不變) ---
        self.violation_table_label = QtWidgets.QLabel("<b>Charts with Anomalies Details</b>")
        self.violation_table_label.setFont(QtGui.QFont("Segoe UI", 12, QtGui.QFont.Weight.Bold))
        self.violation_table_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
        summary_layout.addWidget(self.violation_table_label)

        self.violation_table = QtWidgets.QTableWidget()
        self.violation_table.setColumnCount(5)
        headers = ["Group Name", "Chart Name", "OOC Count", "WE Rules", "OOB Rules"]
        self.violation_table.setHorizontalHeaderLabels(headers)
        self.violation_table.horizontalHeader().setStretchLastSection(True)
        self.violation_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.violation_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.violation_table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.violation_table.verticalHeader().setVisible(False)
        self.violation_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.violation_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.violation_table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        summary_layout.addWidget(self.violation_table)

        summary_layout.addStretch()

        summary_layout.setContentsMargins(20, 20, 20, 20)
        summary_layout.setSpacing(20)
        
        # 注意: setup_summary_dashboard_tab 不再呼叫 self.addTab，
        # 因為它會被 oob_system_tabs 呼叫
    # --- 繪製圖表的輔助方法 ---

    def create_status_pie_chart(self, processed, skipped):
        fig = Figure(figsize=(4, 4))
        ax = fig.add_subplot(111)

        labels = ['Processed', 'No Data']
        sizes = [processed, skipped]
        colors = ['#577BC1', '#cccccc'] # Blue and Grey

        # 甜甜圈圖設定
        wedgeprops = {'width': 0.3, 'edgecolor': 'white'} # 設定甜甜圈厚度

        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
               shadow=False, startangle=140, wedgeprops=wedgeprops, 
               pctdistance=0.85, labeldistance=1.1,
               textprops={'fontsize': 10})
        ax.axis('equal')
        ax.set_title('Overall Processing Status', fontsize=12, pad=15)

        # 確保圖表邊界有足夠空間
        fig.subplots_adjust(left=0.1, right=0.9, top=0.85, bottom=0.15)

        fig.patch.set_alpha(0)

        canvas = FigureCanvas(fig)
        return canvas
    def create_processed_violation_pie_chart(self, processed_count, violating_count):
        fig = Figure(figsize=(4, 4))
        ax = fig.add_subplot(111)

        # 計算未違規的已處理圖表數量
        non_violating_count = processed_count - violating_count

        labels = ['Violating', 'Normal']
        sizes = [violating_count, non_violating_count]
        colors = ['#ff6666', '#99ff99'] # Red and Green

        # 如果沒有成功處理的圖表，或者所有都已處理但都未違規
        if processed_count == 0 or (processed_count > 0 and violating_count == 0):
             labels = ['All Normal']
             sizes = [processed_count if processed_count > 0 else 1] # 如果 processed_count=0，給個非零大小繪圖
             colors = ['#99ff99']
             if processed_count == 0: # 如果 processed_count=0，顯示 N/A 或無數據
                  labels = ['N/A']
                  sizes = [1]
                  colors = ['#cccccc']


        # 甜甜圈圖設定
        wedgeprops = {'width': 0.3, 'edgecolor': 'white'}

        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
               shadow=False, startangle=140, wedgeprops=wedgeprops, 
               pctdistance=0.85, labeldistance=1.1,
               textprops={'fontsize': 10})
        ax.axis('equal')
        ax.set_title('Violation Rate (Processed Charts)', fontsize=12, pad=15) # <--- 增加標題

        # 確保圖表邊界有足夠空間
        fig.subplots_adjust(left=0.1, right=0.9, top=0.85, bottom=0.15)

        fig.patch.set_alpha(0)

        canvas = FigureCanvas(fig)
        return canvas
    def create_anomaly_bar_chart(self, ooc_count, we_count, oob_count):
        fig = Figure(figsize=(5, 4))
        ax = fig.add_subplot(111)

        categories = ['OOC', 'WE_Rule', 'OOB']
        counts = [ooc_count, we_count, oob_count]
        colors = ['#ff9999','#66b3ff','#99ff99'] # Red, Blue, Green

        bars = ax.bar(categories, counts, color=colors)

        for bar in bars:
            yval = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2.0, yval, int(yval), va='bottom', ha='center', fontsize=8)

        ax.set_ylabel('Number of Charts', fontsize=10)
        ax.set_title('Charts with Anomalies', fontsize=12)
        ax.set_ylim(0, max(counts) * 1.2 or 1)

        fig.tight_layout()
        fig.patch.set_alpha(0)

        canvas = FigureCanvas(fig)
        return canvas
    # --- 清理 Summary Tab 圖表和表格的方法 ---
    def clear_summary_charts(self):
        print("Clearing summary charts and table...")
        # 清理 charts_horizontal_layout 中的所有項目
        if self.charts_horizontal_layout is not None:
             while self.charts_horizontal_layout.count():
                 item = self.charts_horizontal_layout.takeAt(0)
                 if item.widget():
                     item.widget().deleteLater()
                 elif item.spacerItem():
                      self.charts_horizontal_layout.removeItem(item.spacerItem())

        # 將 Canvas 屬性設為 None，雖然 deleteLater() 已經標記刪除，但好習慣是解除引用
        self.status_pie_canvas = None
        self.processed_violation_pie_canvas = None # <--- 清理第二個圓餅圖 Canvas
        self.anomaly_bar_canvas = None


        # 清理表格內容
        if self.violation_table:
             self.violation_table.setRowCount(0) # 將行數設為 0 清空表格
        print("Summary charts and table cleared.")

    def update_summary_dashboard(self, total, processed, skipped):
        print("\nUpdating Summary Dashboard...")
        ooc_count = 0
        we_count = 0
        oob_count = 0

        violating_charts = []

        for result in self.results: # self.results 已經是成功處理的圖表結果列表
            has_ooc = result.get('ooc_cnt', 0) > 0
            has_we = result.get('WE_Rule', '') and result.get('WE_Rule', '') != 'N/A'
            has_oob = result.get('OOB_Rule', '') and result.get('OOB_Rule', '') != 'N/A'

            if has_ooc:
                 ooc_count += 1
            if has_we:
                 we_count += 1
            if has_oob:
                 oob_count += 1

            if has_ooc or has_we or has_oob:
                violating_charts.append(result)


        print(f"DEBUG: Calculated OOC chart count: {ooc_count}")
        print(f"DEBUG: Calculated WE_Rule chart count: {we_count}")
        print(f"DEBUG: Calculated OOB chart count: {oob_count}")
        print(f"DEBUG: Number of violating charts: {len(violating_charts)}")


        # 更新統計數字 Label 文本
        self.total_charts_label_summary.setText(f"Total Charts: {total}")
        self.processed_charts_label_summary.setText(f"Processed Successfully: {processed}")
        self.skipped_charts_label_summary.setText(f"No Data: {skipped}")
        self.ooc_charts_label_summary.setText(f"Charts with OOC: {ooc_count}")
        self.we_count_charts_label_summary.setText(f"Charts with WE_Rule: {we_count}")
        self.oob_charts_label_summary.setText(f"Charts with OOB: {oob_count}")


        # --- 清理舊圖表並添加新圖表 ---
        self.clear_summary_charts() # 清理 Summary Tab 中的舊圖表和表格

        if total > 0:
             self.status_pie_canvas = self.create_status_pie_chart(processed, skipped)
             self.charts_horizontal_layout.addStretch() # 左邊添加彈性空間
             self.charts_horizontal_layout.addWidget(self.status_pie_canvas)


        # 添加成功處理圖表違規比例甜甜圈圖 (中間圖)
        # 這個圖只需要在有成功處理的圖表時顯示
        if processed > 0:
             # violating_charts 列表已經是從 self.results (已處理圖表) 中篩選的
             violating_count_in_processed = len(violating_charts)
             self.processed_violation_pie_canvas = self.create_processed_violation_pie_chart(processed, violating_count_in_processed)
             self.charts_horizontal_layout.addStretch() # 圓餅圖1和圓餅圖2之間添加彈性空間
             self.charts_horizontal_layout.addWidget(self.processed_violation_pie_canvas)

        if ooc_count > 0 or we_count > 0 or oob_count > 0:
             self.anomaly_bar_canvas = self.create_anomaly_bar_chart(ooc_count, we_count, oob_count)

             if self.charts_horizontal_layout.count() > 0 and not isinstance(self.charts_horizontal_layout.itemAt(self.charts_horizontal_layout.count()-1).spacerItem(), type(None)):
                  # 如果最後一個 item 不是 stretch，則在其前面加 stretch
                  self.charts_horizontal_layout.addStretch()
             elif self.charts_horizontal_layout.count() == 0:
                 # 如果目前為空，先加 stretch
                 self.charts_horizontal_layout.addStretch()

             self.charts_horizontal_layout.addWidget(self.anomaly_bar_canvas)
             self.charts_horizontal_layout.addStretch() # 最右邊添加彈性空間

        self.violation_table.setRowCount(len(violating_charts))


        for row_index, result in enumerate(violating_charts):
            # 分別取得 group_name 和 chart_name
            group_name = result.get('group_name', 'N/A')
            chart_name = result.get('chart_name', 'N/A')

            ooc_cnt = result.get('ooc_cnt', 0)
            we_rules = result.get('WE_Rule', 'N/A')
            oob_rules = result.get('OOB_Rule', 'N/A')

            # 在表格中插入新的一行
            self.violation_table.insertRow(row_index)

            # 創建 QTableWidgetItem 並設定 alignment
            item_group_name = QtWidgets.QTableWidgetItem(group_name)
            item_chart_name = QtWidgets.QTableWidgetItem(chart_name)
            item_ooc_cnt = QtWidgets.QTableWidgetItem(str(ooc_cnt))
            item_we_rules = QtWidgets.QTableWidgetItem(we_rules)
            item_oob_rules = QtWidgets.QTableWidgetItem(oob_rules)

            # 將所有欄位內容都置中顯示
            for item in [item_group_name, item_chart_name, item_ooc_cnt, item_we_rules, item_oob_rules]:
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

            # 將項目設定到正確的欄位
            self.violation_table.setItem(row_index, 0, item_group_name)
            self.violation_table.setItem(row_index, 1, item_chart_name)
            self.violation_table.setItem(row_index, 2, item_ooc_cnt)
            self.violation_table.setItem(row_index, 3, item_we_rules)
            self.violation_table.setItem(row_index, 4, item_oob_rules)

        print("Summary Dashboard updated.")

    # --- UI部件 (需要確保這些方法在類別定義內) ---
    def create_start_button(self):
        button = QtWidgets.QPushButton("Start Processing", self)
        button.setFont(QtGui.QFont("Segoe UI", 14))
        button.clicked.connect(self.process_charts)
        return button

    def create_progress_bar(self):
        progress_bar = QtWidgets.QProgressBar(self)
        progress_bar.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        return progress_bar

    def create_image_label(self, image_path: str, max_width=450, max_height=350, keep_original_size=False):
        try:
            image = Image.open(image_path)
            qt_image = ImageQt(image)
            pixmap = QtGui.QPixmap.fromImage(qt_image)
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            label = QtWidgets.QLabel("Image Not Found", self)
            label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            label.setStyleSheet("color: red;")
            return label

        label = QtWidgets.QLabel(self)
        label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

        if keep_original_size:
            label.setPixmap(pixmap)
            label.setMaximumSize(pixmap.width(), pixmap.height())
            label.setMinimumSize(pixmap.width(), pixmap.height())
        else:
            scaled_pixmap = pixmap.scaled(
                max_width,
                max_height,
                QtCore.Qt.AspectRatioMode.KeepAspectRatio,
                QtCore.Qt.TransformationMode.SmoothTransformation
            )
            label.setPixmap(scaled_pixmap)
            label.setMaximumSize(max_width, max_height)

        return label

    def get_cached_csv(self, filepath):
        """使用快取讀取 CSV 檔案，提升性能"""
        try:
            if filepath not in self.csv_cache:
                print(f"  - 讀取並快取 CSV: {os.path.basename(filepath)}")
                df = pd.read_csv(filepath)
                self.csv_cache[filepath] = df
            else:
                print(f"  - 使用快取的 CSV: {os.path.basename(filepath)}")
            
            # 返回副本避免修改快取的資料
            return self.csv_cache[filepath].copy()
        except Exception as e:
            print(f"[Error] 讀取檔案 {filepath} 失敗: {str(e)}")
            return None

    def process_charts(self):
        import time
        self.results = []
        total_charts_count = 0
        skipped_charts_count = 0
        processed_charts_count = 0

        try:
            self.validate_files_and_directories()
            self.progress_bar.setValue(0)
            self.clear_image_grid()

            all_charts_info = load_chart_information(self.filepath)
            total_charts_count = len(all_charts_info)
            self.progress_bar.setMaximum(100)

            # 性能優化：預處理所有圖表的數據類型
            print("=== 性能優化：開始預處理數據類型 ===")
            self.preprocess_chart_types(all_charts_info)
            
            # 清空 CSV 快取（如果之前有的話）
            self.csv_cache.clear()
            print("=== 預處理完成，開始處理圖表 ===")

            if self.display_gui_checkbox.isChecked():
                self.add_column_headers()

            execution_time = load_execution_time(self.filepath)

            for i, (_, chart_info) in enumerate(all_charts_info.iterrows()):
                group_name = str(chart_info['GroupName'])
                chart_name = str(chart_info['ChartName'])
                chart_key = f"{group_name}_{chart_name}"
                print(f"\n正在處理圖表: GroupName={group_name}, ChartName={chart_name}")

                try:
                    filepath = find_matching_file(self.raw_data_directory, group_name, chart_name)
                    
                    if filepath and os.path.exists(filepath):
                        # 性能優化：使用快取讀取 CSV
                        raw_df = self.get_cached_csv(filepath)
                        
                        if raw_df is not None:
                            print(f" - 原始資料 shape: {raw_df.shape}")

                            # 性能優化：使用預處理的數據類型
                            data_type = self.chart_types_cache.get(chart_key, 'continuous')
                            chart_info = chart_info.copy()  # 避免修改原始數據
                            chart_info['data_type'] = data_type
                            print(f" - 使用快取的數據類型: {data_type}")

                            if 'point_time' in raw_df.columns:
                                raw_df['point_time'] = pd.to_datetime(raw_df['point_time'], errors='coerce')
                                raw_df.dropna(subset=['point_time'], inplace=True)

                            is_successful, processed_df, updated_chart_info, _full_df = preprocess_data(chart_info, raw_df)

                            if not is_successful or processed_df is None or processed_df.empty:
                                print(f"[Info] 圖表 {group_name}/{chart_name} 預處理失敗或資料為空，跳過。")
                                skipped_charts_count += 1
                            else:
                                print(f" - 預處理後資料 shape: {processed_df.shape}")
                                print(f" - 準備分析圖表: {group_name}/{chart_name}")

                                # 性能優化：減少假進度條的步數，降低 GUI 更新頻率
                                fake_steps = 5  # 從 10 減少到 5
                                for fake_step in range(fake_steps):
                                    percent = int(((i + fake_step / fake_steps) / total_charts_count) * 100)
                                    self.progress_bar.setValue(percent)
                                    if fake_step % 2 == 0:  # 只在偶數步驟更新 GUI
                                        QtWidgets.QApplication.processEvents()
                                    time.sleep(0.005)  # 從 0.01 減少到 0.005

                                result = self.analyze_chart(execution_time, processed_df, updated_chart_info)

                                if result:
                                    self.results.append(result)
                                    processed_charts_count += 1

                                    if self.display_gui_checkbox.isChecked():
                                        if 'chart_path' in result and 'weekly_chart_path' in result:
                                            self.display_image(result, len(self.results) - 1)
                                            print(f" - 顯示圖表完成: {group_name}/{chart_name}")
                                        else:
                                            print(f"[Warning] 圖表 {group_name}/{chart_name} 缺少圖片路徑，無法顯示。")
                                    else:
                                        print(f" - GUI 顯示已禁用，跳過顯示圖表: {group_name}/{chart_name}")
                                else:
                                    print(f"[Info] 圖表 {group_name}/{chart_name} 分析返回 None，跳過結果記錄。")
                                    skipped_charts_count += 1
                        else:
                            print(f"[Error] 無法讀取檔案: {filepath}")
                            skipped_charts_count += 1
                    else:
                        print(f"[Info] 圖表 {group_name}/{chart_name} 對應檔案 {filepath} 不存在，跳過處理。")
                        skipped_charts_count += 1

                except FileNotFoundError:
                    print(f"[Warning] 檔案未找到，跳過圖表: {group_name}/{chart_name}")
                    skipped_charts_count += 1
                except Exception as e:
                    print(f"[Error] 處理圖表 {group_name}/{chart_name} 時發生錯誤: {str(e)}")
                    traceback.print_exc()
                    skipped_charts_count += 1

                # 性能優化：減少 GUI 更新頻率
                if i % 2 == 0:  # 每3個圖表更新一次進度條
                    percent = int(((i + 1) / total_charts_count) * 100)
                    self.progress_bar.setValue(percent)
                    QtWidgets.QApplication.processEvents()

            # 最終更新進度條
            self.progress_bar.setValue(100)
            QtWidgets.QApplication.processEvents()

            self.update_summary_dashboard(total_charts_count, processed_charts_count, skipped_charts_count)

            if self.results:
                self.save_results()
                QtWidgets.QMessageBox.information(self, "Processing Complete", "Results have been saved to result_with_images.xlsx")
            else:
                QtWidgets.QMessageBox.information(self, "Processing Complete", "No charts were processed successfully to save.")

            # 清理快取（可選）
            print(f"處理完成，清理快取。CSV 快取大小: {len(self.csv_cache)}")
            # 如果記憶體有限，可以清空快取
            # self.csv_cache.clear()

        except FileNotFoundError as e:
            self.show_error("File Error", str(e))
        except NotADirectoryError as e:
            self.show_error("Directory Error", str(e))
        except Exception as e:
            self.show_error("Processing Error", str(e))
            traceback.print_exc()

    # --- 新增清理 Grid Layout 的方法 (針對第一個分頁) ---
    def clear_image_grid(self):
        print("Clearing image grid...")
        if self.header_container and self.processing_tab_widget.layout() and self.processing_tab_widget.layout().indexOf(self.header_container) != -1:
             while self.header_container.count():
                 item = self.header_container.takeAt(0)
                 if item.widget():
                     item.widget().deleteLater()
             if self.processing_tab_widget.layout():
                  self.processing_tab_widget.layout().removeItem(self.header_container)
             self.header_container = None

        if self.image_grid_layout:
            while self.image_grid_layout.count():
                item = self.image_grid_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
                elif item.layout():
                     self.clear_layout(item.layout())
        print("Image grid cleared.")

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
                elif item.layout():
                    self.clear_layout(item.layout())

    def validate_files_and_directories(self):
        if not os.path.isdir(self.raw_data_directory):
            print(f"Creating directory: {self.raw_data_directory}")
            os.makedirs(self.raw_data_directory, exist_ok=True)

        if not os.path.exists(self.filepath):
            print(f"[Error] 缺少必要的 All_Chart_Information.xlsx 檔案於 {self.filepath}，請先準備好檔案再執行。")
            raise FileNotFoundError(f"{self.filepath} does not exist. Please provide the required Excel file.")

        print("Files and directories validated.")

    def add_column_headers(self):
        if self.header_container and self.processing_tab_widget.layout() and self.processing_tab_widget.layout().indexOf(self.header_container) != -1:
            return

        self.header_container = QtWidgets.QHBoxLayout()
        for header in HEADERS:
            label = QtWidgets.QLabel(f"<b>{header}</b>", self)
            label.setFont(QtGui.QFont("Segoe UI", 12))
            label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.header_container.addWidget(label)

        if self.processing_tab_widget.layout():
             self.processing_tab_widget.layout().insertLayout(3, self.header_container)
        else:
             print("[Warning] Processing tab layout is not initialized when trying to add headers.")

    def update_filter_values(self, filter_type):
        self.filter_value_combo.clear()
        if filter_type == "FabPhase":
            self.filter_value_combo.addItems(["T14P5", "T14P6", "T12P4"])
        elif filter_type == "Production Line":
            self.filter_value_combo.addItems(["7", "11", "5", "8", "6", "9", "4", "10"])

    def preprocess_chart_types(self, all_charts_info):
        """
        性能優化：一次性預處理所有圖表的數據類型，避免重複計算
        """
        print("正在預處理圖表數據類型...")
        chart_types = {}
        processed_files = set()
        
        for _, chart_info in all_charts_info.iterrows():
            group_name = str(chart_info.get('GroupName', 'Unknown'))
            chart_name = str(chart_info.get('ChartName', 'Unknown'))
            chart_key = f"{group_name}_{chart_name}"
            
            # 找到對應的 CSV 文件
            filepath = find_matching_file(self.raw_data_directory, group_name, chart_name)
            
            if filepath and os.path.exists(filepath) and filepath not in processed_files:
                try:
                    # 快速讀取部分數據來判斷類型（只需要 point_val 欄位）
                    raw_df = pd.read_csv(filepath, usecols=['point_val'] if 'point_val' in pd.read_csv(filepath, nrows=1).columns else None, nrows=1000)
                    
                    if 'point_val' in raw_df.columns:
                        data_type = determine_data_type(raw_df['point_val'].dropna())
                        chart_types[chart_key] = data_type
                        processed_files.add(filepath)
                        print(f"  預處理完成: {chart_key} -> {data_type}")
                    else:
                        chart_types[chart_key] = 'continuous'  # 預設值
                        
                except Exception as e:
                    print(f"  預處理錯誤 {chart_key}: {e}")
                    chart_types[chart_key] = 'continuous'  # 預設值
            else:
                chart_types[chart_key] = 'continuous'  # 預設值
        
        self.chart_types_cache = chart_types
        print(f"數據類型預處理完成，共處理 {len(chart_types)} 個圖表")
        return chart_types

    def get_cached_csv(self, filepath):
        """
        性能優化：使用快取讀取 CSV 文件，避免重複讀取
        """
        if filepath not in self.csv_cache:
            try:
                self.csv_cache[filepath] = pd.read_csv(filepath)
                print(f"  CSV 文件已快取: {os.path.basename(filepath)}")
            except Exception as e:
                print(f"  CSV 讀取錯誤 {filepath}: {e}")
                return None
        
        return self.csv_cache[filepath].copy() if self.csv_cache[filepath] is not None else None


    def analyze_chart(self, execution_time, raw_df, chart_info, task_output_dir: str = 'output'):
        # 補齊 rule_list，確保每個 chart 都有正確的 WE 規則清單
        if 'rule_list' not in chart_info or not chart_info['rule_list']:
            rule_list = []
            for rule in ['WE1','WE2','WE3','WE4','WE5','WE6','WE7','WE8','WE9','WE10']:
                if chart_info.get(rule, 'N') == 'Y':
                    rule_list.append(rule)
        chart_info['rule_list'] = rule_list
        group_name = str(chart_info.get('group_name', chart_info.get('GroupName', 'Unknown')))
        chart_name = str(chart_info.get('chart_name', chart_info.get('ChartName', 'Unknown')))
        print(f" - analyze_chart 開始處理 {group_name}/{chart_name}")
        print(f" - analyze_chart: 接收到的 raw_df shape: {raw_df.shape}")

        if 'point_time' not in raw_df.columns or not pd.api.types.is_datetime64_any_dtype(raw_df['point_time']):
                print(f" - analyze_chart: 'point_time' column missing or not datetime type for {group_name}/{chart_name}. Skipping analysis.")
                return None

        latest_raw_data_time = raw_df['point_time'].max()

        if execution_time is None or pd.isna(execution_time):
            print(" - analyze_chart: execution_time is None or NaT, using latest data time as weekly end date.")
            weekly_end_date = latest_raw_data_time
        else:
            print(f" - analyze_chart: execution_time is provided ({execution_time}), using it as weekly end date.")
            weekly_end_date = execution_time

        if pd.isna(weekly_end_date):
            print(f" - analyze_chart: Unable to determine weekly end date (latest_raw_data_time is also invalid). Skipping analysis.")
            return None

        weekly_start_date = weekly_end_date - pd.Timedelta(days=6)
        baseline_end_date = weekly_start_date - pd.Timedelta(seconds=1)
        # 這裡使用初始的一年基線範圍
        initial_baseline_start_date = baseline_end_date - pd.Timedelta(days=365)

        print(f" - analyze_chart: 計算出的時間範圍 - Weekly: {weekly_start_date} to {weekly_end_date}, Initial Baseline: {initial_baseline_start_date} to {baseline_end_date}")

        try:
            # === 提前進行數據類型判斷 ===
            if raw_df is None or raw_df.empty or 'point_val' not in raw_df.columns:
                print(" - analyze_chart: raw_df 無效或為空，預設為連續型")
                data_type = 'continuous'
            else:
                # 使用全部 point_val（移除 NaN）來判斷是否為離散
                data_type = determine_data_type(raw_df['point_val'].dropna())
                print(f" - analyze_chart: 數據類型判斷結果: {data_type}")
            
            chart_info['data_type'] = data_type

            # === 根據數據類型分流處理 ===
            if data_type == 'discrete':
                print(f" - analyze_chart: 執行離散型專用流程 for {group_name}/{chart_name}")
                result = self._process_discrete_chart(raw_df, chart_info, weekly_start_date, weekly_end_date, 
                                                    initial_baseline_start_date, baseline_end_date)
            else:
                print(f" - analyze_chart: 執行連續型流程 for {group_name}/{chart_name}")
                result = process_single_chart(chart_info.copy(), raw_df, initial_baseline_start_date, 
                                            baseline_end_date, weekly_start_date, weekly_end_date)
                if result:
                    result['data_type'] = 'continuous'

            if result is None:
                print(f" - analyze_chart: 處理返回 None for {group_name}/{chart_name}")
                return None

            # === 共同的後處理步驟 ===
            print(f" - analyze_chart: 準備生成圖表 for {group_name}/{chart_name}")
            
            # 生成 SPC 圖表
            image_path, violated_rules = plot_spc_chart(raw_df, chart_info, weekly_start_date, weekly_end_date, output_dir=task_output_dir)
            print(f" - analyze_chart: plot_spc_chart 完成，image_path: {image_path}")

            # 生成週圖表
            weekly_image_path = plot_weekly_spc_chart(raw_df, chart_info, weekly_start_date, weekly_end_date, output_dir=task_output_dir)
            print(f" - analyze_chart: plot_weekly_spc_chart 完成，weekly_image_path: {weekly_image_path}")

            # Cpk 計算
            weekly_data = raw_df[(raw_df['point_time'] >= weekly_start_date) & 
                               (raw_df['point_time'] <= weekly_end_date)].copy()
            cpk_result = calculate_cpk(weekly_data, chart_info)
            result['Cpk'] = cpk_result.get('Cpk', np.nan) if cpk_result else np.nan

            # 序列化 chart_data 供 UI Plotly hover 使用 (All Data SPC)
            try:
                _plot_cols = ['point_time', 'point_val']
                if 'Matching' in raw_df.columns:
                    _plot_cols.append('Matching')
                _plot_df = raw_df[_plot_cols].copy()
                _plot_df['point_time'] = _plot_df['point_time'].astype(str)
                _plot_df['point_val'] = pd.to_numeric(_plot_df['point_val'], errors='coerce')
                _plot_df = _plot_df.dropna(subset=['point_val'])
                if 'Matching' in _plot_df.columns:
                    _plot_df['Matching'] = _plot_df['Matching'].fillna('Unknown').astype(str)
                result['chart_data'] = _plot_df.to_dict(orient='records')
            except Exception as _chart_data_err:
                print(f" - analyze_chart: chart_data 序列化失敗 ({_chart_data_err})，略過")
                result['chart_data'] = []

            # 更新結果
            result['violated_rules'] = violated_rules if violated_rules is not None else {}
            self.build_result(result, image_path, weekly_image_path)

            print(f" - analyze_chart 處理完成並返回結果 for {group_name}/{chart_name}")
            return result

        except Exception as e:
                print(f"[Error] analyze_chart 處理圖表 {group_name}/{chart_name} 時發生錯誤: {str(e)}")
                traceback.print_exc()
                return None

    def _process_discrete_chart(self, raw_df, chart_info, weekly_start_date, weekly_end_date, 
                              initial_baseline_start_date, baseline_end_date):
        """
        離散型數據的專用處理流程，包含 record high low 判斷
        """
        group_name = chart_info.get('group_name', 'Unknown')
        chart_name = chart_info.get('chart_name', 'Unknown')
        
        print(f" - _process_discrete_chart: 開始離散型專用處理 {group_name}/{chart_name}")
        
        try:
            # === 基線範圍選擇邏輯 ===
            baseline_data_one_year = raw_df[(raw_df['point_time'] >= initial_baseline_start_date) & 
                                          (raw_df['point_time'] <= baseline_end_date)].copy()
            baseline_count_one_year = len(baseline_data_one_year)
            print(f" - _process_discrete_chart: 初始一年基線數據點數量: {baseline_count_one_year}")

            baseline_insufficient = False
            if baseline_count_one_year < 10:
                actual_baseline_start_date = baseline_end_date - pd.Timedelta(days=365 * 2)
                print(f" - _process_discrete_chart: 基線數據不足，擴展至兩年: {actual_baseline_start_date}")
                
                baseline_data_two_year = raw_df[(raw_df['point_time'] >= actual_baseline_start_date) & 
                                              (raw_df['point_time'] <= baseline_end_date)].copy()
                baseline_count_two_year = len(baseline_data_two_year)
                
                if baseline_count_two_year < 10:
                    print(f" - _process_discrete_chart: 擴展至兩年後仍少於10點，標記為基線不足")
                    baseline_insufficient = True
            else:
                actual_baseline_start_date = initial_baseline_start_date

            # 篩選最終數據
            baseline_data = raw_df[(raw_df['point_time'] >= actual_baseline_start_date) & 
                                 (raw_df['point_time'] <= baseline_end_date)].copy()
            weekly_data = raw_df[(raw_df['point_time'] >= weekly_start_date) & 
                               (raw_df['point_time'] <= weekly_end_date)].copy()

            baseline_empty = baseline_data.empty
            if baseline_empty:
                print(f" - _process_discrete_chart: 基線數據為空，但仍繼續處理 WE Rule 和圖表生成")
                baseline_insufficient = True
                
            if weekly_data.empty:
                print(f" - _process_discrete_chart: 週數據為空，跳過處理")
                return None

            # === 計算統計數據 ===
            def calculate_statistics(data):
                if data.shape[0] <= 1:
                    sigma = 0.0
                else:
                    sigma = data['point_val'].std()
                if np.isnan(sigma):
                    sigma = 0.0
                return {
                    'values': data['point_val'].values,
                    'cnt': data.shape[0],
                    'mean': data['point_val'].mean(),
                    'sigma': sigma
                }

            base_data_dict = calculate_statistics(baseline_data) if not baseline_empty else None
            weekly_data_dict = calculate_statistics(weekly_data)

            if not baseline_empty:
                print(f" - _process_discrete_chart: 基線統計 - cnt={base_data_dict['cnt']}, mean={base_data_dict['mean']}")
            else:
                print(f" - _process_discrete_chart: 基線數據為空，跳過基線統計輸出")
            print(f" - _process_discrete_chart: 週統計 - cnt={weekly_data_dict['cnt']}, mean={weekly_data_dict['mean']}")

            # === 初始化結果字典 ===
            result = {
                'data_cnt': weekly_data_dict['cnt'],
                'ooc_cnt': 0,
                'WE_Rule': '',
                'OOB_Rule': '',
                'Material_no': chart_info.get('material_no', 'N/A'),
                'group_name': chart_info.get('group_name', 'N/A'),
                'chart_name': chart_info.get('chart_name', 'N/A'),
                'chart_ID': chart_info.get('ChartID', 'N/A'),
                'Characteristics': chart_info.get('Characteristics', 'N/A'),
                'USL': chart_info.get('USL', 'N/A'),
                'LSL': chart_info.get('LSL', 'N/A'),
                'UCL': chart_info.get('UCL', 'N/A'),
                'LCL': chart_info.get('LCL', 'N/A'),
                'Target': chart_info.get('Target', 'N/A'),
                'Resolution': chart_info.get('Resolution', 'N/A'),
                'baseline_insufficient': baseline_insufficient,
                'baseline_empty': baseline_empty,  # 新增標記
                'data_type': 'discrete'
            }

            if not baseline_insufficient and not baseline_empty:
                # === OOC 計算 ===
                print(" - _process_discrete_chart: 計算 OOC...")
                weekly_df = pd.DataFrame({'point_val': weekly_data['point_val']})
                ooc_results = ooc_calculator(weekly_df, chart_info.get('UCL'), chart_info.get('LCL'))
                ooc_highlight = review_ooc_results(ooc_results[1], ooc_results[2])
                result['ooc_cnt'] = ooc_results[1]
                
                # === 離散型 OOB 計算 ===
                print(" - _process_discrete_chart: 計算離散型 OOB...")
                discrete_oob_result = discrete_oob_calculator(
                    base_data_dict, weekly_data_dict, chart_info,
                    raw_df, weekly_start_date, weekly_end_date,
                    actual_baseline_start_date, baseline_end_date
                )
                
                # === Record High Low 計算 ===
                print(" - _process_discrete_chart: 計算 record high low...")
                # DEBUG: 輸出時間範圍信息
                print(f" - DEBUG: 基線時間範圍 - 從 {actual_baseline_start_date} 到 {baseline_end_date}")
                print(f" - DEBUG: 當週時間範圍 - 從 {weekly_start_date} 到 {weekly_end_date}")
                print(f" - DEBUG: 基線結束與當週開始間隔 = {weekly_start_date - baseline_end_date}")
                # 基線先過濾 OOC 點再進行 record high/low 比較
                _ucl_d = chart_info.get('UCL')
                _lcl_d = chart_info.get('LCL')
                _baseline_for_record_d = baseline_data.copy()
                if _ucl_d is not None and not pd.isna(_ucl_d):
                    _baseline_for_record_d = _baseline_for_record_d[_baseline_for_record_d['point_val'] <= _ucl_d]
                if _lcl_d is not None and not pd.isna(_lcl_d):
                    _baseline_for_record_d = _baseline_for_record_d[_baseline_for_record_d['point_val'] >= _lcl_d]
                if _baseline_for_record_d.empty:
                    _baseline_for_record_d = baseline_data  # fallback：若全部被過濾則還原
                record_results = record_high_low_calculator(
                    weekly_data['point_val'].values,
                    _baseline_for_record_d['point_val'].values
                )
                
                # === 更新結果 ===
                result.update({
                    'HL_P95_shift': discrete_oob_result.get('HL_P95_shift', 'NO_HIGHLIGHT'),
                    'HL_P50_shift': discrete_oob_result.get('HL_P50_shift', 'NO_HIGHLIGHT'),
                    'HL_P05_shift': discrete_oob_result.get('HL_P05_shift', 'NO_HIGHLIGHT'),
                    'HL_sticking_shift': discrete_oob_result.get('HL_sticking_shift', 'NO_HIGHLIGHT'),
                    'HL_trending': discrete_oob_result.get('HL_trending', 'NO_HIGHLIGHT'),
                    'HL_high_OOC': ooc_highlight,
                    'HL_category_LT_shift': discrete_oob_result.get('HL_category_LT_shift', 'NO_HIGHLIGHT'),
                    'HL_record_high_low': record_results.get('highlight_status', 'NO_HIGHLIGHT'),
                    'record_high': record_results.get('record_high', False),
                    'record_low': record_results.get('record_low', False)
                })
                
                print(f" - _process_discrete_chart: 離散型 OOB 計算完成")
                
            else:
                # 基線不足時設置所有 OOB 為 NO_HIGHLIGHT
                result.update({
                    'HL_P95_shift': 'NO_HIGHLIGHT',
                    'HL_P50_shift': 'NO_HIGHLIGHT',
                    'HL_P05_shift': 'NO_HIGHLIGHT',
                    'HL_sticking_shift': 'NO_HIGHLIGHT',
                    'HL_trending': 'NO_HIGHLIGHT',
                    'HL_high_OOC': 'NO_HIGHLIGHT',
                    'HL_category_LT_shift': 'NO_HIGHLIGHT',
                    'HL_record_high_low': 'NO_HIGHLIGHT',
                    'record_high': False,
                    'record_low': False
                })
                print(f" - _process_discrete_chart: 基線數據不足，所有 OOB 設為 NO_HIGHLIGHT")

            print(f" - _process_discrete_chart: 離散型處理完成 {group_name}/{chart_name}")
            return result

        except Exception as e:
            print(f" - _process_discrete_chart: 處理錯誤 {group_name}/{chart_name}: {e}")
            traceback.print_exc()
            return None

    def build_result(self, result, image_path, weekly_image_path):
        violated_rules = result.get('violated_rules', {})
        we_true_keys = [k for k, v in violated_rules.items() if v]
        result['WE_Rule'] = ', '.join(we_true_keys) if we_true_keys else 'N/A'
        # 只要當週有任何點違規就亮 HL
        result['HL_WE'] = 'HIGHLIGHT' if we_true_keys else 'NO_HIGHLIGHT'

        oob_true_keys = [k for k in OOB_KEYS if result.get(k) == 'HIGHLIGHT']
        result['OOB_Rule'] = ', '.join(oob_true_keys) if oob_true_keys else 'N/A'

        for key in OOB_KEYS:
            result.pop(key, None)
        result.pop('violated_rules', None) # 移除原始的 violated_rules 字典

        result['chart_path'] = image_path
        result['weekly_chart_path'] = weekly_image_path

        # --- 強制補齊 group_name/chart_name 欄位，來源優先順序 ---
        if 'group_name' not in result or not result['group_name']:
            result['group_name'] = result.get('GroupName', 'N/A')
        if 'chart_name' not in result or not result['chart_name']:
            result['chart_name'] = result.get('ChartName', 'N/A')

        # 確保 Cpk 在 result 中，即使是 N/A 或 NaT
        if 'Cpk' not in result:
            result['Cpk'] = np.nan

        print(f" - build_result 完成更新 result for {result.get('group_name', 'Unknown')}/{result.get('chart_name', 'Unknown')}")

    def save_results(self):
        results_df = pd.DataFrame(self.results)

        # 確保所有預期的列都存在，包括新增的數據類型欄位
        expected_cols = ['data_cnt', 'ooc_cnt', 'WE_Rule', 'OOB_Rule', 'data_type', 'Material_no',
                         'group_name', 'chart_name', 'chart_ID', 'Characteristics',
                         'USL', 'LSL', 'UCL', 'LCL', 'Target', 'Cpk', 'Resolution',
                         'HL_record_high_low', 'record_high', 'record_low',
                         'chart_path', 'weekly_chart_path']
        for col in expected_cols:
            if col not in results_df.columns:
                results_df[col] = np.nan

        cols_to_order = [col for col in expected_cols if col in results_df.columns]
        results_df = results_df[cols_to_order]

        results_df = results_df.replace([np.nan, np.inf, -np.inf], 'N/A')

        try:
             save_results_to_excel(results_df)
             print("Results saved to Excel.")
        except Exception as e:
             print(f"Error saving results to Excel: {e}")
             self.show_error("Save Error", f"Failed to save results to Excel: {e}")

    def display_image(self, result, index):
        spc_image_path = result.get('chart_path')
        weekly_image_path = result.get('weekly_chart_path')

        if not spc_image_path or not os.path.exists(spc_image_path):
             print(f"[Warning] SPC chart image path invalid: {spc_image_path}")
             spc_chart_layout = QtWidgets.QVBoxLayout()
             spc_chart_layout.addWidget(QtWidgets.QLabel("SPC Chart Not Available"))
        else:
            spc_chart_layout = self.create_image_layout(spc_image_path)

        if not weekly_image_path or not os.path.exists(weekly_image_path):
             print(f"[Warning] Weekly chart image path invalid: {weekly_image_path}")
             weekly_chart_layout = QtWidgets.QVBoxLayout()
             weekly_chart_layout.addWidget(QtWidgets.QLabel("Weekly Chart Not Available"))
        else:
             weekly_chart_layout = self.create_image_layout(weekly_image_path)

        info_layout = QtWidgets.QVBoxLayout()
        info_layout.addWidget(self.create_info_label(result))

        self.image_grid_layout.addLayout(spc_chart_layout, index, 0)
        self.image_grid_layout.addLayout(weekly_chart_layout, index, 1)
        self.image_grid_layout.addLayout(info_layout, index, 2)

    def create_image_layout(self, image_path):
        layout = QtWidgets.QVBoxLayout()
        layout.setContentsMargins(30, 0, 0, 0)

        image_label = self.create_image_label(image_path, max_width=500, max_height=800)
        view_button = self.create_view_button(image_path)

        layout.addWidget(image_label)
        layout.addWidget(view_button, alignment=QtCore.Qt.AlignmentFlag.AlignCenter)
        return layout

    def create_info_label(self, result):
        table_style = """
            <style>
                .info-container {
                    margin-left: 45px;
                    margin-top: 5px;
                }
                table {
                    border-collapse: collapse;
                    width: 450px;
                }
                td, th {
                    padding: 2.5px;
                    text-align: left;
                    border: 1px solid #ddd;
                }
                th {
                    background-color: #344CB7;
                    color: white;
                    font-weight: bold;
                }
                tr:nth-child(even) {
                    background-color: #f4f4f4;
                }
                tr:hover {
                    background-color: #e0e0e0;
                }
                td {
                    color: #000957;
                }
                .title {
                    font-size: 16px;
                    font-weight: bold;
                    margin-bottom: 6px;
                    color: #344CB7;
                }
            </style>
        """

        info_text = f"""
            <html>
                {table_style}
                <div class="info-container">
                    <div class="title"> </div>
                    <table>
                        <thead>
                            <tr>
                                <th>Property</th>
                                <th>Value</th>
                            </tr>
                        </thead>
                        <tbody>
                            {''.join(self.create_table_row(key, result) for key in [
                                'data_cnt', 'ooc_cnt', 'WE_Rule', 'OOB_Rule', 'data_type', 'Material_no',
                                'group_name', 'chart_name', 'Cpk'
                            ])}
                        </tbody>
                    </table>
                </div>
            </html>
        """

        label = QtWidgets.QLabel(info_text, self)
        label.setFont(QtGui.QFont("Segoe UI", 10))
        label.setTextFormat(QtCore.Qt.TextFormat.RichText)
        label.setWordWrap(True)
        label.setAlignment(QtCore.Qt.AlignmentFlag.AlignTop)

        return label

    def create_table_row(self, key, result):
            value = result.get(key, 'N/A')

            if key == 'group_name' and value == 'Default':
                value = ''

            display_key = key.replace('_', ' ').title()

            if key in ['WE_Rule', 'OOB_Rule']:
                 value = value.replace(', ', '<br>')

            return f"<tr><td>{display_key}:</td><td>{value}</td></tr>"

    def create_view_button(self, image_path):
        button = QtWidgets.QPushButton("Zoom +", self)
        button.setFont(QtGui.QFont("Segoe UI", 7))
        button.setFixedWidth(85)
        button.setStyleSheet("""
            QPushButton {
                padding: 4px 8px;
                background-color: #344CB7;
                color: white;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #577BC1;
            }
        """)
        button.clicked.connect(lambda checked, path=image_path: self.show_full_image(path))
        return button

    def show_full_image(self, image_path):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("View Details")
        dialog.setGeometry(100, 100, 1400, 600)

        layout = QtWidgets.QVBoxLayout(dialog)

        if not os.path.exists(image_path):
             error_label = QtWidgets.QLabel("Image file not found.", dialog)
             error_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
             error_label.setStyleSheet("color: red;")
             layout.addWidget(error_label)
        else:
             scroll_area = QtWidgets.QScrollArea(dialog)
             scroll_area.setWidgetResizable(True)

             image_label = self.create_image_label(image_path, keep_original_size=True)

             scroll_area.setWidget(image_label)

             layout.addWidget(scroll_area)

        dialog.setLayout(layout)
        dialog.exec()

    def show_error(self, title, message, warning=False):
        if warning:
            QtWidgets.QMessageBox.warning(self, title, message)
        else:
            QtWidgets.QMessageBox.critical(self, title, message)

class SplitDataWidget(QtWidgets.QWidget):
    """
    一個 PyQt 介面，用於選擇輸入/輸出路徑和處理 CSV 檔案。
    這個 Widget 將會被加入到主視窗的 QStackedWidget 中。
    採用更現代的 UI 設計。
    """
    
    COMMON_ENCODINGS = ["utf-8-sig", "utf-8", "big5", "cp950", "latin1", "cp1252"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        self.apply_styles() # 應用樣式表

    def init_ui(self):
        """初始化使用者介面元素。"""
        # 主佈局採用網格佈局，提供更多控制
        main_layout = QtWidgets.QGridLayout(self)
        main_layout.setContentsMargins(30, 30, 30, 30) # 增加邊距
        main_layout.setSpacing(15) # 增加元件間距

        # --- 說明文字 (使用更簡潔的標題和引導) ---
        description_label = QtWidgets.QLabel(
            "<h2 style='color:#34495E;'>CSV File Splitter Tool</h2>"
            "<p style='color:#5D6D7E;'>This tool can split CSV files in specific formats into multiple independent CSV files.</p>"
            "<p style='color:#5D6D7E;'>If SPC Chart format is vertically arranged, please choose **Type2** splitting method.</p>"
            "<p style='color:#5D6D7E;'>For horizontal arrangement, choose **Type3** splitting method.</p>"
        )
        description_label.setWordWrap(True)
        description_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignTop)
        main_layout.addWidget(description_label, 0, 0, 1, 2) # Row 0, Col 0, Span 1 row, 2 columns

        # ...existing code...

        # --- 輸入檔案選擇區塊 ---
        # 使用 QFormLayout 在 QGroupBox 內組織標籤和輸入/按鈕對
        input_group_box = QtWidgets.QGroupBox("1. Select Input File")
        input_layout = QtWidgets.QFormLayout(input_group_box)
        input_layout.setContentsMargins(15, 20, 15, 15) # 調整 GroupBox 內部邊距
        input_layout.setHorizontalSpacing(10)

        self.input_path_entry = QtWidgets.QLineEdit()
        self.input_path_entry.setPlaceholderText("Please select one or multiple CSV files (separated by ';')...")
        self.input_path_entry.setReadOnly(True) # 設為只讀，只能通過按鈕選擇
        
        input_button = QtWidgets.QPushButton("Browse Files...")
        # *** PyQt6 兼容性修正 START ***
        input_button.setIcon(self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_DialogOpenButton))
        # *** PyQt6 兼容性修正 END ***
        input_button.clicked.connect(self.select_input_files)

        input_row_layout = QtWidgets.QHBoxLayout()
        input_row_layout.addWidget(self.input_path_entry)
        input_row_layout.addWidget(input_button)
        input_layout.addRow(input_row_layout)
        main_layout.addWidget(input_group_box, 1, 0, 1, 2) # Row 2, Col 0, Span 1 row, 2 columns

        # --- 輸出資料夾選擇區塊 ---
        output_group_box = QtWidgets.QGroupBox("2. Select Output Folder")
        output_layout = QtWidgets.QFormLayout(output_group_box)
        output_layout.setContentsMargins(15, 20, 15, 15)
        output_layout.setHorizontalSpacing(10)

        self.output_folder_entry = QtWidgets.QLineEdit()
        self.output_folder_entry.setPlaceholderText("Split files will be saved in 'raw_charts' subfolder under this directory...")
        self.output_folder_entry.setReadOnly(True) # 設為只讀
        
        output_button = QtWidgets.QPushButton("Browse Folder...")
        # *** PyQt6 兼容性修正 START ***
        output_button.setIcon(self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_DialogOpenButton))
        # *** PyQt6 兼容性修正 END ***
        output_button.clicked.connect(self.select_output_folder)

        output_row_layout = QtWidgets.QHBoxLayout()
        output_row_layout.addWidget(self.output_folder_entry)
        output_row_layout.addWidget(output_button)
        output_layout.addRow(output_row_layout)
        main_layout.addWidget(output_group_box, 2, 0, 1, 2) # Row 2, Col 0, Span 1 row, 2 columns

        # --- 處理模式選擇區塊 ---
        mode_group_box = QtWidgets.QGroupBox("3. Select Processing Mode")
        mode_layout = QtWidgets.QHBoxLayout(mode_group_box)
        mode_layout.setContentsMargins(15, 20, 15, 15)

        mode_label = QtWidgets.QLabel("Select File Type:")
        self.processing_mode_combo = QtWidgets.QComboBox()
        self.processing_mode_combo.addItems(["Type3_Horizontal (Horizontal Layout)", "Type2_Vertical (Vertical Stack)"])
        self.processing_mode_combo.setFixedWidth(250)  # 設定較窄寬度
        self.processing_mode_combo.currentIndexChanged.connect(self._update_processing_mode)
        self._current_processing_mode = "Type3_Horizontal" # 預設內部模式

        mode_layout.addWidget(mode_label)
        mode_layout.addWidget(self.processing_mode_combo)


        # --- 下載範例按鈕（垂直排列，推到最右側）---
        mode_layout.addStretch(1)
        example_buttons_layout = QtWidgets.QVBoxLayout()
        self.download_example_button = QtWidgets.QPushButton("Type3 Example")
        self.download_example_button.setIcon(self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_DialogSaveButton))
        self.download_example_button.setFixedSize(185, 36)
        self.download_example_button.clicked.connect(self.download_type3_example)
        example_buttons_layout.addWidget(self.download_example_button)

        self.download_type2_example_button = QtWidgets.QPushButton("Type2 Example")
        self.download_type2_example_button.setIcon(self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_DialogSaveButton))
        self.download_type2_example_button.setFixedSize(185, 36)
        self.download_type2_example_button.clicked.connect(self.download_type2_example)
        example_buttons_layout.addWidget(self.download_type2_example_button)

        mode_layout.addLayout(example_buttons_layout)

        main_layout.addWidget(mode_group_box, 3, 0, 1, 2) # Row 3, Col 0, Span 1 row, 2 columns

        # --- 處理按鈕 ---
        process_button = QtWidgets.QPushButton("Start Processing")
        process_button.setFixedSize(200, 50) # 更大的按鈕
        process_button.clicked.connect(self.run_processing)
        process_button.setObjectName("processButton") # 設定物件名稱，用於QSS

        # 將按鈕放在佈局中央
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch(1)
        button_layout.addWidget(process_button)
        button_layout.addStretch(1)
        main_layout.addLayout(button_layout, 4, 0, 1, 2) # Row 4, Col 0, Span 1 row, 2 columns

        # --- 進度條與狀態訊息 ---
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("Processing Progress: %p%")
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False) # 預設隱藏

        self.status_label = QtWidgets.QLabel("Ready.")
        self.status_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet("color: #607D8B; font-style: italic;")

        main_layout.addWidget(self.progress_bar, 5, 0, 1, 2) # Row 5
        main_layout.addWidget(self.status_label, 6, 0, 1, 2) # Row 6
        main_layout.setRowStretch(7, 1) # 將所有內容推到頂部

    def download_type2_example(self):
        import csv
        from PyQt6 import QtWidgets
        columns = ["GroupName", "ChartName", "point_time", "Batch_ID", "point_val"]
        data = [
            ["Group1", "A", "2025/3/10 00:45", 123, 56.5],
            ["Group1", "A", "2025/3/11 00:45", 123, 56.6],
            ["Group1", "A", "2025/3/12 00:45", 123, 56.5],
            ["Group1", "B", "2025/3/10 00:45", 123, 84],
            ["Group1", "B", "2025/3/11 00:45", 123, 84.2],
            ["Group1", "B", "2025/3/12 00:45", 123, 83.8],
        ]
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "儲存 type2 範例檔", "type2_example.csv", "CSV 檔案 (*.csv)")
        if save_path:
            with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                writer.writerows(data)
            QtWidgets.QMessageBox.information(self, "Complete", f"Saved to {save_path}")
    def download_type3_example(self):
        import csv
        # 範例資料
        data = [
            ["2025/3/10 00:45", 123, "", 56.5, 84, 123.3, 140, 0.0065, 16820, 16811, -0.11, -0.07, -0.06, 9044],
            ["2025/3/11 00:45", 123, "", 56.6, 84.2, 124, 140, 0.0065, 16748, 16813, -0.11, -0.06, -0.03, 9065],
            ["2025/3/12 00:45", 123, "", 56.5, 83.8, 123, 139.7, 0.0065, 16822, 16822, -0.1, -0.05, -0.13, 9030],
        ]
        columns1 = ["point_time", "Batch_ID", "GroupName", "Group1", "Group1", "Group1", "Group1", "Group1", "Group1", "Group1", "Group1", "Group1", "Group1", "Group1"]
        columns2 = ["", "", "ChartName", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]

        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "儲存 type3 範例檔", "type3_example.csv", "CSV 檔案 (*.csv)")
        if save_path:
            with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=",")
                writer.writerow(columns1)
                writer.writerow(columns2)
                for row in data:
                    writer.writerow(row)
            QtWidgets.QMessageBox.information(self, "Complete", f"Saved to {save_path}")


    def apply_styles(self):
        """應用 QSS 樣式表。"""
        self.setStyleSheet("""
            QWidget {
                font-family: "Segoe UI", "Microsoft YaHei", sans-serif;
                font-size: 14px;
                color: #333;
            }
            QLabel {
                color: #333;
            }
            QGroupBox {
                font-size: 15px;
                font-weight: bold;
                color: #2C3E50;
                margin-top: 10px;
                border: 1px solid #D1D1D1;
                border-radius: 8px;
                padding-top: 20px;
                padding-bottom: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
                left: 10px;
                margin-left: 5px;
                color: #2C3E50;
            }
            QLineEdit {
                border: 1px solid #BDC3C7;
                border-radius: 5px;
                padding: 8px;
                background-color: #ECF0F1;
                selection-background-color: #3498DB;
            }
            QPushButton {
                background-color: #344CB7;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #2980B9;
            }
            QPushButton:pressed {
                background-color: #1F618D;
            }
            #processButton {
                background-color: #344CB7;
                font-size: 16px;
                padding: 12px 25px;
                border-radius: 8px;
            }
            #processButton:hover {
                background-color: #2980B9;
                color: #fff;
            }
            #processButton:pressed {
                background-color: #1F618D;
            }
            QComboBox {
                border: 1px solid #BDC3C7;
                border-radius: 5px;
                padding: 8px;
                background-color: #ECF0F1;
                selection-background-color: #3498DB;
            }


            QProgressBar {
                border: 1px solid #BDC3C7;
                border-radius: 5px;
                text-align: center;
                background-color: #ECF0F1;
            }
            QProgressBar::chunk {
                background-color: #344CB7;
                border-radius: 5px;
            }
        """)

    def _update_processing_mode(self, index):
        """根據下拉選單的選擇更新內部處理模式。"""
        selected_text = self.processing_mode_combo.currentText()
        if "Type3_Horizontal" in selected_text:
            self._current_processing_mode = "Type3_Horizontal"
        elif "Type2_Vertical" in selected_text:
            self._current_processing_mode = "Type2_Vertical"

    def select_input_files(self):
        """開啟檔案對話框，選擇多個輸入 CSV 檔案。"""
        file_paths, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, "Select Input CSV Files (Multiple Selection)", "", "CSV Files (*.csv);;All Files (*.*)"
        )
        if file_paths:
            self.input_path_entry.setText(";".join(file_paths))
            self.status_label.setText(f"Selected {len(file_paths)} files.")

    def select_output_folder(self):
        """開啟資料夾對話框，選擇輸出資料夾。"""
        folder_path = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Select Output Folder"
        )
        if folder_path:
            self.output_folder_entry.setText(folder_path)
            self.status_label.setText(f"Selected output folder: {os.path.basename(folder_path)}")

    def sanitize_filename(self, name):
        """輔助函數：清理字串，使其適用於檔案名稱。"""
        invalid_chars = '<>:"/\\|?*\''
        for char in invalid_chars:
            name = name.replace(char, '')
        return name.strip() 

    def _read_csv_with_encoding_fallback(self, filepath, header_val=None):
        """
        嘗試用多種編碼讀取 CSV 檔案，直到成功為止。
        """
        for enc in self.COMMON_ENCODINGS:
            try:
                df = pd.read_csv(filepath, header=header_val, encoding=enc)
                print(f"檔案 '{os.path.basename(filepath)}' 已成功使用 '{enc}' 編碼讀取。")
                return df
            except (UnicodeDecodeError, pd.errors.ParserError) as e:
                print(f"嘗試使用 '{enc}' 編碼讀取 '{os.path.basename(filepath)}' 失敗: {e}")
                continue
            except Exception as e:
                print(f"讀取檔案 '{os.path.basename(filepath)}' 時發生意外錯誤: {e}")
                raise 

        raise ValueError(f"無法使用任何嘗試的編碼讀取檔案 '{os.path.basename(filepath)}'。")

    def _process_type3_horizontal_csv(self, input_path, final_output_folder):
        """處理 Type3 (水平展開) 的 CSV 檔案。"""
        try:
            self.new_method(input_path)
            print(f"\n--- 正在處理 Type3 (水平展開) 檔案：{os.path.basename(input_path)} ---")
            df = self._read_csv_with_encoding_fallback(input_path, header_val=None)

            new_columns = []
            for col1, col2 in zip(df.iloc[0], df.iloc[1]):
                if pd.isna(col2):
                    new_columns.append(str(col1))
                elif pd.isna(col1):
                    new_columns.append(str(col2))
                else:
                    new_columns.append(f"{col1}_{col2}")

            df = df.iloc[2:].copy()
            df.columns = new_columns

            chartname_col_name = None
            for col in df.columns:
                if 'GroupName' in col and 'ChartName' in col:
                    chartname_col_name = col
                    break
            
            if chartname_col_name is None:
                QtWidgets.QMessageBox.critical(self, "Error", f"File {os.path.basename(input_path)}: Cannot find 'GroupName' and 'ChartName' combination column. This file will be skipped.")
                return False 
            
            chartname_idx = df.columns.get_loc(chartname_col_name)

            universal_info_columns = df.columns[:chartname_idx + 1].tolist()
            chart_columns = df.columns[(chartname_idx + 1):]

            for chart_col in chart_columns:
                temp_df = df[universal_info_columns].copy()
                temp_df['point_val'] = df[chart_col]
                
                if '_' in chart_col:
                    groupname, chartname = chart_col.split('_', 1)
                else:
                    groupname = ''
                    chartname = chart_col

                temp_df['GroupName'] = groupname
                temp_df['ChartName'] = chartname

                if 'point_time' in temp_df.columns:
                    try:
                        temp_df['point_time'] = pd.to_datetime(temp_df['point_time'], errors='coerce') 
                        temp_df['point_time'] = temp_df['point_time'].dt.strftime('%Y/%#m/%#d %H:%M') 
                    except Exception as time_e:
                        print(f"警告：處理檔案 {os.path.basename(input_path)} 中 'point_time' 欄位時發生錯誤：{time_e}")
                        print("該欄位將以原始格式輸出或包含 NaT 值。")
                        
                final_columns_order = ['GroupName', 'ChartName', 'point_time', 'point_val']
                for col in universal_info_columns:
                    if col not in final_columns_order and col != chartname_col_name:
                        final_columns_order.append(col)
                
                existing_final_columns_order = [col for col in final_columns_order if col in temp_df.columns]
                temp_df = temp_df[existing_final_columns_order]

                safe_groupname = self.sanitize_filename(groupname)
                safe_chartname = self.sanitize_filename(chartname)
                
                output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
                
                if not temp_df.empty: 
                    temp_df.to_csv(output_file, index=False, encoding='utf-8-sig')
                    print(f"已輸出：{os.path.basename(output_file)}")
            return True 

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error processing file {os.path.basename(input_path)}: {e}")
            return False 

    def new_method(self, input_path):
        self.status_label.setText(f"Processing: Type3 file {os.path.basename(input_path)}...")

    def _process_type2_vertical_csv(self, input_path, final_output_folder):
        """處理 Type2 (垂直堆疊 / long format) 的 CSV 檔案。"""
        try:
            self.status_label.setText(f"Processing: Type2 file {os.path.basename(input_path)}...")
            print(f"\n--- 正在處理 Type2 (垂直堆疊) 檔案：{os.path.basename(input_path)} ---")
            df = self._read_csv_with_encoding_fallback(input_path, header_val='infer') 
            
            required_cols = ['GroupName', 'ChartName', 'point_time', 'point_val']
            if not all(col in df.columns for col in required_cols):
                missing_cols = [col for col in required_cols if col not in df.columns]
                QtWidgets.QMessageBox.critical(self, "Error", 
                                     f"Type2 (Vertical Stack) file {os.path.basename(input_path)} is missing required columns: {', '.join(missing_cols)}.\n"
                                     "Please ensure the file contains 'GroupName', 'ChartName', 'point_time', 'point_val' columns.")
                return False

            if 'point_time' in df.columns:
                try:
                    df['point_time'] = pd.to_datetime(df['point_time'], errors='coerce') 
                    df['point_time'] = df['point_time'].dt.strftime('%Y/%#m/%#d %H:%M')
                except Exception as time_e:
                    print(f"警告：處理檔案 {os.path.basename(input_path)} 中 'point_time' 欄位時發生錯誤：{time_e}")
                    print("該欄位將以原始格式輸出或包含 NaT 值。")

            unique_combinations = df[['GroupName', 'ChartName']].drop_duplicates()

            processed_any_sub_file = False
            for i, (index, row) in enumerate(unique_combinations.iterrows()):
                groupname = row['GroupName']
                chartname = row['ChartName']

                temp_df = df[(df['GroupName'] == groupname) & (df['ChartName'] == chartname)].copy()

                other_cols = [col for col in temp_df.columns if col not in ['GroupName', 'ChartName', 'point_time', 'point_val']]
                final_cols_order = ['GroupName', 'ChartName', 'point_time', 'point_val'] + other_cols
                
                existing_final_cols = [col for col in final_cols_order if col in temp_df.columns]
                temp_df = temp_df[existing_final_cols]

                safe_groupname = self.sanitize_filename(str(groupname)) 
                safe_chartname = self.sanitize_filename(str(chartname)) 

                output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
                
                if not temp_df.empty: 
                    temp_df.to_csv(output_file, index=False, encoding='utf-8-sig')
                    print(f"已輸出：{os.path.basename(output_file)}")
                    processed_any_sub_file = True
                
                # 更新進度條（針對單一 Type2 檔案內的多個子檔案）
                progress = int((i + 1) / len(unique_combinations) * 100)
                self.progress_bar.setValue(progress)
                self.status_label.setText(f"Processing: {os.path.basename(input_path)} - {i+1}/{len(unique_combinations)} sub-charts")
                QtWidgets.QApplication.processEvents() # 強制更新 UI

            if not processed_any_sub_file:
                QtWidgets.QMessageBox.warning(self, "Warning", f"File {os.path.basename(input_path)} generated no split files. Please check its content.")
                return False

            return True 

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error processing file {os.path.basename(input_path)}: {e}")
            return False 

    def run_processing(self):
        """
        執行 CSV 處理邏輯，處理多個 CSV 檔案。
        """
        input_paths_str = self.input_path_entry.text()
        input_paths = [path.strip() for path in input_paths_str.split(';') if path.strip()]
        
        base_output_folder = self.output_folder_entry.text()
        # 從內部模式變數獲取，而不是直接從下拉選單獲取顯示文本
        processing_mode = self._current_processing_mode 

        if not input_paths or not base_output_folder:
            QtWidgets.QMessageBox.warning(self, "Warning", "Please select at least one input file and output folder!")
            return

        final_output_folder = os.path.join(base_output_folder, "raw_charts")
        
        try:
            os.makedirs(final_output_folder, exist_ok=True)
            self.status_label.setText(f"Created output folder: {os.path.basename(final_output_folder)}")
            print(f"已建立輸出資料夾：{final_output_folder}")
        except OSError as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Cannot create 'raw_charts' folder: {final_output_folder}\nError message: {e}")
            return
        
        # 顯示進度條
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting file processing...")
        QtWidgets.QApplication.processEvents() # 強制更新 UI

        processed_count = 0
        failed_files = []

        total_files = len(input_paths)
        for i, input_path in enumerate(input_paths):
            success = False
            self.status_label.setText(f"Processing file {i+1}/{total_files}: {os.path.basename(input_path)}")
            # 更新整體進度條 (每個檔案平均分配進度)
            overall_progress = int((i / total_files) * 100) 
            self.progress_bar.setValue(overall_progress)
            QtWidgets.QApplication.processEvents() # 強制更新 UI

            try: 
                if processing_mode == "Type3_Horizontal":
                    success = self._process_type3_horizontal_csv(input_path, final_output_folder)
                elif processing_mode == "Type2_Vertical":
                    # Type2 內部會自己更新進度條，這裡只需確保它的起始值正確
                    success = self._process_type2_vertical_csv(input_path, final_output_folder)
            except ValueError as ve: 
                QtWidgets.QMessageBox.critical(self, "Error", f"File {os.path.basename(input_path)} read failed: {ve}\nPlease ensure the file is a valid CSV format.")
                failed_files.append(os.path.basename(input_path))
                continue
            except Exception as e: 
                QtWidgets.QMessageBox.critical(self, "Error", f"Unexpected error processing file {os.path.basename(input_path)}: {e}")
                failed_files.append(os.path.basename(input_path))
                continue
            
            if success:
                processed_count += 1
            else:
                failed_files.append(os.path.basename(input_path))
        
        # 處理完成，設定進度條為100%
        self.progress_bar.setValue(100)
        self.progress_bar.setVisible(False) # 處理完畢後隱藏進度條

        if processed_count > 0:
            if not failed_files:
                QtWidgets.QMessageBox.information(self, "Complete", f"Successfully processed all {processed_count} files!")
                self.status_label.setText("All files processed successfully.")
            else:
                QtWidgets.QMessageBox.warning(self, "Partially Complete", f"Processed {processed_count} files. The following files failed:\n{', '.join(failed_files)}")
                self.status_label.setText("Some files failed processing, please check messages.")
        else:
            QtWidgets.QMessageBox.critical(self, "Error", "No files were successfully processed.")
            self.status_label.setText("Processing failed: No files were successfully processed.")

if __name__ == "__main__" and UI_AVAILABLE:
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("windowsvista")
    spc_app = SPCApp()
    spc_app.show()
    sys.exit(app.exec())
