import os

# 強制限制底層 C 語言庫的執行緒數量，防止 Linux 發生 Segmentation fault
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

import gc
import math
import base64
import tempfile
import shutil
import uuid
import asyncio
import threading
import json
from datetime import datetime, date, timedelta
from io import BytesIO
from typing import List, Optional, Dict, Any, Union

# 確保伺服器端繪圖不會觸發 GUI 報錯 (Headless backend)
os.environ.setdefault("MPLBACKEND", "Agg")

import pandas as pd
import numpy as np
from scipy import stats

# Matplotlib 全域設定
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.transforms as mtransforms
import matplotlib.gridspec as gridspec
from matplotlib import cm

# Excel 匯出套件
import xlsxwriter
import openpyxl
from openpyxl.drawing.image import Image as XLImage

from fastapi import FastAPI, HTTPException
from contextlib import asynccontextmanager
from pydantic import BaseModel, Field

import multiprocessing as mp
from multiprocessing import Process, Manager
try:
    mp.set_start_method('spawn', force=True)
except RuntimeError:
    pass

# ==========================================
# 內部模組 Imports (oob_eng & tool matching)
# ==========================================
try:
    from oob_eng import (
        resource_path,
        load_execution_time,
        load_chart_information,
        preprocess_data,
        find_matching_file,
        determine_data_type,
        process_single_chart,
        plot_spc_chart,
        plot_weekly_spc_chart,
        plot_spc_by_tool_color,
        plot_spc_by_tool_group,
        plot_qq_plot,
        calculate_cpk,
        ooc_calculator,
        review_ooc_results,
        discrete_oob_calculator,
        record_high_low_calculator,
        save_results_to_excel,
    )
except Exception as e:
    raise RuntimeError(
        "Failed to import required functions from oob_eng.py. "
        "Please ensure all dependencies are installed. Error: " + str(e)
    )

try:
    from tool_matching_widget_osat import analyze_tool_matching_data
except Exception as e:
    print(f"Warning: Could not import tool matching functions: {e}")
    # 定義 dummy function 避免程式崩潰
    def analyze_tool_matching_data(*args, **kwargs):
        raise HTTPException(status_code=500, detail="Tool matching functions not available")


import cpk_eng

# ==========================================
# FastAPI 實體與任務狀態儲存
# ==========================================

# 主進程的任務狀態字典，由 lifespan 初始化為 Manager().dict()
# 此處先為 plain dict 佔位，避免子進程 import 時意外觸發 Manager()
task_status_db: Dict[str, Any] = {}

def update_task_status(task_id: str, updates: dict, db=None):
    """跨進程安全更新狀態的幫手函數，解決 Manager.dict 的巢狀更新問題。
    worker 進程需傳入 db=shared_db（Manager DictProxy）以確保跨進程可見。"""
    _db = db if db is not None else task_status_db
    if task_id in _db:
        current_data = dict(_db[task_id])
        current_data.update(updates)
        _db[task_id] = current_data

# 任務超過此時間後即可被清除 (24 小時)
_TASK_TTL_HOURS = 24
# Processing 狀態超過此時間視為卡死，watchdog 會強制 kill (2 小時)
_TASK_PROCESSING_TIMEOUT_HOURS = 2


def _result_json_path(task_id: str) -> str:
    """返回此 task 的結果 JSON 檔案絕對路徑（寫入子進程、讀取主進程）。"""
    return os.path.abspath(os.path.join("output", task_id, "result.json"))


async def _cleanup_expired_tasks() -> None:
    """每小時執行一次：清除 task_status_db 中已完成/失敗且超過 TTL 的任務，
    以及卡在 processing 超過 _TASK_PROCESSING_TIMEOUT_HOURS 的殭屍任務。"""
    while True:
        await asyncio.sleep(3600)  # 每小時清一次
        try:
            now = datetime.now()
            expired = [
                tid for tid, t in list(task_status_db.items())
                if (
                    t.get("status") in ("completed", "failed")
                    and t.get("expires_at")
                    and datetime.fromisoformat(t["expires_at"]) <= now
                ) or (
                    t.get("status") == "processing"
                    and t.get("created_at")
                    and (now - datetime.strptime(t["created_at"], "%Y-%m-%d %H:%M:%S")).total_seconds()
                        > _TASK_PROCESSING_TIMEOUT_HOURS * 3600
                )
            ]
            for tid in expired:
                out_dir = os.path.join("output", tid)
                if os.path.isdir(out_dir):
                    shutil.rmtree(out_dir, ignore_errors=True)
                task_status_db.pop(tid, None)
            if expired:
                print(f"[Cleanup] 清除 {len(expired)} 筆過期任務 (含卡死)")
        except asyncio.CancelledError:
            raise
        except Exception as e:
            print(f"[Cleanup] 清理週期發生例外，跳過此次：{e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    global task_status_db
    _mgr = Manager()
    task_status_db = _mgr.dict()  # 僅在主進程 lifespan 中初始化一次
    _cleanup_task = asyncio.create_task(_cleanup_expired_tasks())
    yield
    _cleanup_task.cancel()
    try:
        await _cleanup_task
    except asyncio.CancelledError:
        pass
    _mgr.shutdown()


app = FastAPI(title="OOB/SPC FastAPI", version="1.0.0", lifespan=lifespan)


# ==========================================
# 資料模型 (Pydantic Models)
# ==========================================
# --- Process / OOB Models ---
class ProcessRequest(BaseModel):
    filepath: Optional[str] = Field(default=None, description="Path to All_Chart_Information.xlsx")
    raw_data_directory: Optional[str] = Field(default=None, description="Directory containing raw chart CSV files")
    split_id: Optional[str] = Field(default=None, description="UUID returned by /split; takes precedence over raw_data_directory")
    save_excel: bool = Field(default=True, description="Save Excel report with images")
    scale_factor: float = Field(default=0.3, description="Image scale factor in Excel")
    limit_charts: Optional[int] = Field(default=None, description="Limit number of charts to process")
    base_date: Optional[str] = Field(default=None, description="Analysis base date (YYYY-MM-DD); used as weekly_end_date")

class ProcessSummary(BaseModel):
    total_charts: int
    processed_charts: int
    skipped_charts: int
    excel_output: Optional[str] = None

class ResultItem(BaseModel):
    data_cnt: Optional[int] = None
    ooc_cnt: Optional[int] = None
    oos_cnt: Optional[int] = None
    WE_Rule: Optional[str] = None
    OOB_Rule: Optional[str] = None
    data_type: Optional[str] = None
    Material_no: Optional[str] = None
    group_name: Optional[str] = None
    chart_name: Optional[str] = None
    chart_ID: Optional[str] = None
    Characteristics: Optional[str] = None
    USL: Optional[float] = None
    LSL: Optional[float] = None
    UCL: Optional[float] = None
    LCL: Optional[float] = None
    Target: Optional[float] = None
    Cpk: Optional[float] = None
    Resolution: Optional[float] = None
    HL_record_high_low: Optional[str] = None
    record_high: Optional[bool] = None
    record_low: Optional[bool] = None
    baseline_insufficient: Optional[bool] = False
    no_data: Optional[bool] = False
    no_data_reason: Optional[str] = None
    chart_path: Optional[str] = None
    weekly_chart_path: Optional[str] = None
    by_tool_color_path: Optional[str] = None
    by_tool_group_path: Optional[str] = None
    qq_plot_path: Optional[str] = None
    chart_data: Optional[List[Dict[str, Any]]] = None
    weekly_start: Optional[str] = None
    weekly_end: Optional[str] = None

class ProcessResponse(BaseModel):
    summary: ProcessSummary
    results: List[ResultItem]

# --- Tool Matching Models ---
class ToolMatchingRequest(BaseModel):
    chart_excel_path: Optional[str] = Field(default=None, description="Path to OOB chart info Excel (e.g. All_Chart_Information.xlsx)")
    raw_data_directory: Optional[str] = Field(default=None, description="Directory containing raw CSV files for batch analysis")
    mean_index_threshold: float = Field(default=1.0, description="Mean index threshold")
    sigma_index_threshold: float = Field(default=2.0, description="Sigma index threshold")
    use_statistical_test: bool = Field(default=False, description="Use statistical test instead of index")
    statistical_method: str = Field(default="unpaired", description="Statistical test method: unpaired or paired")
    alpha_level: float = Field(default=0.05, description="Significance level for statistical test")
    fill_sample_size: int = Field(default=5, description="Minimum sample size per tool for 1M mean window")
    filter_mode: str = Field(default="latest_data", description="Data filter mode: all_data, specified_date, latest_data")
    base_date: Optional[str] = Field(default=None, description="Base date for 1M/6M window analysis (YYYY-MM-DD); defaults to today")

class ToolMatchingSummary(BaseModel):
    total_groups: int
    abnormal_groups: int

class ToolMatchingResultItem(BaseModel):
    gname: str
    cname: str
    group: str
    group_all: str
    mean_index: Union[str, float]
    sigma_index: Union[str, float]
    k_value: Union[str, float]
    mean: float
    std: float
    mean_median: Union[str, float]
    sigma_median: Union[str, float]
    n: int
    characteristic: str
    need_matching: bool = False
    abnormal_type: str = ""
    spc_chart_path: Optional[str] = None
    boxplot_chart_path: Optional[str] = None
    timeline_chart_path: Optional[str] = None

class ToolMatchingResponse(BaseModel):
    summary: ToolMatchingSummary
    results: List[ToolMatchingResultItem]
    excel_output: Optional[str] = None

# --- SPC CPK Dashboard Models ---
class SPCCpkRequest(BaseModel):
    chart_excel_path: Optional[str] = Field(default=None, description="Path to All_Chart_Information.xlsx")
    raw_data_directory: Optional[str] = Field(default=None, description="Directory containing raw chart CSV files")
    start_date: Optional[date] = Field(default=None, description="Start date for analysis")
    end_date: Optional[date] = Field(default=None, description="End date for analysis")
    custom_mode: bool = Field(default=False, description="Whether to use custom time range mode")
    selected_chart: Optional[str] = Field(default=None, description="Specific chart to analyze")

class SPCCpkMetrics(BaseModel):
    cpk: Optional[float] = None
    cpk_l1: Optional[float] = None
    cpk_l2: Optional[float] = None
    custom_cpk: Optional[float] = None
    r1: Optional[float] = None
    r2: Optional[float] = None
    k_value: Optional[float] = None

class SPCChartInfo(BaseModel):
    group_name: str
    chart_name: str
    characteristics: str
    usl: Optional[float] = None
    lsl: Optional[float] = None
    target: Optional[float] = None
    metrics: SPCCpkMetrics
    chart_image: Optional[str] = None  
    mean_current: Optional[float] = None
    sigma_current: Optional[float] = None
    mean_last_month: Optional[float] = None
    sigma_last_month: Optional[float] = None
    mean_last2_month: Optional[float] = None
    sigma_last2_month: Optional[float] = None
    mean_all: Optional[float] = None
    sigma_all: Optional[float] = None

class SPCCpkResponse(BaseModel):
    charts: List[SPCChartInfo]
    summary: Dict[str, Any]
    excel_path: Optional[str] = None

# --- Split Models ---
class SplitRequest(BaseModel):
    mode: str = Field(description="Split mode: 'Type3_Horizontal', 'Type2_Vertical', 'Vendor_Vertical', or 'Test_Horizontal'", pattern="^(Type3_Horizontal|Type2_Vertical|Vendor_Vertical|Test_Horizontal)$")
    input_files: List[str] = Field(description="List of CSV file paths to split")
    output_folder: Optional[str] = Field(default=None, description="Base output folder")


# ==========================================
# 共用工具函數 (Utils)
# ==========================================
def _default_paths() -> Dict[str, str]:
    filepath = resource_path("input/All_Chart_Information.xlsx")
    raw_dir = resource_path("input/raw_charts/")
    return {
        "filepath": filepath,
        "raw_dir": raw_dir,
        "chart_excel_path": filepath,
        "raw_data_directory": raw_dir,
    }

def _read_csv_cached(cache: Dict[str, pd.DataFrame], filepath: str) -> Optional[pd.DataFrame]:
    try:
        if filepath not in cache:
            cache[filepath] = pd.read_csv(filepath)
        return cache[filepath].copy()
    except Exception as e:
        print(f"[Error] Failed to read CSV {filepath}: {e}")
        return None

def _sanitize_filename(name: str) -> str:
    invalid_chars = '<>:"/\\|?*\''
    for ch in invalid_chars:
        name = name.replace(ch, "")
    return name.strip()

def _read_csv_with_encoding_fallback(filepath: str, header_val=None) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "big5", "cp950", "latin1", "cp1252"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            return pd.read_csv(filepath, header=header_val, encoding=enc)
        except Exception as e:
            last_err = e
            continue
    raise ValueError(f"Failed to read {os.path.basename(filepath)} with common encodings: {last_err}")


# ==========================================
# 資料切割模組 (Split Logic)
# ==========================================
def _split_type3_horizontal(input_path: str, final_output_folder: str) -> bool:
    try:
        df = _read_csv_with_encoding_fallback(input_path, header_val=None)
        new_columns: List[str] = []
        for col1, col2 in zip(df.iloc[0], df.iloc[1]):
            if pd.isna(col2):
                new_columns.append(str(col1))
            elif pd.isna(col1):
                new_columns.append(str(col2))
            else:
                new_columns.append(f"{col1}_{col2}")

        df = df.iloc[2:].copy()
        df.columns = new_columns

        chartname_col_name = None
        for col in df.columns:
            if "GroupName" in col and "ChartName" in col:
                chartname_col_name = col
                break
        if chartname_col_name is None:
            raise ValueError("Cannot find combined 'GroupName' and 'ChartName' header column")

        chartname_idx = df.columns.get_loc(chartname_col_name)
        universal_info_columns = df.columns[: chartname_idx + 1].tolist()
        chart_columns = df.columns[(chartname_idx + 1) :]

        for chart_col in chart_columns:
            temp_df = df[universal_info_columns].copy()
            temp_df["point_val"] = df[chart_col]
            if "_" in chart_col:
                groupname, chartname = chart_col.split("_", 1)
            else:
                groupname = ""
                chartname = chart_col
            temp_df["GroupName"] = groupname
            temp_df["ChartName"] = chartname
            if "point_time" in temp_df.columns:
                try:
                    temp_df["point_time"] = pd.to_datetime(temp_df["point_time"], errors="coerce")
                    temp_df["point_time"] = temp_df["point_time"].dt.strftime("%Y/%m/%d %H:%M")
                except Exception:
                    pass

            final_columns_order = ["GroupName", "ChartName", "point_time", "point_val"]
            for col in universal_info_columns:
                if col not in final_columns_order and col != chartname_col_name:
                    final_columns_order.append(col)
            existing_cols = [c for c in final_columns_order if c in temp_df.columns]
            temp_df = temp_df[existing_cols]

            safe_groupname = _sanitize_filename(str(groupname))
            safe_chartname = _sanitize_filename(str(chartname))
            output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
            if not temp_df.empty:
                temp_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"[Error] Type3 split failed for {os.path.basename(input_path)}: {e}")
        return False

def _split_type2_vertical(input_path: str, final_output_folder: str) -> bool:
    try:
        df = _read_csv_with_encoding_fallback(input_path, header_val="infer")
        required_cols = ["GroupName", "ChartName", "point_time", "point_val"]
        if not all(col in df.columns for col in required_cols):
            missing = [c for c in required_cols if c not in df.columns]
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        if "point_time" in df.columns:
            try:
                df["point_time"] = pd.to_datetime(df["point_time"], errors="coerce")
                df["point_time"] = df["point_time"].dt.strftime("%Y/%m/%d %H:%M")
            except Exception:
                pass
        uniq = df[["GroupName", "ChartName"]].drop_duplicates()
        for _, row in uniq.iterrows():
            groupname = row["GroupName"]
            chartname = row["ChartName"]
            temp_df = df[(df["GroupName"] == groupname) & (df["ChartName"] == chartname)].copy()
            other_cols = [c for c in temp_df.columns if c not in ["GroupName", "ChartName", "point_time", "point_val"]]
            final_cols = ["GroupName", "ChartName", "point_time", "point_val"] + other_cols
            existing = [c for c in final_cols if c in temp_df.columns]
            temp_df = temp_df[existing]
            safe_groupname = _sanitize_filename(str(groupname))
            safe_chartname = _sanitize_filename(str(chartname))
            output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
            if not temp_df.empty:
                temp_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"[Error] Type2 split failed for {os.path.basename(input_path)}: {e}")
        return False

def _split_vendor_vertical(input_path: str, final_output_folder: str) -> bool:
    try:
        df = _read_csv_with_encoding_fallback(input_path, header_val="infer")
        # 欄位對應：廠商格式 -> 標準格式（Lot Mean / Lot Mean Valid 均相容）
        lot_mean_col = "Lot Mean Valid" if "Lot Mean Valid" in df.columns else "Lot Mean"
        vendor_col_map = {
            "Part ID":     "GroupName",
            "Item Name":   "ChartName",
            "Report Time": "point_time",
            lot_mean_col:  "point_val",
            "Vendor Site": "Matching",
        }
        required_cols = ["Part ID", "Item Name", "Report Time", lot_mean_col, "Vendor Site"]
        missing = [src for src in required_cols if src not in df.columns]
        if missing:
            raise ValueError(f"Missing required vendor columns: {', '.join(missing)}")
        df = df.rename(columns=vendor_col_map)
        if "point_time" in df.columns:
            try:
                df["point_time"] = pd.to_datetime(df["point_time"], errors="coerce")
                df["point_time"] = df["point_time"].dt.strftime("%Y/%m/%d %H:%M")
            except Exception:
                pass
        uniq = df[["GroupName", "ChartName"]].drop_duplicates()
        for _, row in uniq.iterrows():
            groupname = row["GroupName"]
            chartname = row["ChartName"]
            temp_df = df[(df["GroupName"] == groupname) & (df["ChartName"] == chartname)].copy()
            other_cols = [c for c in temp_df.columns if c not in ["GroupName", "ChartName", "point_time", "point_val"]]
            final_cols = ["GroupName", "ChartName", "point_time", "point_val"] + other_cols
            existing = [c for c in final_cols if c in temp_df.columns]
            temp_df = temp_df[existing]
            safe_groupname = _sanitize_filename(str(groupname))
            safe_chartname = _sanitize_filename(str(chartname))
            output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
            if not temp_df.empty:
                temp_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"[Error] Vendor split failed for {os.path.basename(input_path)}: {e}")
        return False

def _split_test_horizontal(input_path: str, final_output_folder: str) -> bool:
    try:
        df = _read_csv_with_encoding_fallback(input_path, header_val="infer")
        # 欄位重命名：測試橫向格式 -> 標準格式
        test_col_map = {
            "Part ID":          "GroupName",
            "FT Test End Time":  "point_time",
            "Test Site":        "Matching",
        }
        missing = [src for src in test_col_map if src not in df.columns]
        if missing:
            raise ValueError(f"Missing required test columns: {', '.join(missing)}")
        df = df.rename(columns=test_col_map)
        if "point_time" in df.columns:
            try:
                df["point_time"] = pd.to_datetime(df["point_time"], errors="coerce")
                df["point_time"] = df["point_time"].dt.strftime("%Y/%m/%d %H:%M")
            except Exception:
                pass
        # Matching 欄位之後的所有欄位都是測試項目（水平展開）
        matching_idx = df.columns.get_loc("Matching")
        id_cols = df.columns[:matching_idx + 1].tolist()  # GroupName, point_time, Matching, ...
        value_cols = df.columns[matching_idx + 1:].tolist()
        if not value_cols:
            raise ValueError("No test item columns found after 'Matching' column")
        # melt: 欄表頭 -> ChartName，欄內對應對比 -> point_val
        df_melted = df.melt(
            id_vars=id_cols,
            value_vars=value_cols,
            var_name="ChartName",
            value_name="point_val",
        )
        df_melted = df_melted.dropna(subset=["point_val"])
        # 依 GroupName + ChartName 切片輸出
        standard_cols = ["GroupName", "ChartName", "point_time", "point_val", "Matching"]
        uniq = df_melted[["GroupName", "ChartName"]].drop_duplicates()
        for _, row in uniq.iterrows():
            groupname = row["GroupName"]
            chartname = row["ChartName"]
            temp_df = df_melted[
                (df_melted["GroupName"] == groupname) & (df_melted["ChartName"] == chartname)
            ].copy()
            existing = [c for c in standard_cols if c in temp_df.columns]
            temp_df = temp_df[existing]
            safe_groupname = _sanitize_filename(str(groupname))
            safe_chartname = _sanitize_filename(str(chartname))
            output_file = os.path.join(final_output_folder, f"{safe_groupname}_{safe_chartname}.csv")
            if not temp_df.empty:
                temp_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"[Error] Test_Horizontal split failed for {os.path.basename(input_path)}: {e}")
        return False
def _preprocess_chart_types(all_charts_info: pd.DataFrame, raw_data_directory: str) -> Dict[str, str]:
    chart_types: Dict[str, str] = {}
    processed_files = set()
    for _, info in all_charts_info.iterrows():
        group_name = str(info.get("GroupName", "Unknown"))
        chart_name = str(info.get("ChartName", "Unknown"))
        chart_key = f"{group_name}_{chart_name}"
        csv_path = find_matching_file(raw_data_directory, group_name, chart_name)
        if csv_path and os.path.exists(csv_path) and csv_path not in processed_files:
            try:
                head = pd.read_csv(csv_path, nrows=1000)
                if "point_val" in head.columns:
                    data_type = determine_data_type(head["point_val"].dropna())
                else:
                    data_type = "continuous"
                chart_types[chart_key] = data_type
                processed_files.add(csv_path)
            except Exception as e:
                chart_types[chart_key] = "continuous"
        else:
            chart_types[chart_key] = "continuous"
    return chart_types

def _build_result_api(result: Dict[str, Any], violated_rules: Dict[str, bool], image_path: str, weekly_image_path: str, qq_plot_path: Optional[str] = None) -> Dict[str, Any]:
    we_true_keys = [k for k, v in (violated_rules or {}).items() if v]
    result["WE_Rule"] = ", ".join(we_true_keys) if we_true_keys else "N/A"
    result["HL_WE"] = "HIGHLIGHT" if we_true_keys else "NO_HIGHLIGHT"

    OOB_KEYS = ["HL_P95_shift", "HL_P50_shift", "HL_P05_shift", "HL_sticking_shift", 
                "HL_trending", "HL_high_OOC", "HL_record_high_low", "HL_category_LT_shift"]
    oob_true_keys = [k for k in OOB_KEYS if result.get(k) == "HIGHLIGHT"]
    result["OOB_Rule"] = ", ".join(oob_true_keys) if oob_true_keys else "N/A"

    for k in OOB_KEYS:
        result.pop(k, None)
    result.pop("violated_rules", None)

    result["chart_path"] = image_path
    result["weekly_chart_path"] = weekly_image_path
    result["by_tool_color_path"] = result.get("by_tool_color_path", None)
    result["by_tool_group_path"] = result.get("by_tool_group_path", None)
    result["qq_plot_path"] = qq_plot_path

    result["group_name"] = str(result.get("group_name", result.get("GroupName", "N/A")))
    result["chart_name"] = str(result.get("chart_name", result.get("ChartName", "N/A")))
    if "Cpk" not in result:
        result["Cpk"] = np.nan

    return result

def _process_discrete_chart_api(
    raw_df: pd.DataFrame, chart_info: Dict[str, Any], weekly_start_date: pd.Timestamp,
    weekly_end_date: pd.Timestamp, initial_baseline_start_date: pd.Timestamp, baseline_end_date: pd.Timestamp,
) -> Optional[Dict[str, Any]]:
    try:
        baseline_one_year = raw_df[(raw_df["point_time"] >= initial_baseline_start_date) & (raw_df["point_time"] <= baseline_end_date)].copy()
        baseline_count_one_year = len(baseline_one_year)
        baseline_insufficient = False

        if baseline_count_one_year < 10:
            actual_baseline_start = baseline_end_date - pd.Timedelta(days=365 * 2)
            baseline_two_year = raw_df[(raw_df["point_time"] >= actual_baseline_start) & (raw_df["point_time"] <= baseline_end_date)].copy()
            if len(baseline_two_year) < 10:
                baseline_insufficient = True
        else:
            actual_baseline_start = initial_baseline_start_date

        baseline_data = raw_df[(raw_df["point_time"] >= actual_baseline_start) & (raw_df["point_time"] <= baseline_end_date)].copy()
        weekly_data = raw_df[(raw_df["point_time"] >= weekly_start_date) & (raw_df["point_time"] <= weekly_end_date)].copy()

        baseline_empty = baseline_data.empty
        if weekly_data.empty:
            return None

        def calc_stats(df: pd.DataFrame) -> Dict[str, Any]:
            sigma = df["point_val"].std() if df.shape[0] > 1 else 0.0
            if np.isnan(sigma): sigma = 0.0
            return {
                "values": df["point_val"].values,
                "cnt": df.shape[0],
                "mean": df["point_val"].mean(),
                "sigma": sigma,
            }

        base_data_dict = calc_stats(baseline_data) if not baseline_empty else None
        weekly_data_dict = calc_stats(weekly_data)

        result: Dict[str, Any] = {
            "data_cnt": weekly_data_dict["cnt"],
            "ooc_cnt": 0,
            "WE_Rule": "", "OOB_Rule": "",
            "Material_no": chart_info.get("material_no", "N/A"),
            "group_name": str(chart_info.get("group_name", "N/A")),
            "chart_name": str(chart_info.get("chart_name", "N/A")),
            "chart_ID": chart_info.get("ChartID", "N/A"),
            "Characteristics": chart_info.get("Characteristics", "N/A"),
            "USL": chart_info.get("UCL", np.nan),
            "LSL": chart_info.get("LCL", np.nan),
            "UCL": chart_info.get("UCL", np.nan),
            "LCL": chart_info.get("LCL", np.nan),
            "Target": chart_info.get("Target", np.nan),
            "Resolution": chart_info.get("Resolution", np.nan),
            "baseline_insufficient": baseline_insufficient,
            "baseline_empty": baseline_empty,
            "data_type": "discrete",
        }

        if not baseline_insufficient and not baseline_empty:
            weekly_df = pd.DataFrame({"point_val": weekly_data["point_val"]})
            ooc_results = ooc_calculator(weekly_df, chart_info.get("UCL"), chart_info.get("LCL"))
            ooc_highlight = review_ooc_results(ooc_results[1], ooc_results[2])
            result["ooc_cnt"] = ooc_results[1]

            discrete_oob_result = discrete_oob_calculator(
                base_data_dict, weekly_data_dict, chart_info, raw_df,
                weekly_start_date, weekly_end_date, actual_baseline_start, baseline_end_date,
            )
            record_results = record_high_low_calculator(weekly_data["point_val"].values, baseline_data["point_val"].values)

            result.update({
                "HL_P95_shift": discrete_oob_result.get("HL_P95_shift", "NO_HIGHLIGHT"),
                "HL_P50_shift": discrete_oob_result.get("HL_P50_shift", "NO_HIGHLIGHT"),
                "HL_P05_shift": discrete_oob_result.get("HL_P05_shift", "NO_HIGHLIGHT"),
                "HL_sticking_shift": discrete_oob_result.get("HL_sticking_shift", "NO_HIGHLIGHT"),
                "HL_trending": discrete_oob_result.get("HL_trending", "NO_HIGHLIGHT"),
                "HL_high_OOC": ooc_highlight,
                "HL_category_LT_shift": discrete_oob_result.get("HL_category_LT_shift", "NO_HIGHLIGHT"),
                "HL_record_high_low": record_results.get("highlight_status", "NO_HIGHLIGHT"),
                "record_high": record_results.get("record_high", False),
                "record_low": record_results.get("record_low", False),
            })
        else:
            result.update({
                "HL_P95_shift": "NO_HIGHLIGHT", "HL_P50_shift": "NO_HIGHLIGHT", "HL_P05_shift": "NO_HIGHLIGHT",
                "HL_sticking_shift": "NO_HIGHLIGHT", "HL_trending": "NO_HIGHLIGHT", "HL_high_OOC": "NO_HIGHLIGHT",
                "HL_category_LT_shift": "NO_HIGHLIGHT", "HL_record_high_low": "NO_HIGHLIGHT",
                "record_high": False, "record_low": False,
            })

        return result
    except Exception as e:
        print(f"[Error] Discrete process failed: {e}")
        return None

def _analyze_chart_api(execution_time: Optional[pd.Timestamp], raw_df: pd.DataFrame, chart_info: Dict[str, Any], output_dir: str = 'output', plot_df: Optional[pd.DataFrame] = None) -> Optional[Dict[str, Any]]:
    if "rule_list" not in chart_info or not chart_info.get("rule_list"):
        chart_info["rule_list"] = [rule for rule in ["WE1", "WE2", "WE3", "WE4", "WE5", "WE6", "WE7", "WE8", "WE9", "WE10"] if chart_info.get(rule, "N") == "Y"]

    if "point_time" not in raw_df.columns or not pd.api.types.is_datetime64_any_dtype(raw_df["point_time"]):
        return None

    # OOS 顯示：繪圖用含 OOS 的完整資料，分析計算仍用過濾後的 raw_df
    _plot_df = plot_df if (plot_df is not None and not plot_df.empty) else raw_df
    latest_raw_time = _plot_df["point_time"].max()
    weekly_end_date = latest_raw_time if execution_time is None or pd.isna(execution_time) else execution_time
    if pd.isna(weekly_end_date):
        return None
    weekly_start_date = weekly_end_date - pd.Timedelta(days=7)
    baseline_end_date = weekly_start_date - pd.Timedelta(seconds=1)
    initial_baseline_start_date = baseline_end_date - pd.Timedelta(days=365)

    data_type = determine_data_type(raw_df["point_val"].dropna()) if not raw_df.empty and "point_val" in raw_df.columns else "continuous"
    chart_info["data_type"] = data_type

    if data_type == "discrete":
        result = _process_discrete_chart_api(raw_df, chart_info, weekly_start_date, weekly_end_date, initial_baseline_start_date, baseline_end_date)
    else:
        result = process_single_chart(chart_info.copy(), raw_df, initial_baseline_start_date, baseline_end_date, weekly_start_date, weekly_end_date)
        if result:
            result["data_type"] = "continuous"

    # Fallback：若指定的週窗口內無點（result is None），改以資料實際最新時間點為窗口尾端重試
    # 確保只要 CSV 有點數，就一定能畫出圖
    if result is None and not raw_df.empty and "point_val" in raw_df.columns:
        fallback_end = latest_raw_time
        if not pd.isna(fallback_end) and fallback_end != weekly_end_date:
            print(f"[Fallback] 指定週窗口無資料，改以資料最新時間 {fallback_end} 為窗口尾端重新分析")
            fallback_start = fallback_end - pd.Timedelta(days=7)
            fallback_baseline_end = fallback_start - pd.Timedelta(seconds=1)
            fallback_baseline_start = fallback_baseline_end - pd.Timedelta(days=365)
            if data_type == "discrete":
                result = _process_discrete_chart_api(raw_df, chart_info, fallback_start, fallback_end, fallback_baseline_start, fallback_baseline_end)
            else:
                result = process_single_chart(chart_info.copy(), raw_df, fallback_baseline_start, fallback_baseline_end, fallback_start, fallback_end)
                if result:
                    result["data_type"] = "continuous"
            if result is not None:
                result["no_data_reason"] = "weekly_window_fallback"
                weekly_start_date = fallback_start
                weekly_end_date = fallback_end
                baseline_end_date = fallback_baseline_end
                initial_baseline_start_date = fallback_baseline_start

    if result is None:
        return None

    image_path, violated_rules = plot_spc_chart(_plot_df, chart_info, weekly_start_date, weekly_end_date, output_dir=output_dir)
    weekly_image_path = plot_weekly_spc_chart(_plot_df, chart_info, weekly_start_date, weekly_end_date, output_dir=output_dir)

    # --- 修正後的機台偵測邏輯（使用含 OOS 的 _plot_df）---
    # 定義所有可能的機台欄位名稱
    possible_tool_cols = ["ByTool", "EQP_id", "Matching", "Tool", "tool_id"]
    target_tool_col = next((c for c in possible_tool_cols if c in _plot_df.columns), None)
    
    _has_tool_data = False
    if target_tool_col:
        # 檢查是否有超過一個以上的機台，且排除空值
        valid_series = _plot_df[target_tool_col].dropna().astype(str).str.strip()
        valid_series = valid_series[valid_series != ""]
        if valid_series.nunique() > 1:
            _has_tool_data = True

    if _has_tool_data:
        # 注意：這裡要把 _plot_df 裡的欄位暫時 rename 給 oob_eng 的繪圖函式用
        temp_df = _plot_df.rename(columns={target_tool_col: "ByTool"})
        try: result["by_tool_color_path"] = plot_spc_by_tool_color(temp_df, chart_info, weekly_start_date, weekly_end_date, output_dir=output_dir)
        except Exception: result["by_tool_color_path"] = None
        try: result["by_tool_group_path"] = plot_spc_by_tool_group(temp_df, chart_info, output_dir=output_dir)
        except Exception: result["by_tool_group_path"] = None
    else:
        result["by_tool_color_path"] = None
        result["by_tool_group_path"] = None

    weekly_data = raw_df[(raw_df["point_time"] >= weekly_start_date) & (raw_df["point_time"] <= weekly_end_date)].copy()
    cpk = calculate_cpk(weekly_data, chart_info)
    result["Cpk"] = cpk.get("Cpk", np.nan) if cpk else np.nan

    try:
        qq_plot_path = plot_qq_plot(_plot_df, chart_info, output_dir=output_dir)
    except Exception:
        qq_plot_path = None

    # 序列化 chart_data 供 UI Plotly hover 使用 (All Data SPC，含 OOS 點)
    result["weekly_start"] = str(weekly_start_date)
    result["weekly_end"]   = str(weekly_end_date)
    try:
        _possible_site_cols = ["Matching", "ByTool", "EQP_id", "Tool", "tool_id", "Vendor_Site"]
        _site_col = next((c for c in _possible_site_cols if c in _plot_df.columns), None)
        _cd_cols = ["point_time", "point_val"]
        if _site_col:
            _cd_cols.append(_site_col)
        _chart_data_df = _plot_df[_cd_cols].copy()
        _chart_data_df["point_time"] = _chart_data_df["point_time"].astype(str)
        _chart_data_df["point_val"] = pd.to_numeric(_chart_data_df["point_val"], errors="coerce")
        _chart_data_df = _chart_data_df.dropna(subset=["point_val"])
        if _site_col:
            _chart_data_df = _chart_data_df.rename(columns={_site_col: "Matching"})
            _chart_data_df["Matching"] = _chart_data_df["Matching"].fillna("Unknown").astype(str)
        # 標記 OOS 點供前端顯示（point_val 在 OOS 過濾後的 raw_df 中不存在）
        if plot_df is not None and not plot_df.empty:
            filtered_times = set(raw_df["point_time"].astype(str))
            _chart_data_df["is_oos"] = ~_chart_data_df["point_time"].isin(filtered_times)
        else:
            _chart_data_df["is_oos"] = False
        # 確保所有值都是 Python 原生型別（避免 numpy 型別造成 JSON 序列化問題）
        result["chart_data"] = [
            {k: (float(v) if hasattr(v, "item") else (bool(v) if isinstance(v, (bool, np.bool_)) else v)) for k, v in row.items()}
            for row in _chart_data_df.to_dict(orient="records")
        ]
    except Exception:
        result["chart_data"] = []

    # 計算 weekly OOS 計數（weekly 窗口內出現於完整資料但被 OOS 過濾掉的點）
    try:
        if plot_df is not None and not plot_df.empty:
            _weekly_all = _plot_df[
                (_plot_df["point_time"] >= weekly_start_date) & (_plot_df["point_time"] <= weekly_end_date)
            ]
            _weekly_raw_times = set(
                raw_df[
                    (raw_df["point_time"] >= weekly_start_date) & (raw_df["point_time"] <= weekly_end_date)
                ]["point_time"]
            )
            result["oos_cnt"] = int((~_weekly_all["point_time"].isin(_weekly_raw_times)).sum())
        else:
            result["oos_cnt"] = 0
    except Exception:
        result["oos_cnt"] = 0

    return _build_result_api(result, violated_rules, image_path, weekly_image_path, qq_plot_path)


# ==========================================
# Tool Matching 分析模組 (Tool Matching Logic)
# ==========================================
def _create_spc_chart(group_df: pd.DataFrame, group_name: str, chart_name: str, focus_group=None):
    from io import BytesIO
    unique_groups = sorted(group_df["matching_group"].unique(), key=lambda x: str(x))

    if group_df.empty or not any(len(grp["point_val"]) > 0 for _, grp in group_df.groupby("matching_group")):
        return None

    has_time = 'point_time' in group_df.columns
    fig, ax = plt.subplots(figsize=(8, 4))
    try:
        colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))
        if has_time:
            group_df = group_df.copy()
            group_df['point_time'] = pd.to_datetime(group_df['point_time'], errors='coerce')

        x_position = 0
        tick_positions = []
        tick_labels = []

        for i, mg in enumerate(unique_groups):
            group_data = group_df[group_df["matching_group"] == mg]
            if has_time:
                group_data = group_data.sort_values("point_time")

            if not group_data.empty:
                x_vals = np.arange(x_position, x_position + len(group_data))
                y_vals = group_data["point_val"].values
                is_focus = focus_group is None or str(mg) == str(focus_group)
                pt_color = colors[i]
                pt_alpha = 0.9 if is_focus else 0.25
                ln_alpha = 0.5 if is_focus else 0.12
                ax.scatter(x_vals, y_vals, color=pt_color, alpha=pt_alpha, s=40, label=f'{mg}', zorder=3)
                ax.plot(x_vals, y_vals, color=pt_color, alpha=ln_alpha, linewidth=1, zorder=2)

                if i < len(unique_groups) - 1:
                    separator_x = x_position + len(group_data) - 0.5
                    ax.axvline(x=separator_x, color='gray', linestyle='-', alpha=0.3, zorder=1)

                # x 軸標籤：有時間則每個資料點顯示日期，否則顯示組名置中
                if has_time and group_data['point_time'].notna().any():
                    times = group_data['point_time'].reset_index(drop=True)
                    for xi, t in zip(x_vals, times):
                        if pd.notna(t):
                            tick_positions.append(xi)
                            tick_labels.append(pd.Timestamp(t).strftime('%Y-%m-%d'))
                else:
                    center = x_position + len(group_data) / 2 - 0.5
                    tick_positions.append(center)
                    tick_labels.append(str(mg))

                x_position += len(group_data)

        ax.set_title(f"SPC Chart: {group_name} - {chart_name}" + (f"  [Focus: {focus_group}]" if focus_group else ""), fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.3, zorder=0)
        # 最多顯示 15 個 tick，均勻取樣
        max_ticks = 15
        if len(tick_positions) > max_ticks:
            step = len(tick_positions) / max_ticks
            indices = [int(i * step) for i in range(max_ticks)]
            tick_positions = [tick_positions[i] for i in indices]
            tick_labels = [tick_labels[i] for i in indices]
        ax.set_xticks(tick_positions)
        ax.set_xticklabels(tick_labels, rotation=90, ha='center', fontsize=8)
        ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
        plt.tight_layout()
        buf = BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        return buf.read()
    finally:
        plt.close(fig)


def _create_timeline_chart(group_df: pd.DataFrame, group_name: str, chart_name: str, focus_group=None):
    """所有 matching_group 混合，依 point_time 排序，各組不同顏色的時序圖。x 軸用等距索引，避免時間疏密不均。"""
    from io import BytesIO
    import matplotlib.dates as mdates
    if group_df.empty or 'point_time' not in group_df.columns:
        return None

    fig, ax = plt.subplots(figsize=(9, 4.5))
    try:
        df_sorted = group_df.copy()
        df_sorted['point_time'] = pd.to_datetime(df_sorted['point_time'], errors='coerce')
        df_sorted = df_sorted.dropna(subset=['point_time']).sort_values('point_time').reset_index(drop=True)
        if df_sorted.empty:
            return None

        # 等距 x 軸（索引），避免時間間距不一造成視覺誤差
        x_vals = np.arange(len(df_sorted))
        unique_groups = sorted(group_df["matching_group"].unique(), key=lambda x: str(x))
        colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))
        color_map = {mg: colors[i] for i, mg in enumerate(unique_groups)}

        # 連接線（灰色底線）
        ax.plot(x_vals, df_sorted['point_val'].values, color='lightgray', linewidth=1, zorder=1)

        for mg in unique_groups:
            idx = df_sorted[df_sorted['matching_group'] == mg].index
            if len(idx) == 0:
                continue
            is_focus = focus_group is None or str(mg) == str(focus_group)
            sc_color = color_map[mg]
            sc_alpha = 0.9 if is_focus else 0.25
            ax.scatter(idx, df_sorted.loc[idx, 'point_val'].values,
                       color=sc_color, alpha=sc_alpha, s=40, label=str(mg), zorder=3)

        # x 軸：等距 tick，顯示日期標籤
        total = len(df_sorted)
        max_ticks = 15
        if total <= max_ticks:
            tick_idx = list(range(total))
        else:
            step = max(1, total // max_ticks)
            tick_idx = list(range(0, total, step))
            if tick_idx[-1] != total - 1:
                tick_idx.append(total - 1)
        ax.set_xticks(tick_idx)
        ax.set_xticklabels(
            [df_sorted['point_time'].iloc[i].strftime('%Y-%m-%d') for i in tick_idx],
            rotation=90, ha='center', fontsize=8
        )

        ax.set_title(f"Timeline: {group_name} - {chart_name}" + (f"  [Focus: {focus_group}]" if focus_group else ""), fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.3, zorder=0)
        ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
        plt.tight_layout()
        buf = BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        return buf.read()
    finally:
        plt.close(fig)

def _create_boxplot_chart(group_df: pd.DataFrame, group_name: str, chart_name: str, focus_group=None):
    from io import BytesIO
    fig, ax = plt.subplots(figsize=(9, 4.5))
    try:
        unique_groups = sorted(group_df["matching_group"].unique(), key=lambda x: str(x))
        labels = [str(mg) for mg in unique_groups]

        box_data = [group_df[group_df["matching_group"] == mg]["point_val"].values for mg in unique_groups]
        group_stats = group_df.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count'])
        colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))

        if box_data:
            bp = ax.boxplot(box_data, labels=labels, patch_artist=True, widths=0.6)
            for patch, color, mg in zip(bp['boxes'], colors, unique_groups):
                is_focus = focus_group is None or str(mg) == str(focus_group)
                patch.set_facecolor(color)
                patch.set_alpha(0.95 if is_focus else 0.25)

            legend_labels = [
                f"{label}: μ={group_stats.loc[mg, 'mean']:.2f}, σ={group_stats.loc[mg, 'std']:.2f}, n={int(group_stats.loc[mg, 'count'])}"
                for label, mg in zip(labels, unique_groups)
            ]
            ax.legend([bp["boxes"][i] for i in range(len(labels))], legend_labels, loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')

        ax.set_title(f"Boxplot: {group_name} - {chart_name}" + (f"  [Focus: {focus_group}]" if focus_group else ""), fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.6)
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=90, ha='center')
        fig.subplots_adjust(right=0.7)
        plt.tight_layout()
        buf = BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        return buf.read()
    finally:
        plt.close(fig)

def _export_tool_matching_to_excel(results_df: pd.DataFrame, chart_bytes: dict, source_path: str) -> str:
    try:
        excel_data = []
        for _, row in results_df.iterrows():
            mean_index = row['mean_index']
            sigma_index = row['sigma_index']
            k_value = row['k_value']
            need_matching = False
            abnormal_type = ""
            
            if isinstance(mean_index, (int, float)) and isinstance(sigma_index, (int, float)) and isinstance(k_value, (int, float)):
                if abs(mean_index) > 2.0: 
                    need_matching = True
                    abnormal_type = "Mean"
                elif sigma_index > k_value: 
                    need_matching = True
                    abnormal_type = "Sigma"
            elif "Significant" in str(mean_index) and "No Significant" not in str(mean_index):
                need_matching = True
                abnormal_type = "Mean"
            
            def format_value(val):
                if pd.isna(val): return ""
                if isinstance(val, float) and (val == float('inf') or val == float('-inf')):
                    return "inf" if val == float('inf') else "-inf"
                return val
            
            excel_data.append({
                'SPC_Chart': "", 
                'BoxPlot': "",
                'Timeline_Chart': "",
                'Need_matching': need_matching,
                'AbnormalType': abnormal_type,
                'GroupName': str(row['gname']),
                'ChartName': str(row['cname']),
                'matching_group': str(row['group']),
                'mean_matching_index': format_value(row['mean_index']),
                'sigma_matching_index': format_value(row['sigma_index']),
                'K': format_value(row['k_value']),
                'mean': format_value(row['mean']),
                'sigma': format_value(row['std']), 
                'mean_median': format_value(row['mean_median']),
                'sigma_median': format_value(row['sigma_median']),
                'samplesize': int(row['n']) if pd.notna(row['n']) else 0,
                'characteristic': str(row['characteristic'])
            })
        
        df = pd.DataFrame(excel_data)
        temp_dir = "temp_uploads"
        os.makedirs(temp_dir, exist_ok=True)
        file_name = os.path.splitext(os.path.basename(source_path))[0]
        output_path = os.path.join(temp_dir, f"{file_name}_matching_results.xlsx")
        img_temp_dir = tempfile.mkdtemp()
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Tool Matching Results', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Tool Matching Results']
                
                header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                header_fill = openpyxl.styles.PatternFill(start_color="344CB7", end_color="344CB7", fill_type="solid")
                header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                worksheet.column_dimensions['A'].width = 70 
                worksheet.column_dimensions['B'].width = 70 
                worksheet.column_dimensions['C'].width = 70 
                abnormal_fill = openpyxl.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                img_display_width, img_display_height = 450, 250
                
                for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
                    if row_data["Need_matching"]:
                        for cell in worksheet[row_idx]:
                            cell.fill = abnormal_fill
                    
                    group_name = str(row_data["GroupName"])
                    chart_name = str(row_data["ChartName"])
                    chart_key = (group_name, chart_name, str(row_data["matching_group"]))
                    
                    if chart_key in chart_bytes:
                        try:
                            scatter_b = chart_bytes[chart_key].get('scatter')
                            if scatter_b:
                                temp_scatter_path = os.path.join(img_temp_dir, f"spc_{group_name}_{chart_name}_{row_idx}.png")
                                with open(temp_scatter_path, 'wb') as _f: _f.write(scatter_b)
                                scatter_img = XLImage(temp_scatter_path)
                                scatter_img.width = img_display_width
                                scatter_img.height = img_display_height
                                worksheet.add_image(scatter_img, f"A{row_idx}")
                            box_b = chart_bytes[chart_key].get('box')
                            if box_b:
                                temp_box_path = os.path.join(img_temp_dir, f"box_{group_name}_{chart_name}_{row_idx}.png")
                                with open(temp_box_path, 'wb') as _f: _f.write(box_b)
                                box_img = XLImage(temp_box_path)
                                box_img.width = img_display_width
                                box_img.height = img_display_height
                                worksheet.add_image(box_img, f"B{row_idx}")
                            timeline_b = chart_bytes[chart_key].get('timeline')
                            if timeline_b:
                                temp_tl_path = os.path.join(img_temp_dir, f"tl_{group_name}_{chart_name}_{row_idx}.png")
                                with open(temp_tl_path, 'wb') as _f: _f.write(timeline_b)
                                tl_img = XLImage(temp_tl_path)
                                tl_img.width = img_display_width
                                tl_img.height = img_display_height
                                worksheet.add_image(tl_img, f"C{row_idx}")
                        except Exception:
                            worksheet.cell(row=row_idx, column=1).value = "Chart failed to load"
                            worksheet.cell(row=row_idx, column=2).value = "Chart failed to load"
                            worksheet.cell(row=row_idx, column=3).value = "Chart failed to load"
                
                for row in range(2, len(df) + 2):
                    worksheet.row_dimensions[row].height = img_display_height * 0.75
        finally:
            shutil.rmtree(img_temp_dir, ignore_errors=True)
        return output_path
    except Exception as e:
        print(f"Excel export failed: {e}")
        return None

def _analyze_tool_matching_with_charts_and_excel(all_charts_info: pd.DataFrame, raw_data_directory: str, config: dict, source_path: str, output_dir: str = "output") -> dict:
    basic_result = analyze_tool_matching_data(all_charts_info, raw_data_directory, config)
    chart_bytes = {}
    chart_paths = {}  # (gname, cname) -> {'spc': path, 'box': path}

    try:
        from oob_eng import find_matching_file as _find_file
        import concurrent.futures

        # 第一階段：蒐集可產圖的任務參數
        task_args = []
        task_meta = {}  # (gname, cname) -> safe_g, safe_c

        for _, chart_row in all_charts_info.iterrows():
            gname = str(chart_row.get('GroupName', ''))
            cname = str(chart_row.get('ChartName', ''))
            try:
                csv_path = _find_file(raw_data_directory, gname, cname)
                if not csv_path or not os.path.isfile(csv_path):
                    continue
                df_chart = pd.read_csv(csv_path)
                for col in ['ByTool', 'EQP_id', 'Matching', 'Tool', 'tool_id']:
                    if col in df_chart.columns and 'matching_group' not in df_chart.columns:
                        df_chart = df_chart.rename(columns={col: 'matching_group'})
                if 'matching_group' not in df_chart.columns:
                    continue
                safe_g = "".join(c if c.isalnum() or c in "-_" else "_" for c in gname)
                safe_c = "".join(c if c.isalnum() or c in "-_" else "_" for c in cname)
                task_args.append((gname, cname, df_chart.to_dict('records')))
                task_meta[(gname, cname)] = (safe_g, safe_c)
            except Exception:
                continue

        # 第二階段：產圖（序列執行，已停用平行運算）
        if task_args:
            # with concurrent.futures.ProcessPoolExecutor(max_workers=4) as executor:
            #     for result in executor.map(_tool_matching_plot_worker, task_args):
            for result in map(_tool_matching_plot_worker, task_args):
                    try:
                        gname, cname, per_site = result
                        safe_g, safe_c = task_meta.get((gname, cname), (gname, cname))
                        for site, (scatter_b, box_b, timeline_b) in per_site.items():
                            safe_site = "".join(c if c.isalnum() or c in "-_" else "_" for c in site)
                            paths = {}
                            if scatter_b:
                                spc_path = os.path.join(output_dir, f"TM_{safe_g}_{safe_c}_{safe_site}_spc.png")
                                with open(spc_path, 'wb') as f:
                                    f.write(scatter_b)
                                paths['spc'] = os.path.abspath(spc_path)
                                chart_bytes.setdefault((gname, cname, site), {})['scatter'] = scatter_b
                            if box_b:
                                box_path = os.path.join(output_dir, f"TM_{safe_g}_{safe_c}_{safe_site}_box.png")
                                with open(box_path, 'wb') as f:
                                    f.write(box_b)
                                paths['box'] = os.path.abspath(box_path)
                                chart_bytes.setdefault((gname, cname, site), {})['box'] = box_b
                            if timeline_b:
                                tl_path = os.path.join(output_dir, f"TM_{safe_g}_{safe_c}_{safe_site}_timeline.png")
                                with open(tl_path, 'wb') as f:
                                    f.write(timeline_b)
                                paths['timeline'] = os.path.abspath(tl_path)
                                chart_bytes.setdefault((gname, cname, site), {})['timeline'] = timeline_b
                            if paths:
                                chart_paths[(gname, cname, site)] = paths
                    except Exception:
                        continue

    except Exception:
        pass

    excel_output = None
    try:
        excel_output = _export_tool_matching_to_excel(basic_result["results"], chart_bytes, source_path)
    except Exception:
        pass

    return {
        "summary": basic_result["summary"],
        "results": basic_result["results"],
        "excel_output": excel_output,
        "chart_paths": chart_paths,
    }


# ==========================================
# FastAPI 路由與端點 (Routes)
# ==========================================

def _watchdog_process(p: Process, task_id: str, timeout_sec: int = _TASK_PROCESSING_TIMEOUT_HOURS * 3600) -> None:
    """Daemon thread：監控子進程，若超時則強制 kill 並標記任務為 failed。"""
    p.join(timeout=timeout_sec)
    if p.is_alive():
        p.kill()
        p.join()
        try:
            update_task_status(task_id, {
                "status": "failed",
                "error": f"Task killed: timed out after {timeout_sec // 3600}h",
                "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat(),
            }, db=task_status_db)
        except Exception as e:
            print(f"[Watchdog] 無法更新任務狀態 (Manager 可能已關閉): {e}")
        print(f"[Watchdog] Task {task_id} killed after {timeout_sec}s timeout")


@app.post("/split")
def split_csvs(req: SplitRequest) -> Dict[str, Any]:
    split_id = str(uuid.uuid4())
    final_output_folder = os.path.join("input", "raw_charts", split_id)
    os.makedirs(final_output_folder, exist_ok=True)

    successes = 0
    failures: List[str] = []
    for path in req.input_files:
        ok = False
        if req.mode == "Type3_Horizontal": ok = _split_type3_horizontal(path, final_output_folder)
        elif req.mode == "Type2_Vertical": ok = _split_type2_vertical(path, final_output_folder)
        elif req.mode == "Vendor_Vertical": ok = _split_vendor_vertical(path, final_output_folder)
        elif req.mode == "Test_Horizontal": ok = _split_test_horizontal(path, final_output_folder)
        if ok: successes += 1
        else: failures.append(os.path.basename(path))

    return {
        "mode": req.mode,
        "split_id": split_id,
        "raw_data_directory": os.path.abspath(final_output_folder),
        "processed": successes,
        "failed": failures,
    }

@app.get("/split-status")
def get_split_status(split_id: Optional[str] = None) -> Dict[str, Any]:
    if not split_id:
        return {"has_split_data": False, "split_folder": None, "csv_file_count": 0, "folder_exists": False, "split_id": None}
    folder = os.path.abspath(os.path.join("input", "raw_charts", split_id))
    if os.path.exists(folder):
        try:
            file_count = len([f for f in os.listdir(folder) if f.endswith(".csv")])
        except Exception:
            file_count = 0
        return {"has_split_data": True, "split_folder": folder, "csv_file_count": file_count, "folder_exists": True, "split_id": split_id}
    return {"has_split_data": False, "split_folder": None, "csv_file_count": 0, "folder_exists": False, "split_id": split_id}

# ==========================================
# /process 背景任務函式
# ==========================================
def _tool_matching_plot_worker(args):
    """頂層 picklable worker：每個 site 各自高亮，生成 per-site 的 scatter + boxplot + timeline PNG bytes。"""
    group_name, chart_name, group_df_records = args
    try:
        import pandas as pd
        group_df = pd.DataFrame(group_df_records)
        unique_sites = sorted(group_df['matching_group'].dropna().astype(str).unique(), key=str)
        per_site = {}
        for site in unique_sites:
            per_site[site] = (
                _create_spc_chart(group_df, group_name, chart_name, focus_group=site),
                _create_boxplot_chart(group_df, group_name, chart_name, focus_group=site),
                _create_timeline_chart(group_df, group_name, chart_name, focus_group=site),
            )
        return (group_name, chart_name, per_site)
    except Exception:
        return (group_name, chart_name, {})


def _spc_cpk_worker(args):
    """頂層 picklable worker：CPK 計算 + SPC 圖表生成（供 ProcessPoolExecutor 使用）。"""
    chart_info_dict, raw_data_directory, start_date, end_date, custom_mode = args
    try:
        import pandas as pd
        import math
        import cpk_eng
        group_name = str(chart_info_dict.get('GroupName', ''))
        chart_name = str(chart_info_dict.get('ChartName', ''))
        raw_path = find_matching_file(raw_data_directory, group_name, chart_name)
        if not raw_path or not os.path.exists(raw_path):
            return None
        raw_df = pd.read_csv(raw_path)
        usl, lsl = chart_info_dict.get('USL'), chart_info_dict.get('LSL')
        if usl is not None and lsl is not None:
            raw_df = raw_df[(raw_df['point_val'] <= usl) & (raw_df['point_val'] >= lsl)]
        elif usl is not None:
            raw_df = raw_df[raw_df['point_val'] <= usl]
        elif lsl is not None:
            raw_df = raw_df[raw_df['point_val'] >= lsl]
        if custom_mode and start_date and end_date:
            cpk_res = cpk_eng._compute_cpk_custom_range(raw_df, chart_info_dict, pd.to_datetime(start_date), pd.to_datetime(end_date))
            all_data_cpk = cpk_eng.calculate_cpk_dashboard(raw_df, chart_info_dict)['Cpk']
            metrics = {'cpk': cpk_res.get('Cpk'), 'cpk_l1': None, 'cpk_l2': None, 'custom_cpk': all_data_cpk,
                       'r1': None, 'r2': None, 'k_value': cpk_eng._calculate_k_value(raw_df, chart_info_dict, start_date, end_date, custom_mode)}
        else:
            end_time = pd.to_datetime(end_date)
            if 'point_time' in raw_df.columns:
                latest = pd.to_datetime(raw_df['point_time']).max()
                if end_time > latest:
                    end_time = latest
            cpk_res = cpk_eng.compute_cpk_windows(raw_df, chart_info_dict, end_time)
            all_data_cpk = cpk_eng.calculate_cpk_dashboard(raw_df, chart_info_dict)['Cpk']
            cpk, l1, l2 = cpk_res.get('Cpk'), cpk_res.get('Cpk_last_month'), cpk_res.get('Cpk_last2_month')
            r1 = r2 = None
            if cpk is not None and l1 is not None and l1 != 0 and cpk <= l1:
                r1 = (1 - (cpk / l1)) * 100
            if cpk is not None and l1 is not None and l2 is not None and l2 != 0 and cpk <= l1 <= l2:
                r2 = (1 - (cpk / l2)) * 100
            metrics = {'cpk': cpk, 'cpk_l1': l1, 'cpk_l2': l2, 'custom_cpk': all_data_cpk,
                       'r1': r1, 'r2': r2, 'k_value': cpk_eng._calculate_k_value(raw_df, chart_info_dict, start_date, end_date, custom_mode)}
        mean_stats = cpk_eng._calculate_period_statistics(raw_df, end_date, custom_mode, start_date)
        for k in ['mean_current', 'sigma_current', 'mean_last_month', 'sigma_last_month', 'mean_last2_month', 'sigma_last2_month']:
            if mean_stats.get(k) is None and cpk_res.get(k) is not None:
                mean_stats[k] = cpk_res[k]
        # 計算 violation 標記供圖片 title 使用
        _r1v = metrics.get('r1'); _r2v = metrics.get('r2')
        def _h(v, thr):
            try: return float(v) >= thr if v is not None else False
            except: return False
        _viol = ("H(R1+R2)" if _h(_r1v, 25) and _h(_r2v, 20) else
                 "H(R1)"    if _h(_r1v, 25) else
                 "H(R2)"    if _h(_r2v, 20) else "")
        metrics_for_title = {**metrics, 'violation': _viol}
        chart_image = cpk_eng.generate_spc_chart_base64(raw_df, chart_info_dict, start_date, end_date, custom_mode, metrics=metrics_for_title)
        def san(x): return None if isinstance(x, float) and (math.isnan(x) or math.isinf(x)) else x
        return {
            'group_name': group_name, 'chart_name': chart_name,
            'characteristics': str(chart_info_dict.get('Characteristics', '')),
            'usl': usl, 'lsl': lsl, 'target': cpk_eng._get_target_value(chart_info_dict),
            'metrics': {k: san(v) for k, v in metrics.items()},
            'chart_image': chart_image,
            'mean_current': san(mean_stats.get('mean_current')), 'sigma_current': san(mean_stats.get('sigma_current')),
            'mean_last_month': san(mean_stats.get('mean_last_month')), 'sigma_last_month': san(mean_stats.get('sigma_last_month')),
            'mean_last2_month': san(mean_stats.get('mean_last2_month')), 'sigma_last2_month': san(mean_stats.get('sigma_last2_month')),
            'mean_all': san(mean_stats.get('mean_all')), 'sigma_all': san(mean_stats.get('sigma_all')),
        }
    except Exception as e:
        print(f"[_spc_cpk_worker] Error {chart_info_dict.get('GroupName')}/{chart_info_dict.get('ChartName')}: {e}")
        return None


def _process_single_chart_worker(args):
    """
    頂層 picklable worker，供 ProcessPoolExecutor 使用。
    args: (chart_info_row_dict, exec_time, raw_dir, task_output_dir)
    回傳分析結果 dict，若失敗或跳過則回傳 None。
    """
    chart_info_row_dict, exec_time, raw_dir, task_output_dir = args
    try:
        import pandas as pd
        group_name = str(chart_info_row_dict.get("GroupName", chart_info_row_dict.get("group_name", "Unknown")))
        chart_name = str(chart_info_row_dict.get("ChartName", chart_info_row_dict.get("chart_name", "Unknown")))

        _minimal = {
            "group_name": group_name,
            "chart_name": chart_name,
            "data_cnt": 0,
            "WE_Rule": None,
            "OOB_Rule": None,
            "no_data": True,
        }

        csv_path = find_matching_file(raw_dir, group_name, chart_name)
        if not csv_path or not os.path.exists(csv_path):
            return {**_minimal, "no_data_reason": "csv_not_found"}

        try:
            raw_df = pd.read_csv(csv_path)
        except Exception:
            return {**_minimal, "no_data_reason": "csv_read_error"}

        if "point_time" in raw_df.columns:
            raw_df["point_time"] = pd.to_datetime(raw_df["point_time"], errors="coerce")
            raw_df.dropna(subset=["point_time"], inplace=True)

        if raw_df.empty:
            return {**_minimal, "no_data_reason": "empty"}

        chart_info_row = pd.Series(chart_info_row_dict)
        is_ok, processed_df, updated_chart_info, full_df = preprocess_data(chart_info_row, raw_df)
        if not is_ok or processed_df is None or processed_df.empty:
            return {**_minimal, "no_data_reason": "preprocess_failed"}

        chart_info = dict(chart_info_row_dict)
        if "Material_no" in chart_info: chart_info["material_no"] = chart_info.pop("Material_no")
        if "GroupName" in chart_info: chart_info["group_name"] = chart_info.pop("GroupName")
        if "ChartName" in chart_info: chart_info["chart_name"] = chart_info.pop("ChartName")
        chart_info.update(updated_chart_info.to_dict() if hasattr(updated_chart_info, "to_dict") else dict(updated_chart_info))

        return _analyze_chart_api(exec_time, processed_df, chart_info, output_dir=task_output_dir, plot_df=full_df)
    except Exception:
        return None


def _run_process_task(task_id: str, req: ProcessRequest, shared_db) -> None:
    """實際的 OOB/SPC 重運算邏輯，在獨立 Process 中執行。"""
    def _upd(updates): update_task_status(task_id, updates, shared_db)
    try:
        defaults = _default_paths()
        filepath = req.filepath or defaults["filepath"]

        if req.split_id:
            raw_dir = os.path.abspath(os.path.join("input", "raw_charts", req.split_id))
            if not os.path.isdir(raw_dir):
                _upd({"status": "failed", "error": f"Split folder not found for split_id: {req.split_id}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
                return
        else:
            raw_dir = req.raw_data_directory or defaults["raw_dir"]

        if not os.path.isfile(filepath):
            _upd({"status": "failed", "error": f"Excel file not found: {filepath}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
            return
        if not os.path.isdir(raw_dir):
            os.makedirs(raw_dir, exist_ok=True)

        try:
            all_charts_info = load_chart_information(filepath)
        except Exception as e:
            _upd({"status": "failed", "error": f"Failed to read chart info: {e}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
            return

        total_charts = len(all_charts_info)
        if req.limit_charts:
            all_charts_info = all_charts_info.head(req.limit_charts)
        if req.base_date:
            exec_time = pd.Timestamp(req.base_date) + pd.Timedelta(hours=23, minutes=59, seconds=59)
        else:
            exec_time = None
        _upd({"progress": 10})

        # 為此 task 建立獨立的圖片輸出資料夾，避免多工時路徑衝突
        task_output_dir = os.path.abspath(os.path.join("output", task_id))
        os.makedirs(task_output_dir, exist_ok=True)

        # 封裝每張圖的參數為可 pickle 的 tuple list
        _upd({"progress": 20})
        task_args = [
            (row.to_dict(), exec_time, raw_dir, task_output_dir)
            for _, row in all_charts_info.iterrows()
        ]

        # 序列執行（已停用平行運算）
        # import concurrent.futures
        # with concurrent.futures.ProcessPoolExecutor(max_workers=4) as executor:
        #     for result in executor.map(_process_single_chart_worker, task_args):
        #         processed_results.append(result)
        processed_results = []
        _total_tasks = len(task_args)
        for _i, result in enumerate(map(_process_single_chart_worker, task_args)):
            processed_results.append(result)
            _upd({"progress": 20 + int(70 * (_i + 1) / max(_total_tasks, 1))})

        results: List[Dict[str, Any]] = [r for r in processed_results if r is not None]
        skipped = len(processed_results) - len(results)

        processed = len(results)
        excel_output = None
        if req.save_excel and results:
            try:
                results_df = pd.DataFrame(results)
                expected_cols = [
                    "data_cnt", "ooc_cnt", "oos_cnt", "WE_Rule", "OOB_Rule", "data_type", "Material_no", "group_name",
                    "chart_name", "chart_ID", "Characteristics", "USL", "LSL", "UCL", "LCL", "Target",
                    "Cpk", "Resolution", "HL_record_high_low", "record_high", "record_low",
                    "chart_path", "weekly_chart_path", "qq_plot_path"
                ]
                for col in expected_cols:
                    if col not in results_df.columns:
                        results_df[col] = np.nan
                results_df = results_df[[c for c in expected_cols if c in results_df.columns]].replace([np.nan, np.inf, -np.inf], "N/A")
                excel_path = os.path.join(task_output_dir, "result_with_images.xlsx")
                save_results_to_excel(results_df, scale_factor=req.scale_factor, output_path=excel_path)
                excel_output = os.path.abspath(excel_path)
            except Exception:
                pass

        summary = ProcessSummary(
            total_charts=total_charts if not req.limit_charts else min(total_charts, req.limit_charts),
            processed_charts=processed,
            skipped_charts=skipped,
            excel_output=excel_output,
        )
        result_items = []
        for r in results:
            if "group_name" in r: r["group_name"] = str(r["group_name"])
            if "chart_name" in r: r["chart_name"] = str(r["chart_name"])
            if "Material_no" in r: r["Material_no"] = str(r["Material_no"]) if r["Material_no"] is not None else None
            if "chart_ID" in r: r["chart_ID"] = str(r["chart_ID"]) if r["chart_ID"] is not None else None
            if "Characteristics" in r: r["Characteristics"] = str(r["Characteristics"]) if r["Characteristics"] is not None else None
            result_items.append(ResultItem(**r))

        # 將完整結果序列化寫入 JSON 檔（不走 Manager.dict，避免大量資料反序列化耗 CPU）
        # Atomic write：先寫 .tmp 再 rename，防止 watchdog kill 到一半產生損壞 JSON
        _result_file = _result_json_path(task_id)
        os.makedirs(os.path.dirname(_result_file), exist_ok=True)
        _result_tmp = _result_file + ".tmp"
        with open(_result_tmp, "w", encoding="utf-8") as _f:
            json.dump(ProcessResponse(summary=summary, results=result_items).model_dump(), _f, ensure_ascii=False, default=str)
        os.replace(_result_tmp, _result_file)

        _upd({
            "status": "completed",
            "progress": 100,
            "excel_output": excel_output,
            "result_json_path": _result_file,
            "completed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat(),
        })

    except Exception as e:
        import traceback
        _upd({"status": "failed", "error": f"{e}\n{traceback.format_exc()}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
    finally:
        plt.close('all')
        gc.collect()


def _run_tool_matching_task(task_id: str, req: ToolMatchingRequest, shared_db) -> None:
    """Tool Matching 批次分析背景任務，在獨立 Process 中執行。"""
    def _upd(updates): update_task_status(task_id, updates, shared_db)
    try:
        defaults = _default_paths()
        chart_excel_path = req.chart_excel_path or defaults.get("chart_excel_path", "")
        raw_data_directory = req.raw_data_directory or defaults.get("raw_data_directory", "")
        _upd({"progress": 5})
        try:
            all_charts_info = load_chart_information(chart_excel_path)
        except Exception as e:
            _upd({"status": "failed", "error": f"Failed to load chart info: {e}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
            return
        if all_charts_info is None or all_charts_info.empty:
            _upd({"status": "failed", "error": "No chart info loaded", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
            return
        config = {
            "mean_index_threshold": req.mean_index_threshold, "sigma_index_threshold": req.sigma_index_threshold,
            "use_statistical_test": req.use_statistical_test, "statistical_method": req.statistical_method,
            "alpha_level": req.alpha_level, "fill_sample_size": req.fill_sample_size,
            "filter_mode": req.filter_mode, "base_date": str(req.base_date) if req.base_date else str(date.today()),
        }
        _upd({"progress": 20})
        task_output_dir = os.path.join("output", task_id)
        os.makedirs(task_output_dir, exist_ok=True)
        analysis_result = _analyze_tool_matching_with_charts_and_excel(all_charts_info, raw_data_directory, config, chart_excel_path, output_dir=task_output_dir)
        _upd({"progress": 90})
        chart_paths = analysis_result.get("chart_paths", {})
        result_items = []
        for _, row in analysis_result["results"].iterrows():
            try:
                def _sf(v): return 0.0 if pd.isna(v) else (999999.0 if isinstance(v, (int, float)) and v in [float('inf'), float('-inf')] else float(v)) if isinstance(v, (int, float)) else 0.0
                def _sv(v, d='N/A'): return d if pd.isna(v) else ('Infinite' if isinstance(v, float) and v == float('inf') else ('Negative Infinite' if isinstance(v, float) and v == float('-inf') else v))
                gname_val = str(row['gname'])
                cname_val = str(row['cname'])
                paths = chart_paths.get((gname_val, cname_val, str(row['group'])), {})
                # 判斷是否違反
                raw_mean = row['mean_index']
                raw_sigma = row['sigma_index']
                raw_k = row['k_value']
                need_matching = False
                abnormal_type = ""
                if isinstance(raw_mean, (int, float)) and isinstance(raw_sigma, (int, float)) and isinstance(raw_k, (int, float)):
                    if abs(raw_mean) > 2.0:
                        need_matching = True; abnormal_type = "Mean"
                    elif raw_sigma > raw_k:
                        need_matching = True; abnormal_type = "Sigma"
                elif isinstance(raw_mean, str) and "Significant" in raw_mean and "No Significant" not in raw_mean:
                    need_matching = True; abnormal_type = "Mean"
                result_items.append(ToolMatchingResultItem(
                    gname=gname_val, cname=cname_val, group=str(row['group']), group_all=str(row['group_all']),
                    mean_index=_sv(row['mean_index']), sigma_index=_sv(row['sigma_index']), k_value=_sv(row['k_value']),
                    mean=_sf(row['mean']), std=_sf(row['std']), mean_median=_sv(row['mean_median']),
                    sigma_median=_sv(row['sigma_median']), n=int(row['n']) if pd.notna(row['n']) and isinstance(row['n'], (int, float)) else 0,
                    characteristic=str(row['characteristic']),
                    need_matching=need_matching, abnormal_type=abnormal_type,
                    spc_chart_path=paths.get('spc'), boxplot_chart_path=paths.get('box'),
                    timeline_chart_path=paths.get('timeline'),
                ).model_dump())
            except Exception:
                continue
        _result_file = _result_json_path(task_id)
        os.makedirs(os.path.dirname(_result_file), exist_ok=True)
        _result_tmp = _result_file + ".tmp"
        with open(_result_tmp, "w", encoding="utf-8") as _f:
            json.dump({"summary": analysis_result["summary"], "results": result_items, "excel_output": analysis_result.get("excel_output")}, _f, ensure_ascii=False, default=str)
        os.replace(_result_tmp, _result_file)

        _upd({
            "status": "completed", "progress": 100,
            "result_json_path": _result_file,
            "completed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat(),
        })
    except Exception as e:
        import traceback
        _upd({"status": "failed", "error": f"{e}\n{traceback.format_exc()}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
    finally:
        plt.close('all')
        gc.collect()


def _run_spc_cpk_task(task_id: str, req: SPCCpkRequest, shared_db) -> None:
    """SPC CPK 分析背景任務，在獨立 Process 中執行。"""
    def _upd(updates): update_task_status(task_id, updates, shared_db)
    try:
        defaults = _default_paths()
        chart_excel_path = req.chart_excel_path or defaults["chart_excel_path"]
        raw_data_directory = req.raw_data_directory or defaults["raw_data_directory"]
        if not os.path.isfile(chart_excel_path):
            _upd({"status": "failed", "error": f"Chart info file not found: {chart_excel_path}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
            return
        _upd({"progress": 5})
        all_charts_info = load_chart_information(chart_excel_path)
        if all_charts_info is None or all_charts_info.empty:
            _upd({"status": "failed", "error": "No chart info loaded", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
            return
        end_date = req.end_date or date.today()
        start_date = req.start_date if req.start_date else (pd.Timestamp(end_date) - pd.DateOffset(months=3)).date()
        if req.selected_chart:
            all_charts_info = all_charts_info[
                all_charts_info.apply(lambda r: f"{r['GroupName']} - {r['ChartName']}" == req.selected_chart, axis=1)
            ]
        task_args = [
            (row.to_dict(), raw_data_directory, start_date, end_date, req.custom_mode)
            for _, row in all_charts_info.iterrows()
        ]
        _upd({"progress": 20})
        # import concurrent.futures
        # with concurrent.futures.ProcessPoolExecutor(max_workers=4) as executor:
        #     for result in executor.map(_spc_cpk_worker, task_args):
        #         raw_results.append(result)
        raw_results = []
        for result in map(_spc_cpk_worker, task_args):
            raw_results.append(result)
        _upd({"progress": 80})
        chart_result_objs = []
        for r in raw_results:
            if r is None:
                continue
            try:
                chart_result_objs.append(SPCChartInfo(
                    group_name=r['group_name'], chart_name=r['chart_name'], characteristics=r['characteristics'],
                    usl=r.get('usl'), lsl=r.get('lsl'), target=r.get('target'),
                    metrics=SPCCpkMetrics(**r['metrics']),
                    chart_image=r.get('chart_image'),
                    mean_current=r.get('mean_current'), sigma_current=r.get('sigma_current'),
                    mean_last_month=r.get('mean_last_month'), sigma_last_month=r.get('sigma_last_month'),
                    mean_last2_month=r.get('mean_last2_month'), sigma_last2_month=r.get('sigma_last2_month'),
                    mean_all=r.get('mean_all'), sigma_all=r.get('sigma_all'),
                ))
            except Exception:
                continue
        cpk_values = [c.metrics.cpk for c in chart_result_objs if c.metrics.cpk is not None]
        all_cpk_values = [c.metrics.custom_cpk for c in chart_result_objs if c.metrics.custom_cpk is not None]
        summary = {
            "total_charts": len(chart_result_objs), "charts_with_cpk": len(cpk_values),
            "avg_cpk": round(float(np.mean(cpk_values)), 4) if cpk_values else None,
            "median_cpk": round(float(np.median(cpk_values)), 4) if cpk_values else None,
            "avg_all_cpk": round(float(np.mean(all_cpk_values)), 4) if all_cpk_values else None,
            "median_all_cpk": round(float(np.median(all_cpk_values)), 4) if all_cpk_values else None,
            "charts_with_mean_current": len([c.mean_current for c in chart_result_objs if c.mean_current is not None]),
            "custom_mode": req.custom_mode, "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        _upd({"progress": 90})
        excel_path = cpk_eng._export_spc_cpk_to_excel(chart_result_objs, summary, start_date, end_date) if chart_result_objs else None
        _result_file = _result_json_path(task_id)
        os.makedirs(os.path.dirname(_result_file), exist_ok=True)
        _result_tmp = _result_file + ".tmp"
        with open(_result_tmp, "w", encoding="utf-8") as _f:
            json.dump(SPCCpkResponse(charts=chart_result_objs, summary=summary, excel_path=excel_path).model_dump(), _f, ensure_ascii=False, default=str)
        os.replace(_result_tmp, _result_file)

        _upd({
            "status": "completed", "progress": 100,
            "result_json_path": _result_file,
            "completed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat(),
        })
    except Exception as e:
        import traceback
        _upd({"status": "failed", "error": f"{e}\n{traceback.format_exc()}", "expires_at": (datetime.now() + timedelta(hours=_TASK_TTL_HOURS)).isoformat()})
    finally:
        plt.close('all')
        gc.collect()


@app.post("/process")
def process_charts_api(req: ProcessRequest) -> Dict[str, Any]:
    # 先做快速的前置驗證，避免把明顯錯誤丟到背景才發現
    defaults = _default_paths()
    filepath = req.filepath or defaults["filepath"]

    if req.split_id:
        raw_dir = os.path.abspath(os.path.join("input", "raw_charts", req.split_id))
        if not os.path.isdir(raw_dir):
            raise HTTPException(status_code=400, detail=f"Split folder not found for split_id: {req.split_id}")
    else:
        raw_dir = req.raw_data_directory or defaults["raw_dir"]

    if not os.path.isfile(filepath):
        raise HTTPException(status_code=400, detail=f"Excel file not found: {filepath}")

    task_id = str(uuid.uuid4())
    task_status_db[task_id] = {
        "task_id": task_id,
        "status": "processing",
        "progress": 0,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "completed_at": None,
        "excel_output": None,
        "error": None,
    }
    p = Process(target=_run_process_task, args=(task_id, req, task_status_db))
    p.start()
    threading.Thread(target=_watchdog_process, args=(p, task_id), daemon=True).start()
    return {"task_id": task_id, "status": "processing"}


@app.get("/process/status/{task_id}")
def get_process_status(task_id: str) -> Dict[str, Any]:
    mp.active_children()  # 自動回收已結束的殭屍子進程
    task = task_status_db.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail=f"Task not found: {task_id}")
    # 回傳輕量狀態（不含大型 result，避免 Manager.dict 反序列化耗 CPU）
    return {k: v for k, v in task.items() if k != "result"}

@app.get("/process/result/{task_id}")
def get_process_result(task_id: str) -> Dict[str, Any]:
    """取得已完成任務的完整分析結果（從磁碟 JSON 讀取，不走 Manager.dict）。"""
    task = task_status_db.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail=f"Task not found: {task_id}")
    if task.get("status") != "completed":
        raise HTTPException(status_code=400, detail=f"Task not completed yet: {task.get('status')}")
    json_path = task.get("result_json_path")
    if not json_path or not os.path.isfile(json_path):
        raise HTTPException(status_code=404, detail="Result file not found on disk")
    with open(json_path, "r", encoding="utf-8") as _f:
        return json.load(_f)

@app.get("/debug/task/{task_id}/chart_data")
def debug_chart_data(task_id: str, idx: int = 0) -> Dict[str, Any]:
    """Debug endpoint: 回傳指定 task 第 idx 筆結果的 chart_data 狀態。"""
    task = task_status_db.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail=f"Task not found: {task_id}")
    json_path = task.get("result_json_path")
    if not json_path or not os.path.isfile(json_path):
        return {"error": "result file not found", "task_status": task.get("status")}
    with open(json_path, "r", encoding="utf-8") as _f:
        result = json.load(_f)
    results = result.get("results", [])
    if not results:
        return {"error": "no results", "task_status": task.get("status"), "keys_in_result": list(result.keys())}
    item = results[idx] if idx < len(results) else results[0]
    cd = item.get("chart_data")
    return {
        "total_results": len(results),
        "item_keys": list(item.keys()),
        "chart_data_present": cd is not None,
        "chart_data_type": str(type(cd)),
        "chart_data_length": len(cd) if isinstance(cd, list) else None,
        "chart_data_first_record": cd[0] if isinstance(cd, list) and cd else None,
        "chart_path": item.get("chart_path"),
    }

@app.post("/tool-matching")
def analyze_tool_matching(request: ToolMatchingRequest) -> Dict[str, Any]:
    defaults = _default_paths()
    chart_excel_path = request.chart_excel_path or defaults.get("chart_excel_path", "")
    raw_data_directory = request.raw_data_directory or defaults.get("raw_data_directory", "")
    if not os.path.isfile(chart_excel_path):
        raise HTTPException(status_code=400, detail=f"Chart info file not found: {chart_excel_path}")
    if not os.path.isdir(raw_data_directory):
        raise HTTPException(status_code=400, detail=f"Raw data directory not found: {raw_data_directory}")
    task_id = str(uuid.uuid4())
    task_status_db[task_id] = {"task_id": task_id, "status": "processing", "progress": 0,
                                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "completed_at": None, "error": None}
    p = Process(target=_run_tool_matching_task, args=(task_id, request, task_status_db))
    p.start()
    threading.Thread(target=_watchdog_process, args=(p, task_id), daemon=True).start()
    return {"task_id": task_id, "status": "processing"}

@app.post("/spc-cpk")
def analyze_spc_cpk(request: SPCCpkRequest) -> Dict[str, Any]:
    defaults = _default_paths()
    chart_excel_path = request.chart_excel_path or defaults["chart_excel_path"]
    raw_data_directory = request.raw_data_directory or defaults["raw_data_directory"]
    if not os.path.isfile(chart_excel_path):
        raise HTTPException(status_code=400, detail=f"Chart info file not found: {chart_excel_path}")
    if not os.path.isdir(raw_data_directory):
        raise HTTPException(status_code=400, detail=f"Raw data dir not found: {raw_data_directory}")
    task_id = str(uuid.uuid4())
    task_status_db[task_id] = {"task_id": task_id, "status": "processing", "progress": 0,
                                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "completed_at": None, "error": None}
    p = Process(target=_run_spc_cpk_task, args=(task_id, request, task_status_db))
    p.start()
    threading.Thread(target=_watchdog_process, args=(p, task_id), daemon=True).start()
    return {"task_id": task_id, "status": "processing"}

@app.get("/health")
async def health() -> Dict[str, str]:
    return {"status": "ok"}

@app.get("/")
def root() -> Dict[str, Any]:
    return {
        "message": "OOB/SPC FastAPI is running",
        "defaults": _default_paths(),
        "usage": {
            "split": "POST /split with {mode, input_files} → returns split_id + raw_data_directory",
            "split-status": "GET /split-status?split_id=<uuid> → check split folder status",
            "process": "POST /process with {split_id, filepath, save_excel} → returns task_id immediately",
            "process-status": "GET /process/status/{task_id} → poll task progress (processing/completed/failed)",
            "tool-matching": "POST /tool-matching with {filepath, mean_index_threshold, sigma_index_threshold, ...}",
            "spc-cpk": "POST /spc-cpk with {chart_excel_path, raw_data_directory, start_date, end_date, custom_mode, selected_chart}",
            "health": "GET /health",
        },
    }