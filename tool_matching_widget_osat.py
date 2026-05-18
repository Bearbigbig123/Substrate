import os
import pandas as pd
# 使 UI 相關依賴為可選，避免在 FastAPI/無介面環境下匯入失敗
try:
    from PyQt6 import QtWidgets, QtCore, QtGui
    UI_AVAILABLE = True
except Exception:
    # 定義最小替代，以便在無 UI 環境下仍可匯入本模組
    UI_AVAILABLE = False
    class _Dummy: pass
    class _QtWidgets:
        class QWidget: pass
        class QMainWindow: pass
        class QApplication: pass
        class QCheckBox: pass
        class QPushButton: pass
        class QProgressBar: pass
        class QScrollArea: pass
        class QGridLayout: pass
        class QVBoxLayout: pass
        class QHBoxLayout: pass
        class QTabWidget: pass
        class QStackedWidget: pass
        class QLabel: pass
        class QTableWidget: pass
        class QHeaderView: pass
        class QAbstractItemView: pass
        class QDialog: pass
        class QGroupBox: pass
        class QFormLayout: pass
        class QComboBox: pass
        class QStyle: pass
        class QFileDialog: 
            @staticmethod
            def getOpenFileName(*args, **kwargs): return ("", "")
            @staticmethod
            def getSaveFileName(*args, **kwargs): return ("", "")
        class QMessageBox:
            @staticmethod
            def warning(*args, **kwargs): pass
            @staticmethod
            def critical(*args, **kwargs): pass
        class QDoubleSpinBox: pass
        class QSpinBox: pass
        class QLineEdit: pass
        class QDateEdit: pass
        class QToolButton: pass
        class QDialogButtonBox: 
            class StandardButton:
                Close = None
    class _QtCore:
        class Qt:
            class TextFormat:
                RichText = None
            class ScrollBarPolicy:
                ScrollBarAsNeeded = None
                ScrollBarAlwaysOff = None
            class ToolButtonStyle:
                ToolButtonTextBesideIcon = None
            class ArrowType:
                RightArrow = None
                DownArrow = None
            class AlignmentFlag:
                AlignCenter = None
        class QDate:
            @staticmethod
            def currentDate(): return None
        class QSize: pass
    class _QtGui:
        class QFont: pass
        class QIcon: pass
    QtWidgets = _QtWidgets()
    QtCore = _QtCore()
    QtGui = _QtGui()

import pickle # 導入 pickle 模組用於深度複製
import numpy as np

# 檢查是否安裝了統計檢定相關套件
try:
    from scipy import stats
    from scipy.stats import ttest_ind, ttest_rel, f_oneway
    try:
        from statsmodels.stats.multicomp import pairwise_tukeyhsd
        TUKEY_AVAILABLE = True
    except ImportError:
        TUKEY_AVAILABLE = False
        pairwise_tukeyhsd = None
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False
    stats = None
    ttest_ind = None
    ttest_rel = None
    f_oneway = None
    TUKEY_AVAILABLE = False
    pairwise_tukeyhsd = None

# 檢查是否安裝了 openpyxl 套件
try:
    import openpyxl
except ImportError:
    openpyxl = None

class ToolMatchingWidget(QtWidgets.QWidget):
    """
    Tool Matching 分析工具：
    - 讀入 CSV 檔案
    - 根據 GroupName + ChartName 分組
    - 根據 characteristic 進行 mean/sigma matching 檢查
    - 顯示不匹配的結果
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        # 設定全域中文字體為微軟正黑體（僅影響本 widget 及其子元件）
        font = QtGui.QFont("Microsoft JhengHei")
        font.setPointSize(10)
        self.setFont(font)
        self.init_ui()


    def init_ui(self):
        self.setWindowTitle("Tool Matching ")
        self.resize(1200, 800)

        # 主佈局
        self.main_layout = QtWidgets.QVBoxLayout(self)
        self.setLayout(self.main_layout)

        # --- 上方控制區 ---
        top_layout_widget = QtWidgets.QWidget()
        top_layout = QtWidgets.QVBoxLayout(top_layout_widget)

        title = QtWidgets.QLabel("<h2 style='color:#34495E;'>Tool Matching Analysis (OSAT) </h2>")
        # 強制套用微軟正黑體於標題（即使是 HTML）
        title_font = QtGui.QFont("Microsoft JhengHei")
        title_font.setPointSize(16)
        title.setFont(title_font)
        top_layout.addWidget(title)

        file_layout = QtWidgets.QHBoxLayout()
        self.file_path_entry = QtWidgets.QLineEdit()
        self.file_path_entry.setPlaceholderText("Please select a CSV file...")
        self.file_path_entry.setReadOnly(True)
        # 加入資料夾符號於「瀏覽檔案...」按鈕
        file_btn = QtWidgets.QPushButton()
        file_btn.setText("📁 Browse File")
        file_btn.setIcon(QtGui.QIcon())  # 移除原本的 QStyle icon
        btn_font = QtGui.QFont("Microsoft JhengHei")
        btn_font.setBold(True)
        btn_font.setPointSize(12)
        file_btn.setFont(btn_font)
        file_btn.setFixedWidth(165)
        file_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.file_path_entry)
        file_layout.addWidget(file_btn)

        # 新增 temp 按鈕
        temp_btn = QtWidgets.QPushButton("💾 Example")
        temp_btn.setFont(btn_font)
        temp_btn.setFixedWidth(145)
        temp_btn.clicked.connect(self.generate_temp_csv)
        file_layout.addWidget(temp_btn)

        top_layout.addLayout(file_layout)

        # 先插入 mean/sigma index門檻，再插入補滿樣本數欄位
        mean_threshold_layout = QtWidgets.QHBoxLayout()
        self.mean_index_checkbox = QtWidgets.QCheckBox()
        self.mean_index_checkbox.setText("")
        self.mean_index_checkbox.setChecked(False)
        mean_threshold_label = QtWidgets.QLabel("Mean Index Threshold:")
        mean_threshold_label.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.mean_index_threshold_spin = QtWidgets.QDoubleSpinBox()
        self.mean_index_threshold_spin.setRange(0, 10)
        self.mean_index_threshold_spin.setValue(1.0)
        self.mean_index_threshold_spin.setSingleStep(0.1)
        self.mean_index_threshold_spin.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.mean_index_threshold_spin.setEnabled(False)
        mean_threshold_layout.addWidget(self.mean_index_checkbox)
        mean_threshold_layout.addWidget(mean_threshold_label)
        mean_threshold_layout.addWidget(self.mean_index_threshold_spin)
        mean_threshold_layout.addStretch(1)
        def on_mean_checkbox_changed(state):
            self.mean_index_threshold_spin.setEnabled(self.mean_index_checkbox.isChecked())
        self.mean_index_checkbox.stateChanged.connect(on_mean_checkbox_changed)

        sigma_threshold_layout = QtWidgets.QHBoxLayout()
        self.sigma_index_checkbox = QtWidgets.QCheckBox()
        self.sigma_index_checkbox.setText("")
        self.sigma_index_checkbox.setChecked(False)
        sigma_threshold_label = QtWidgets.QLabel("Sigma Index Threshold:")
        sigma_threshold_label.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.sigma_index_threshold_spin = QtWidgets.QDoubleSpinBox()
        self.sigma_index_threshold_spin.setRange(0, 10)
        self.sigma_index_threshold_spin.setValue(2.0)
        self.sigma_index_threshold_spin.setSingleStep(0.1)
        self.sigma_index_threshold_spin.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.sigma_index_threshold_spin.setEnabled(False)
        sigma_threshold_layout.addWidget(self.sigma_index_checkbox)
        sigma_threshold_layout.addWidget(sigma_threshold_label)
        sigma_threshold_layout.addWidget(self.sigma_index_threshold_spin)
        sigma_threshold_layout.addStretch(1)
        def on_sigma_checkbox_changed(state):
            self.sigma_index_threshold_spin.setEnabled(self.sigma_index_checkbox.isChecked())
        self.sigma_index_checkbox.stateChanged.connect(on_sigma_checkbox_changed)

        top_layout.addLayout(mean_threshold_layout)
        
        # 新增統計檢定方法選擇
        stats_method_layout = QtWidgets.QHBoxLayout()
        self.use_statistical_test_checkbox = QtWidgets.QCheckBox()
        self.use_statistical_test_checkbox.setText("t-test")
        self.use_statistical_test_checkbox.setChecked(False)
        self.use_statistical_test_checkbox.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        
        self.stats_method_combo = QtWidgets.QComboBox()
        self.stats_method_combo.addItems(["Unpaired t-test", "Paired t-test"])
        self.stats_method_combo.setFixedWidth(220)
        self.stats_method_combo.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        self.stats_method_combo.setEnabled(False)
        
        self.alpha_level_label = QtWidgets.QLabel("Significance Level:")
        self.alpha_level_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        self.alpha_level_spin = QtWidgets.QDoubleSpinBox()
        self.alpha_level_spin.setRange(0.001, 0.1)
        self.alpha_level_spin.setValue(0.05)
        self.alpha_level_spin.setSingleStep(0.01)
        self.alpha_level_spin.setDecimals(3)
        self.alpha_level_spin.setFixedWidth(80)
        self.alpha_level_spin.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        self.alpha_level_spin.setEnabled(False)
        
        stats_method_layout.addWidget(self.use_statistical_test_checkbox)
        stats_method_layout.addWidget(self.stats_method_combo)
        stats_method_layout.addWidget(self.alpha_level_label)
        stats_method_layout.addWidget(self.alpha_level_spin)
        stats_method_layout.addStretch(1)
        
        def on_stats_checkbox_changed(state):
            enabled = self.use_statistical_test_checkbox.isChecked()
            self.stats_method_combo.setEnabled(enabled)
            self.alpha_level_spin.setEnabled(enabled)
            # 當使用統計檢定時，禁用 Mean Index
            self.mean_index_checkbox.setEnabled(not enabled)
            if enabled:
                self.mean_index_checkbox.setChecked(False)
                self.mean_index_threshold_spin.setEnabled(False)
        
        self.use_statistical_test_checkbox.stateChanged.connect(on_stats_checkbox_changed)
        top_layout.addLayout(stats_method_layout)
        
        top_layout.addLayout(sigma_threshold_layout)


        fillnum_layout = QtWidgets.QHBoxLayout()
        self.fillnum_checkbox = QtWidgets.QCheckBox()
        self.fillnum_checkbox.setText("")
        self.fillnum_checkbox.setChecked(False)
        fillnum_label = QtWidgets.QLabel("Fill Sample Size:")
        fillnum_label.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.fillnum_spin = QtWidgets.QSpinBox()
        self.fillnum_spin.setMinimum(1)
        self.fillnum_spin.setMaximum(100)
        self.fillnum_spin.setValue(5)
        self.fillnum_spin.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.fillnum_spin.setEnabled(False)
        fillnum_layout.addWidget(self.fillnum_checkbox)
        fillnum_layout.addWidget(fillnum_label)
        fillnum_layout.addWidget(self.fillnum_spin)
        fillnum_layout.addStretch(1)
        def on_fillnum_checkbox_changed(state):
            self.fillnum_spin.setEnabled(self.fillnum_checkbox.isChecked())
        self.fillnum_checkbox.stateChanged.connect(on_fillnum_checkbox_changed)
        top_layout.addLayout(fillnum_layout)

        # 新增資料篩選模式選擇
        filter_layout = QtWidgets.QHBoxLayout()
        self.filter_mode_combo = QtWidgets.QComboBox()
        self.filter_mode_combo.addItems(["All Data", "Specified Date (1M / Dormant)", "Latest Data (1M / Dormant)"])
        self.filter_mode_combo.setFixedWidth(260)
        self.filter_mode_combo.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        filter_layout.addWidget(QtWidgets.QLabel("Data Filter Mode:"))
        filter_layout.addWidget(self.filter_mode_combo)

        self.date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate())
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setFont(QtGui.QFont("Microsoft JhengHei", 11))
        self.date_edit.setEnabled(False)
        filter_layout.addWidget(QtWidgets.QLabel("Base Date:"))
        filter_layout.addWidget(self.date_edit)
        filter_layout.addStretch(1)
        top_layout.addLayout(filter_layout)

        def on_filter_mode_changed(idx):
            self.date_edit.setEnabled(idx == 1)
        self.filter_mode_combo.currentIndexChanged.connect(on_filter_mode_changed)

        self.run_btn = QtWidgets.QPushButton("🚀 Run Analysis")
        run_btn_font = QtGui.QFont("Microsoft JhengHei")
        run_btn_font.setBold(True)
        run_btn_font.setPointSize(12)
        self.run_btn.setFont(run_btn_font)
        self.run_btn.clicked.connect(self.run_analysis)
        top_layout.addWidget(self.run_btn)

        # 狀態標籤
        self.status_label = QtWidgets.QLabel("Please select a file and click run analysis.")
        self.status_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        top_layout.addWidget(self.status_label)


        # 可收合的標準說明區塊
        # 1. 摘要說明
        summary_label = QtWidgets.QLabel(
            """
<div style='background-color:#f5f5f5; padding:8px 12px; border-radius:6px; font-size:14px; margin-bottom:4px; font-family:Microsoft JhengHei;'>
  <strong>⚠ Note:</strong> The table shows all analysis results.
  <ul style='margin:8px 0 8px 20px; padding-left:0;'>
    <li><span style='color:#d9534f;'><strong>mean_matching_index ≥ 1</strong></span>：Mean Not Matched (Traditional Method)</li>
    <li><span style='color:#d9534f;'><strong>Statistical Test Significant</strong></span>：p-value &lt; 0.05 (Statistical Test Method)</li>
    <li><span style='color:#d9534f;'><strong>sigma_matching_index ≥ K</strong></span>：Sigma Not Matched</li>
    <li><span style='color:#8a6d3b;'><strong>Insufficient Data</strong></span>：Sample size &lt; 5, no comparison performed</li>
  </ul>
  <span style='color:#5bc0de;'>You can choose to use statistical tests (t-test/ANOVA) instead of the traditional Mean Index method</span><br>
  <span style='color:#d9534f;'>Click "Calculation Formula" below to expand/collapse detailed instructions.</span>
</div>
            """
        )
        summary_label.setWordWrap(True)
        summary_label.setTextFormat(QtCore.Qt.TextFormat.RichText)
        summary_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        top_layout.addWidget(summary_label)

        # 2. 展開/收合按鈕
        self.criterion_toggle_btn = QtWidgets.QToolButton()
        self.criterion_toggle_btn.setText("📘 Calculation Formula (Click to Expand)")
        self.criterion_toggle_btn.setCheckable(True)
        self.criterion_toggle_btn.setChecked(False)
        self.criterion_toggle_btn.setStyleSheet("QToolButton { font-size:13px; color:#344CB7; text-align:left; padding:4px 0; }")
        self.criterion_toggle_btn.setToolButtonStyle(QtCore.Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.criterion_toggle_btn.setArrowType(QtCore.Qt.ArrowType.RightArrow)
        top_layout.addWidget(self.criterion_toggle_btn)

        # 3. 詳細公式內容（預設隱藏，加入滾輪）
        self.criterion_detail_scroll = QtWidgets.QScrollArea()
        self.criterion_detail_scroll.setWidgetResizable(True)
        self.criterion_detail_scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.criterion_detail_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.criterion_detail_scroll.setMaximumHeight(250)
        # 內容 widget
        detail_content = QtWidgets.QWidget()
        detail_layout = QtWidgets.QVBoxLayout(detail_content)
        detail_layout.setContentsMargins(0, 0, 0, 0)
        detail_label = QtWidgets.QLabel(
            """
<div style="background-color:#f5f5f5; padding:12px; border-radius:6px; font-size:14px; line-height:1.6; font-family:Microsoft JhengHei;">
  <strong>📘 Calculation Formula:</strong>
  <table style="font-size:13px; margin-top:8px; font-family:Microsoft JhengHei;">
    <tr>
      <td style="vertical-align:top; padding-right:8px;"><strong>Mean Matching Index:</strong></td>
      <td>
        <u>Two Groups Comparison:</u><br>
        <code>|μ₁ − μ₂| / min(σ₁, σ₂)</code><br>
        <u>Multiple Groups Comparison:</u><br>
        <code>|μ − median(μ)| / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding-right:8px;"><strong>Sigma Matching Index:</strong></td>
      <td>
        <u>Two Groups Comparison:</u><br>
        <code>σ / min(σ₁, σ₂)</code><br>
        <u>Multiple Groups Comparison:</u><br>
        <code>σ / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding-right:8px;"><strong>K Value:</strong></td>
      <td>
        <code>
          n = Sample size<br>
          n ≤ 4: No comparison<br>
          5 ≤ n ≤ 10: K = 1.73<br>
          11 ≤ n ≤ 120: K = 1.414<br>
          n > 120: K = 1.15
        </code>
      </td>
    </tr>
  </table>
  <div style="margin-top:12px; font-size:13px; color:#344CB7;">
    <strong>【Filter Mode Calculation Description】</strong><br>
    <ul style="margin:8px 0 8px 20px; padding-left:0;">
      <li><b>Dormant</b>: If a group has 0 data points in the last 7 days, it is excluded entirely from analysis.</li>
      <li><b>Force Fill</b>: If a group has &ge;1 point in the last 7 days but &lt;5 in the last 30 days, fetch the latest 5 historical records regardless of time limit.</li>
      <li><b>Normal</b>: If a group has &ge;1 point in the last 7 days and &ge;5 in the last 30 days, use all data within 30 days as-is.</li>
      <li>Both mean and sigma windows use the same 30-day dataset.</li>
    </ul>
    <span style="color:#8a6d3b;">(All data mode directly uses all data for group calculation, no filling)</span>
  </div>
</div>
            """
        )
        detail_label.setWordWrap(True)
        detail_label.setTextFormat(QtCore.Qt.TextFormat.RichText)
        detail_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        detail_layout.addWidget(detail_label)
        self.criterion_detail_scroll.setWidget(detail_content)
        self.criterion_detail_scroll.setVisible(False)
        top_layout.addWidget(self.criterion_detail_scroll)

        # 4. 綁定展開/收合事件（滾輪版）
        def toggle_criterion_detail(checked):
            self.criterion_detail_scroll.setVisible(checked)
            if checked:
                self.criterion_toggle_btn.setText("📘 Calculation Formula (Click to Collapse)")
                self.criterion_toggle_btn.setArrowType(QtCore.Qt.ArrowType.DownArrow)
            else:
                self.criterion_toggle_btn.setText("📘 Calculation Formula (Click to Expand)")
                self.criterion_toggle_btn.setArrowType(QtCore.Qt.ArrowType.RightArrow)
        self.criterion_toggle_btn.toggled.connect(toggle_criterion_detail)

        self.main_layout.addWidget(top_layout_widget)
        
    def generate_temp_csv(self):
        # 預設範例資料
        data = {
            "GroupName": ["GroupA"],
            "ChartName": ["X"],
            "point_time": ["2023/5/15 14:39"],
            "matching_group": ["A"],
            "point_val": [99.88135943],
            "characteristic": ["Nominal"]
        }
        df = pd.DataFrame(data)
        # 彈出儲存檔案對話框
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Example CSV File",
            "tool_matching_input_example.csv",
            "CSV Files (*.csv);;All Files (*.*)"
        )
        if not save_path:
            self.status_label.setText("Save example file cancelled.")
            return
        try:
            df.to_csv(save_path, index=False, encoding="utf-8-sig")
            self.status_label.setText(f"Example CSV saved to: {save_path}")
        except Exception as e:
            self.status_label.setText(f"Failed to generate example CSV: {e}")
    def select_file(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select CSV File", "", "CSV Files (*.csv);;All Files (*.*)"
        )
        if file_path:
            self.file_path_entry.setText(file_path)
            self.status_label.setText(f"File selected: {os.path.basename(file_path)}")

    def get_k_value(self, n):
        """根據樣本數量 n 返回 K 值"""
        if n <= 4:  # 樣本數量太少，不進行比較
            return "No Compare"  # 返回特殊標記，表示不進行比較
        elif 5 <= n <= 10:
            return 1.73
        elif 11 <= n <= 120:
            return 1.414
        else:
            return 1.15

    def calculate_mean_index(self, mean1, mean2, min_sigma, characteristic):
        """計算 mean matching index，考慮方向性"""
        if min_sigma <= 0:
            return float('inf')
        
        if characteristic == 'Bigger':  # Bigger is better
            return (mean2 - mean1) / min_sigma
        elif characteristic in ['Smaller', 'Sigma']:  # Smaller is better, Sigma 與 Smaller 邏輯相同
            return (mean1 - mean2) / min_sigma
        else:  # Nominal
            return abs(mean1 - mean2) / min_sigma

    def perform_statistical_test(self, data_groups, method="auto", alpha=0.05):
        """
        執行統計檢定
        
        Parameters:
        - data_groups: dict, key 為群組名稱，value 為數據 list
        - method: str, "unpaired", "paired", "auto"
        - alpha: float, 顯著水準
        
        Returns:
        - dict: 包含檢定結果的字典
        """
        if not SCIPY_AVAILABLE:
            return {"error": "SciPy package not installed, cannot perform statistical tests"}
        
        group_names = list(data_groups.keys())
        group_data = list(data_groups.values())
        
        # 過濾掉空群組
        valid_groups = [(name, data) for name, data in zip(group_names, group_data) if len(data) >= 2]
        
        if len(valid_groups) < 2:
            return {"error": "Insufficient number of valid groups, cannot perform statistical tests"}
        
        valid_names, valid_data = zip(*valid_groups)
        
        result = {
            "method": method,
            "alpha": alpha,
            "groups": valid_names,
            "n_groups": len(valid_groups),
            "significant": False,
            "p_value": None,
            "test_statistic": None,
            "post_hoc": None
        }
        
        try:
            if len(valid_groups) == 2:
                # 兩群比較
                group1_data, group2_data = valid_data[0], valid_data[1]
                
                if method == "paired":
                    # 檢查是否能配對（樣本數相等）
                    if len(group1_data) == len(group2_data):
                        # 配對 t 檢定
                        stat, p_val = ttest_rel(group1_data, group2_data)
                        result["method"] = "Paired t-test"
                    else:
                        return {
                            "error": f"配對 t 檢定需要兩組樣本數相等。目前：{valid_names[0]}={len(group1_data)} vs {valid_names[1]}={len(group2_data)}"
                        }
                elif method == "unpaired":
                    # 非配對 t 檢定
                    stat, p_val = ttest_ind(group1_data, group2_data)
                    result["method"] = "Unpaired t-test"
                else:
                    # 預設使用非配對 t 檢定
                    stat, p_val = ttest_ind(group1_data, group2_data)
                    result["method"] = "Unpaired t-test"
                
                result["test_statistic"] = stat
                result["p_value"] = p_val
                result["significant"] = p_val < alpha
                
            else:
                # 三群以上：使用 ANOVA
                stat, p_val = f_oneway(*valid_data)
                result["method"] = "One-way ANOVA"
                result["test_statistic"] = stat
                result["p_value"] = p_val
                result["significant"] = p_val < alpha
                
                # 如果 ANOVA 顯著且有 Tukey HSD 可用，進行事後檢定
                if result["significant"] and TUKEY_AVAILABLE:
                    # 準備 Tukey HSD 的數據格式
                    all_data = []
                    all_groups = []
                    for i, (name, data) in enumerate(valid_groups):
                        all_data.extend(data)
                        all_groups.extend([name] * len(data))
                    
                    try:
                        tukey_result = pairwise_tukeyhsd(all_data, all_groups, alpha=alpha)
                        result["post_hoc"] = {
                            "method": "Tukey HSD",
                            "summary": str(tukey_result),
                            "significant_pairs": []
                        }
                        
                        # 解析顯著的配對
                        for i, row in enumerate(tukey_result.summary().data[1:]):  # 跳過標題行
                            group1, group2, meandiff, p_adj, lower, upper, reject = row
                            if reject:  # reject=True indicates significant difference
                                result["post_hoc"]["significant_pairs"].append({
                                    "group1": group1,
                                    "group2": group2, 
                                    "p_adj": p_adj,
                                    "mean_diff": meandiff
                                })
                    except Exception as e:
                        result["post_hoc"] = {"error": f"Tukey HSD failed: {str(e)}"}
                
        except Exception as e:
            result["error"] = f"Statistical test failed: {str(e)}"
        
        return result

    def run_analysis(self):
        file_path = self.file_path_entry.text()
        if not file_path or not os.path.exists(file_path):
            self.status_label.setText("Please select a valid CSV file first!")
            return

        try:
            df = pd.read_csv(file_path)
        except Exception as e:
            self.status_label.setText(f"Failed to read file: {e}")
            return

        # 檢查必要欄位
        required_cols = ["GroupName", "ChartName", "matching_group", "point_val", "characteristic", "point_time"]
        for col in required_cols:
            if col not in df.columns:
                self.status_label.setText(f"Missing required column: {col}")
                return

        # 轉換 point_time 為 datetime
        try:
            df["point_time"] = pd.to_datetime(df["point_time"])
        except Exception as e:
            self.status_label.setText(f"Failed to convert point_time column: {e}")
            return

        filter_mode = self.filter_mode_combo.currentIndex()
        base_date = self.date_edit.date().toPyDate() if filter_mode == 1 else None

        # 取得補滿筆數
        fill_num = self.fillnum_spin.value()

        results = []

        if filter_mode == 0:
            # 全算
            grouped = df.groupby(["GroupName", "ChartName"])
            print("\n[DEBUG] All unique (GroupName, ChartName) pairs:")
            for pair in grouped.groups.keys():
                print("  ", pair)
            for (gname, cname), subdf in grouped:
                print(f"[DEBUG] Now processing group: GroupName='{gname}', ChartName='{cname}' | subdf.shape={subdf.shape}")
                characteristic = subdf["characteristic"].dropna().unique()
                if len(characteristic) != 1:
                    self.status_label.setText(f"Group: {gname}-{cname} characteristic is not unique or missing")
                    continue
                group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count']).reset_index()
                n_groups = len(group_stats)
                
                # 檢查有效群組數量（樣本數 >= 5）
                valid_groups = group_stats[group_stats['count'] >= 5]
                n_valid_groups = len(valid_groups)
                
                if n_valid_groups == 2 and n_groups > 2:
                    # 原本是多群，但只有兩群有效，改為兩群比較
                    print(f"[INFO] {gname}-{cname}: Originally {n_groups} groups, but only {n_valid_groups} groups are valid, switching to two-group comparison")
                    self._analyze_two_groups(valid_groups, gname, cname, characteristic[0], results, subdf)
                elif n_groups == 2:
                    self._analyze_two_groups(group_stats, gname, cname, characteristic[0], results, subdf)
                else:
                    self._analyze_multiple_groups(subdf, group_stats, gname, cname, characteristic[0], results)
            self._create_boxplots(grouped)
        elif filter_mode == 1:
            # 指定日期模式
            grouped = df.groupby(["GroupName", "ChartName"])
            print("\n[DEBUG] All unique (GroupName, ChartName) pairs:")
            for pair in grouped.groups.keys():
                print("  ", pair)
            sigma_df_all = []  # 收集所有 30 天資料
            mean_df_all = []   # 收集所有 30 天資料（三段式邏輯處理後）
            for (gname, cname), subdf in grouped:
                print(f"[DEBUG] Now processing group: GroupName='{gname}', ChartName='{cname}' | subdf.shape={subdf.shape}")
                characteristic = subdf["characteristic"].dropna().unique()
                if len(characteristic) != 1:
                    self.status_label.setText(f"Group: {gname}-{cname} characteristic is not unique or missing")
                    continue
                mean_end = pd.Timestamp(base_date)
                mean_start = mean_end - pd.DateOffset(days=30)
                week_start = mean_end - pd.Timedelta(days=7)
                MIN_FILL = 5
                # 先抓 30 天初始區間（mean 與 sigma 共用同一資料集）
                mean_df = subdf[(subdf["point_time"] > mean_start) & (subdf["point_time"] <= mean_end)].copy()
                sigma_df = mean_df.copy()
                # 三段式補點邏輯：依當週活躍度決定處理方式
                for mg in subdf["matching_group"].unique():
                    mg_week = subdf[
                        (subdf["matching_group"] == mg) &
                        (subdf["point_time"] > week_start) &
                        (subdf["point_time"] <= mean_end)
                    ]
                    if len(mg_week) == 0:
                        # Dormant：當週無資料，完全排除此 group
                        mean_df = mean_df[mean_df["matching_group"] != mg]
                        sigma_df = sigma_df[sigma_df["matching_group"] != mg]
                        print(f"[INFO] {gname}-{cname}: group '{mg}' Dormant (no data in last 7 days), excluded.")
                        continue
                    mg_30d = mean_df[mean_df["matching_group"] == mg]
                    if len(mg_30d) < MIN_FILL:
                        # 強制補足：取歷史最新 MIN_FILL 筆（不限時間）
                        mg_fill = (
                            subdf[subdf["matching_group"] == mg]
                            .sort_values("point_time", ascending=False)
                            .head(MIN_FILL)
                        )
                        mean_df = pd.concat([mean_df[mean_df["matching_group"] != mg], mg_fill]).drop_duplicates()
                        sigma_df = pd.concat([sigma_df[sigma_df["matching_group"] != mg], mg_fill]).drop_duplicates()
                    # else: 正常使用 30 天內所有資料（不做任何補點）
                mean_df_all.append(mean_df.assign(GroupName=gname, ChartName=cname))
                sigma_df_all.append(sigma_df.assign(GroupName=gname, ChartName=cname))
                mean_stats = mean_df.groupby("matching_group")["point_val"].agg(['mean', 'count']).reset_index()
                sigma_stats = sigma_df.groupby("matching_group")["point_val"].agg(['std']).reset_index()
                group_stats = pd.merge(mean_stats, sigma_stats, on="matching_group", how="outer")
                group_stats = group_stats.fillna({"mean": 0, "std": 0, "count": 0})
                n_groups = len(group_stats)
                
                # 檢查有效群組數量（樣本數 >= 5）
                valid_groups = group_stats[group_stats['count'] >= 5]
                n_valid_groups = len(valid_groups)
                
                if n_valid_groups == 2 and n_groups > 2:
                    # 原本是多群，但只有兩群有效，改為兩群比較
                    print(f"[INFO] {gname}-{cname}: 原有{n_groups}群，但只有{n_valid_groups}群有效，切換為兩群比較")
                    combined_data = pd.concat([mean_df, sigma_df]).drop_duplicates()
                    self._analyze_two_groups(valid_groups, gname, cname, characteristic[0], results, combined_data)
                elif n_groups == 2:
                    # 對於指定日期模式，使用合併的數據進行統計檢定
                    combined_data = pd.concat([mean_df, sigma_df]).drop_duplicates()
                    self._analyze_two_groups(group_stats, gname, cname, characteristic[0], results, combined_data)
                else:
                    self._analyze_multiple_groups_time(mean_df, sigma_df, group_stats, gname, cname, characteristic[0], results)
            if mean_df_all:
                mean_df_concat = pd.concat(mean_df_all, ignore_index=True)
                mean_grouped = mean_df_concat.groupby(["GroupName", "ChartName"])
                self._create_boxplots(mean_grouped)
            else:
                self._create_boxplots(grouped)
        elif filter_mode == 2:
            # 最新進點模式
            grouped = df.groupby(["GroupName", "ChartName"])
            sigma_df_all = []
            mean_df_all = []
            for (gname, cname), subdf in grouped:
                characteristic = subdf["characteristic"].dropna().unique()
                if len(characteristic) != 1:
                    self.status_label.setText(f"Group: {gname}-{cname} 的 characteristic 不唯一或缺失")
                    continue
                latest_time = subdf["point_time"].max()
                mean_end = latest_time
                mean_start = mean_end - pd.DateOffset(days=30)
                week_start = mean_end - pd.Timedelta(days=7)
                MIN_FILL = 5
                # 先抓 30 天初始區間（mean 與 sigma 共用同一資料集）
                mean_df = subdf[(subdf["point_time"] > mean_start) & (subdf["point_time"] <= mean_end)].copy()
                sigma_df = mean_df.copy()
                # 三段式補點邏輯：依當週活躍度決定處理方式
                for mg in subdf["matching_group"].unique():
                    mg_week = subdf[
                        (subdf["matching_group"] == mg) &
                        (subdf["point_time"] > week_start) &
                        (subdf["point_time"] <= mean_end)
                    ]
                    if len(mg_week) == 0:
                        # Dormant：當週無資料，完全排除此 group
                        mean_df = mean_df[mean_df["matching_group"] != mg]
                        sigma_df = sigma_df[sigma_df["matching_group"] != mg]
                        print(f"[INFO] {gname}-{cname}: group '{mg}' Dormant (no data in last 7 days), excluded.")
                        continue
                    mg_30d = mean_df[mean_df["matching_group"] == mg]
                    if len(mg_30d) < MIN_FILL:
                        # 強制補足：取歷史最新 MIN_FILL 筆（不限時間）
                        mg_fill = (
                            subdf[subdf["matching_group"] == mg]
                            .sort_values("point_time", ascending=False)
                            .head(MIN_FILL)
                        )
                        mean_df = pd.concat([mean_df[mean_df["matching_group"] != mg], mg_fill]).drop_duplicates()
                        sigma_df = pd.concat([sigma_df[sigma_df["matching_group"] != mg], mg_fill]).drop_duplicates()
                    # else: 正常使用 30 天內所有資料（不做任何補點）
                mean_df_all.append(mean_df.assign(GroupName=gname, ChartName=cname))
                sigma_df_all.append(sigma_df.assign(GroupName=gname, ChartName=cname))
                mean_stats = mean_df.groupby("matching_group")["point_val"].agg(['mean', 'count']).reset_index()
                sigma_stats = sigma_df.groupby("matching_group")["point_val"].agg(['std']).reset_index()
                group_stats = pd.merge(mean_stats, sigma_stats, on="matching_group", how="outer")
                group_stats = group_stats.fillna({"mean": 0, "std": 0, "count": 0})
                n_groups = len(group_stats)
                
                # 檢查有效群組數量（樣本數 >= 5）
                valid_groups = group_stats[group_stats['count'] >= 5]
                n_valid_groups = len(valid_groups)
                
                if n_valid_groups == 2 and n_groups > 2:
                    # 原本是多群，但只有兩群有效，改為兩群比較
                    print(f"[INFO] {gname}-{cname}: Originally {n_groups} groups, but only {n_valid_groups} groups are valid, switching to two-group comparison")
                    combined_data = pd.concat([mean_df, sigma_df]).drop_duplicates()
                    self._analyze_two_groups(valid_groups, gname, cname, characteristic[0], results, combined_data)
                elif n_groups == 2:
                    # 對於最新進點模式，使用合併的數據進行統計檢定
                    combined_data = pd.concat([mean_df, sigma_df]).drop_duplicates()
                    self._analyze_two_groups(group_stats, gname, cname, characteristic[0], results, combined_data)
                else:
                    self._analyze_multiple_groups_time(mean_df, sigma_df, group_stats, gname, cname, characteristic[0], results)
            if mean_df_all:
                mean_df_concat = pd.concat(mean_df_all, ignore_index=True)
                mean_grouped = mean_df_concat.groupby(["GroupName", "ChartName"])
                self._create_boxplots(mean_grouped)
            else:
                self._create_boxplots(grouped)

        self._display_results(results)

    def _analyze_multiple_groups_time(self, mean_df, sigma_df, group_stats, gname, cname, characteristic, results):
        """
        多組分析（mean/std/count 與 median(sigma) 共用同一 30 天 window，三段式補點邏輯處理後）
        - mean, std, count: 來自 mean_df（30 天 window，三段式補點）
        - median_sigma: 來自 sigma_df（同 mean_df，三段式補點）
        """
        # 只納入樣本數 >= 5 的 group 計算 median
        valid_mean_df = mean_df.groupby("matching_group").filter(lambda x: len(x) >= 5)
        sigma_by_group = sigma_df.groupby("matching_group")["point_val"].std()
        valid_groups = group_stats[group_stats['count'] >= 5]['matching_group']
        valid_sigma = sigma_by_group[valid_groups] if not valid_groups.empty else pd.Series(dtype=float)
        # 防呆：如果有效 group 只有一個，全部標記資料不足
        if len(valid_groups) <= 1:
            for i, row in group_stats.iterrows():
                group = row["matching_group"]
                mean = row["mean"]
                std = row["std"]
                n = row["count"]
                results.append([
                    gname, cname, group, "group_all",
                    'Insufficient Data', 'Insufficient Data', 
                    self.get_k_value(n), mean, std, 
                    '-', '-', n, characteristic
                ])
            return
        mean_median = valid_mean_df["point_val"].median() if not valid_mean_df.empty else 0
        median_sigma = valid_sigma.median() if not valid_sigma.empty else 0
        for i, row in group_stats.iterrows():
            group = row["matching_group"]
            mean = row["mean"]
            std = row["std"]  # 這是來自 mean_df（一個月 window）
            n = row["count"]
            if n < 5:
                results.append([
                    gname, cname, group, "group_all",
                    'Insufficient Data', 'Insufficient Data', 
                    self.get_k_value(n), mean, std, 
                    mean_median, median_sigma, n, characteristic
                ])
                continue
            if median_sigma > 0:
                if characteristic == 'Bigger':
                    mean_index = (mean_median - mean) / median_sigma
                elif characteristic in ['Smaller', 'Sigma']:
                    mean_index = (mean - mean_median) / median_sigma
                else:
                    mean_index = abs(mean - mean_median) / median_sigma
                sigma_index = std / median_sigma
            else:
                # 分母為零時，判斷所有 mean 是否相等
                all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    mean_index = 0
                    sigma_index = 0
                else:
                    mean_index = float('inf')
                    sigma_index = float('inf')
            K = self.get_k_value(n)
            if K == "No Compare":
                results.append([
                    gname, cname, group, "group_all",
                    'Insufficient Data', 'Insufficient Data', 
                    'No Compare', round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n, characteristic
                ])
            else:
                results.append([
                    gname, cname, group, "group_all",
                    round(mean_index, 2), round(sigma_index, 2), 
                    round(K, 2), round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n, characteristic
                ])

    def _analyze_two_groups(self, group_stats, gname, cname, characteristic, results, raw_data=None):
        """分析兩台設備的匹配情況"""
        row1 = group_stats.iloc[0]
        row2 = group_stats.iloc[1]

        group1 = row1["matching_group"]
        group2 = row2["matching_group"]
        mean1, std1, n1 = row1["mean"], row1["std"], row1["count"]
        mean2, std2, n2 = row2["mean"], row2["std"], row2["count"]

        min_sigma = min(std1, std2)

        # 檢查是否使用統計檢定
        use_stats = self.use_statistical_test_checkbox.isChecked()
        
        if use_stats and raw_data is not None and SCIPY_AVAILABLE:
            # 準備統計檢定的數據
            group1_data = raw_data[raw_data["matching_group"] == group1]["point_val"].tolist()
            group2_data = raw_data[raw_data["matching_group"] == group2]["point_val"].tolist()
            
            if len(group1_data) >= 2 and len(group2_data) >= 2:
                # 執行統計檢定
                method_map = {
                    0: "unpaired",
                    1: "paired"
                }
                method = method_map.get(self.stats_method_combo.currentIndex(), "unpaired")
                alpha = self.alpha_level_spin.value()
                
                data_groups = {group1: group1_data, group2: group2_data}
                stats_result = self.perform_statistical_test(data_groups, method, alpha)
                
                # 根據統計檢定結果決定 mean matching 狀態
                if "error" in stats_result:
                    error_msg = stats_result['error']
                    mean_status_1 = f"Statistical test failed: {error_msg}"
                    mean_status_2 = f"Statistical test failed: {error_msg}"
                    
                    # 如果是配對檢定樣本數不一致的錯誤，顯示詳細的前端提示
                    if "配對 t 檢定需要兩組樣本數相等" in error_msg:
                        QtWidgets.QMessageBox.warning(
                            self, 
                            "Paired Test Error", 
                            f"Paired t-test requires equal sample sizes!\n\n"
                            f"Problem group: {gname} - {cname}\n"
                            f"Current sample sizes:\n"
                            f"• {group1}: {len(group1_data)} samples\n"
                            f"• {group2}: {len(group2_data)} samples\n\n"
                            f"Suggested solutions:\n"
                            f"1. Use 'Unpaired t-test' instead\n"
                            f"2. Ensure data has proper pairing relationship"
                        )
                else:
                    p_val = stats_result.get("p_value", 1.0)
                    if stats_result.get("significant", False):
                        mean_status_1 = f"Significant (p={p_val:.4f})"
                        mean_status_2 = f"Significant (p={p_val:.4f})"
                    else:
                        mean_status_1 = f"No Significant (p={p_val:.4f})"
                        mean_status_2 = f"No Significant (p={p_val:.4f})"
            else:
                mean_status_1 = "Statistical test - Insufficient data"
                mean_status_2 = "Statistical test - Insufficient data"
        else:
            # 使用原有的 Mean Index 計算（考慮方向性）
            mean_index_1 = self.calculate_mean_index(mean1, mean2, min_sigma, characteristic)
            mean_index_2 = self.calculate_mean_index(mean2, mean1, min_sigma, characteristic)
            
            mean_status_1 = round(mean_index_1, 2)
            mean_status_2 = round(mean_index_2, 2)

        # 計算 sigma index (保持原邏輯)
        if min_sigma > 0:
            sigma_index_1 = std1 / min_sigma
            sigma_index_2 = std2 / min_sigma
        else:
            sigma_index_1 = 0 if std1 == std2 else float('inf')
            sigma_index_2 = 0 if std1 == std2 else float('inf')

        # 統一格式：第4欄都用 'group_all'，與多群分析一致
        # mean_median, sigma_median 欄位（兩組時用 mean2, min_sigma 或 mean1, min_sigma）
        # 這裡用 mean2, min_sigma for group1, mean1, min_sigma for group2

        if n1 < 5 or n2 < 5:
            results.append([
                gname, cname, group1, 'group_all',
                'Insufficient Data', 'Insufficient Data',
                self.get_k_value(n1), mean1, std1,
                mean2, min_sigma, n1, characteristic
            ])
            results.append([
                gname, cname, group2, 'group_all',
                'Insufficient Data', 'Insufficient Data',
                self.get_k_value(n2), mean2, std2,
                mean1, min_sigma, n2, characteristic
            ])
            return

        k1 = self.get_k_value(n1)
        k2 = self.get_k_value(n2)

        if k1 == "No Compare":
            results.append([
                gname, cname, group1, 'group_all',
                'Insufficient Data', 'Insufficient Data',
                'No Compare', round(mean1, 2), round(std1, 2),
                round(mean2, 2), round(min_sigma, 2), n1, characteristic
            ])
        else:
            results.append([
                gname, cname, group1, 'group_all',
                mean_status_1, round(sigma_index_1, 2),
                round(k1, 2), round(mean1, 2), round(std1, 2),
                round(mean2, 2), round(min_sigma, 2), n1, characteristic
            ])

        if k2 == "No Compare":
            results.append([
                gname, cname, group2, 'group_all',
                'Insufficient Data', 'Insufficient Data',
                'No Compare', round(mean2, 2), round(std2, 2),
                round(mean1, 2), round(min_sigma, 2), n2, characteristic
            ])
        else:
            results.append([
                gname, cname, group2, 'group_all',
                mean_status_2, round(sigma_index_2, 2),
                round(k2, 2), round(mean2, 2), round(std2, 2),
                round(mean1, 2), round(min_sigma, 2), n2, characteristic
            ])

    def _analyze_multiple_groups(self, subdf, group_stats, gname, cname, characteristic, results):
        """分析多台設備的匹配情況 (mean matching index 分母都用 median_sigma)"""
        # 只納入樣本數 >= 5 的 group 計算 median
        valid_stats = group_stats[group_stats['count'] >= 5]
        if valid_stats.shape[0] <= 1:
            # 只有一個有效群組，全部標記資料不足
            for i, row in group_stats.iterrows():
                group = row["matching_group"]
                mean = row["mean"]
                std = row["std"]
                n = row["count"]
                results.append([
                    gname, cname, group, "group_all",
                    'Insufficient Data', 'Insufficient Data', 
                    self.get_k_value(n), mean, std, 
                    '-', '-', n, characteristic
                ])
            return

        mean_median = valid_stats['mean'].median() if not valid_stats.empty else 0
        median_sigma = valid_stats['std'].median() if not valid_stats.empty else 0

        # 檢查是否使用統計檢定
        use_stats = self.use_statistical_test_checkbox.isChecked()
        stats_result = None
        
        if use_stats and SCIPY_AVAILABLE and len(valid_stats) >= 3:
            # 準備 ANOVA 的數據
            data_groups = {}
            for _, row in valid_stats.iterrows():
                group_name = row["matching_group"]
                group_data = subdf[subdf["matching_group"] == group_name]["point_val"].tolist()
                if len(group_data) >= 2:
                    data_groups[group_name] = group_data
            
            if len(data_groups) >= 3:
                alpha = self.alpha_level_spin.value()
                stats_result = self.perform_statistical_test(data_groups, "unpaired", alpha)

        for i, row in group_stats.iterrows():
            group = row["matching_group"]
            mean = row["mean"]
            std = row["std"]
            n = row["count"]

            # 計算 mean matching index（考慮方向性）
            if n < 5:  # 樣本數不足5個，不進行比較
                results.append([
                    gname, cname, group, "group_all",
                    'Insufficient Data', 'Insufficient Data', 
                    self.get_k_value(n), mean, std, 
                    mean_median, median_sigma, n
                ])
                continue

            # 決定 mean matching 狀態
            if use_stats and stats_result and "error" not in stats_result:
                # 使用統計檢定結果
                if stats_result.get("significant", False):
                    # ANOVA 顯著，檢查事後檢定結果
                    mean_status = "ANOVA Significant"
                    if stats_result.get("post_hoc") and "significant_pairs" in stats_result["post_hoc"]:
                        # 檢查這個群組是否在顯著配對中
                        significant_pairs = stats_result["post_hoc"]["significant_pairs"]
                        group_in_significant = any(
                            pair["group1"] == group or pair["group2"] == group 
                            for pair in significant_pairs
                        )
                        if group_in_significant:
                            # 找出與此群組有顯著差異的其他群組
                            different_groups = []
                            for pair in significant_pairs:
                                if pair["group1"] == group:
                                    different_groups.append(pair["group2"])
                                elif pair["group2"] == group:
                                    different_groups.append(pair["group1"])
                            if different_groups:
                                mean_status = f"Significant vs {','.join(different_groups)}"
                            else:
                                mean_status = "Participate in Significant"
                        else:
                            mean_status = "No Significant"
                    else:
                        mean_status = f"ANOVA Significant(p={stats_result.get('p_value', 0):.4f})"
                else:
                    mean_status = f"No Significant(p={stats_result.get('p_value', 1):.4f})"
            else:
                # 使用原有的 Mean Index 計算
                if median_sigma > 0:
                    if characteristic == 'Bigger':
                        mean_index = (mean_median - mean) / median_sigma
                    elif characteristic in ['Smaller', 'Sigma']:
                        mean_index = (mean - mean_median) / median_sigma
                    else:
                        mean_index = abs(mean - mean_median) / median_sigma
                else:
                    # 分母為零時，判斷所有 mean 是否相等
                    all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                    if len(set([round(m, 8) for m in all_means])) == 1:
                        mean_index = 0
                    else:
                        mean_index = float('inf')
                
                mean_status = round(mean_index, 2)

            # 計算 sigma index (保持原邏輯)
            if median_sigma > 0:
                sigma_index = std / median_sigma
            else:
                all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    sigma_index = 0
                else:
                    sigma_index = float('inf')

            K = self.get_k_value(n)

            # 檢查 K 值是否為字串 "No Compare"
            if K == "No Compare":
                # 樣本數不足，使用 "Insufficient Data" 標記
                results.append([
                    gname, cname, group, "group_all",
                    'Insufficient Data', 'Insufficient Data', 
                    'No Compare', round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n, characteristic
                ])
            else:
                # 正常比較情況
                # 無論是否匹配都添加結果，保證所有比較都出現在報表中
                results.append([
                    gname, cname, group, "group_all",
                    mean_status, round(sigma_index, 2), 
                    round(K, 2), round(mean, 2), round(std, 2), 
                    round(mean_median, 2), round(median_sigma, 2), n, characteristic
                ])

    def _display_results(self, results):
        """以新格式顯示分析結果，並在表格中添加按鈕以查看詳情。"""
        # 儲存報告數據以供彈出視窗使用
        self.report_data = {}
        
        # 遍歷結果，整理報表資料
        for row in results:
            gname, cname = row[0], row[1]
            key = f"{gname}_{cname}"
            
            if key not in self.report_data:
                self.report_data[key] = {
                    "GroupName": gname,
                    "ChartName": cname,
                    "groups": {}
                }
            group1, group2 = row[2], row[3]
            mean_index = row[4]
            sigma_index = row[5]
            
            if len(row) >= 13:
                k_value, mean, sigma, mean_median, sigma_median, n, characteristic = row[6:13]
            else:
                k_value, mean, sigma, mean_median, sigma_median, n, characteristic = [""] * 6 + [row[6] if len(row) > 6 else ""]
            
            if group2 == "group_all":
                self.report_data[key]["groups"][group1] = {
                    "mean_matching_index": mean_index,
                    "sigma_matching_index": sigma_index,
                    "K": k_value,
                    "mean": mean,
                    "sigma": sigma,
                    "mean_median": mean_median,
                    "sigma_median": sigma_median,
                    "samplesize": n,
                    "characteristic": characteristic
                }
            else:
                if group1 not in self.report_data[key]["groups"]:
                    self.report_data[key]["groups"][group1] = {}
                self.report_data[key]["groups"][group1][group2] = {
                    "mean_matching_index": mean_index,
                    "sigma_matching_index": sigma_index,
                    "K": k_value,
                    "mean": mean,
                    "sigma": sigma,
                    "mean_median": mean_median,
                    "sigma_median": sigma_median,
                    "samplesize": n,
                    "characteristic": characteristic
                }

        all_table_rows = []
        all_ui_rows = []  # 改為顯示所有項目，不只異常項目
        
        for key, data in self.report_data.items():
            gname = data["GroupName"]
            cname = data["ChartName"]
            
            for group_id, stats in data["groups"].items():
                mean_index = stats.get("mean_matching_index", "")
                sigma_index = stats.get("sigma_matching_index", "")
                k_value = stats.get("K", "")
                
                is_abnormal = False
                is_data_insufficient = mean_index == 'Insufficient Data' or sigma_index == 'Insufficient Data' or k_value == 'No Compare'
                abnormal_type = ""
                
                # 檢查是否為統計檢定的顯著差異
                is_statistical_significant = False
                if isinstance(mean_index, str) and ("Significant" in mean_index or "ANOVA" in mean_index):
                    # 排除「No Significant」的情況
                    if "No Significant" not in mean_index:
                        is_statistical_significant = True
                    
                if not is_data_insufficient:
                    # 先檢查 Sigma Index 是否異常
                    sigma_abn = False
                    try:
                        sigma_threshold = self.sigma_index_threshold_spin.value() if self.sigma_index_checkbox.isChecked() else float(k_value) if k_value not in [None, '', 'No Compare'] else 2.0
                        sigma_abn = float(sigma_index) >= sigma_threshold
                    except (ValueError, TypeError):
                        sigma_abn = False
                    
                    if is_statistical_significant:
                        # 統計檢定顯著，但也要檢查 Sigma Index
                        is_abnormal = True
                        if sigma_abn:
                            abnormal_type = "Mean, Sigma"
                        else:
                            abnormal_type = "Mean"
                    else:
                        try:
                            # 只有勾選核取方塊才啟用門檻判斷，否則用預設值
                            mean_threshold = self.mean_index_threshold_spin.value() if self.mean_index_checkbox.isChecked() else 1.0
                            mean_abn = float(mean_index) >= mean_threshold
                            if mean_abn or sigma_abn:
                                is_abnormal = True
                                if mean_abn and sigma_abn:
                                    abnormal_type = "Mean, Sigma"
                                elif mean_abn:
                                    abnormal_type = "Mean"
                                elif sigma_abn:
                                    abnormal_type = "Sigma"
                        except (ValueError, TypeError):
                            pass
                else:
                    abnormal_type = ""
                
                # 樣本數 n 強制轉為 int 顯示
                samplesize_val = stats.get("samplesize", "")
                try:
                    if samplesize_val != '' and samplesize_val is not None:
                        samplesize_val = int(float(samplesize_val))
                except Exception:
                    pass
                row_data = [
                    gname, cname, group_id,
                    stats.get("mean_matching_index", ""), stats.get("sigma_matching_index", ""),
                    stats.get("K", ""), stats.get("mean", ""), stats.get("sigma", ""),
                    stats.get("mean_median", ""), stats.get("sigma_median", ""),
                    samplesize_val, stats.get("characteristic", "")
                ]
                
                all_row_data = [is_abnormal, abnormal_type] + row_data
                all_table_rows.append(all_row_data)
                
                # 將所有項目加入 UI 顯示列表，不只異常項目
                all_ui_rows.append({
                    "key": (gname, cname),
                    "group_id": group_id,
                    "data": [abnormal_type] + row_data,
                    "is_abnormal": is_abnormal,
                    "is_data_insufficient": is_data_insufficient
                })


        # 匯出全部結果到 Excel 檔案
        if all_table_rows and hasattr(self, 'file_path_entry') and self.file_path_entry.text():
            self._export_to_excel(all_table_rows, self.file_path_entry.text())
            
        # 統計異常項目數量
        abnormal_count = sum(1 for item in all_ui_rows if item["is_abnormal"])
        insufficient_count = sum(1 for item in all_ui_rows if item["is_data_insufficient"])
        total_count = len(all_ui_rows)
        
        status_msg = f"Analysis completed. Total: {total_count} items."
        if abnormal_count > 0:
            status_msg += f" Abnormal: {abnormal_count} items."
        if insufficient_count > 0:
            status_msg += f" Insufficient data: {insufficient_count} items."
        if abnormal_count == 0 and insufficient_count == 0:
            status_msg += " No issues found."
        
        self.status_label.setText(status_msg)

    def _show_details_dialog(self, chart_key, group_id):
        """彈出一個視窗，顯示詳細資訊和圖表，上方為數據，下方為圖表。"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"Detailed Information: {chart_key[0]} - {chart_key[1]} | Group: {group_id}")
        dialog.setMinimumSize(1400, 450) # 調整視窗大小以適應新佈局 (高度減少)

        main_layout = QtWidgets.QVBoxLayout(dialog)
        main_layout.setSpacing(10)

        # --- 上方：數據表格 (水平排列) ---
        try:
            stats = self.report_data[f"{chart_key[0]}_{chart_key[1]}"]["groups"][group_id]
        except KeyError:
            QtWidgets.QMessageBox.critical(self, "Error", "Cannot find detailed data for this item.")
            return


        info_group = QtWidgets.QGroupBox("Analysis Data")
        info_v_layout = QtWidgets.QVBoxLayout(info_group)

        # 取得異常類型
        # 這裡需重算異常類型，與 UI/Excel 一致
        mean_index = stats.get("mean_matching_index", "")
        sigma_index = stats.get("sigma_matching_index", "")
        k_value = stats.get("K", "")
        abnormal_type = ""
        is_data_insufficient = mean_index == 'Insufficient Data' or sigma_index == 'Insufficient Data' or k_value == 'No Compare'
        
        # 檢查是否為統計檢定的顯著差異
        is_statistical_significant = False
        if isinstance(mean_index, str) and ("Significant" in mean_index or "ANOVA" in mean_index):
            # 排除「No Significant」的情況
            if "No Significant" not in mean_index:
                is_statistical_significant = True
            
        if not is_data_insufficient:
            if is_statistical_significant:
                # 統計檢定顯著
                abnormal_type = "Statistical Test Significant"
            else:
                try:
                    mean_threshold = self.mean_index_threshold_spin.value() if self.mean_index_checkbox.isChecked() else 1.0
                    sigma_threshold = self.sigma_index_threshold_spin.value() if self.sigma_index_checkbox.isChecked() else float(k_value) if k_value not in [None, '', 'No Compare'] else 2.0
                    mean_abn = float(mean_index) >= mean_threshold
                    sigma_abn = float(sigma_index) >= sigma_threshold
                    if mean_abn and sigma_abn:
                        abnormal_type = "Mean, Sigma"
                    elif mean_abn:
                        abnormal_type = "Mean"
                    elif sigma_abn:
                        abnormal_type = "Sigma"
                except (ValueError, TypeError):
                    pass

        # 新增異常類型欄位
        headers = [
            "Abnormal Type", "Group Name", "Chart Name", "Matching Group", "Mean Index", "Sigma Index",
            "K", "Mean", "Sigma", "Mean Median", "Sigma Median", "Sample Size"
        ]
        gname, cname = chart_key
        # 樣本數 n 強制轉為 int 顯示
        samplesize_val = stats.get("samplesize", "")
        try:
            if samplesize_val != '' and samplesize_val is not None:
                samplesize_val = int(float(samplesize_val))
        except Exception:
            pass
        row_values = [
            abnormal_type,
            gname, cname, group_id,
            mean_index, sigma_index,
            stats.get("K", ""), stats.get("mean", ""), stats.get("sigma", ""),
            stats.get("mean_median", ""), stats.get("sigma_median", ""),
            samplesize_val
        ]

        info_table = QtWidgets.QTableWidget()
        info_table.setColumnCount(len(headers))
        info_table.setHorizontalHeaderLabels(headers)
        info_table.setRowCount(1)
        info_table.verticalHeader().setVisible(False)
        info_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)

        for j, value in enumerate(row_values):
            if j in [4,5,6,7,8,9,10]:  # Mean Index, Sigma Index, K, Mean, Sigma, Mean Median, Sigma Median
                try:
                    if value != 'Insufficient Data' and value != 'No Compare' and value != '' and value is not None:
                        value = float(value)
                        value = f"{value:.2f}"
                except Exception:
                    pass
            item = QtWidgets.QTableWidgetItem(str(value))
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            info_table.setItem(0, j, item)

        info_table.resizeColumnsToContents()
        info_table.setFixedHeight(info_table.horizontalHeader().height() + info_table.rowHeight(0) + 5)
        info_v_layout.addWidget(info_table)
        main_layout.addWidget(info_group)

        # --- 下方：圖表區塊 ---
        charts_container_widget = QtWidgets.QWidget()
        charts_layout = QtWidgets.QHBoxLayout(charts_container_widget)

        site_key = (chart_key[0], chart_key[1], str(group_id))
        if hasattr(self, 'chart_figures') and site_key in self.chart_figures:
            figures = self.chart_figures[site_key]
            
            if figures['scatter'] and figures['box']:
                scatter_pixmap = QtGui.QPixmap()
                scatter_pixmap.loadFromData(figures['scatter'])
                scatter_label = QtWidgets.QLabel()
                scatter_label.setPixmap(scatter_pixmap)
                scatter_label.setScaledContents(True)
                scatter_label.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Expanding)
                scatter_label.setMinimumSize(300, 200)

                box_pixmap = QtGui.QPixmap()
                box_pixmap.loadFromData(figures['box'])
                box_label = QtWidgets.QLabel()
                box_label.setPixmap(box_pixmap)
                box_label.setScaledContents(True)
                box_label.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Expanding)
                box_label.setMinimumSize(300, 200)

                charts_layout.addWidget(scatter_label)
                charts_layout.addWidget(box_label)
            else:
                charts_layout.addWidget(QtWidgets.QLabel("This item's chart was not generated due to insufficient data."))
        else:
            charts_layout.addWidget(QtWidgets.QLabel("Cannot find corresponding chart."))
        
        main_layout.addWidget(charts_container_widget)

        # 設定佈局伸展因子，讓圖表區域佔用更多空間
        main_layout.setStretchFactor(info_group, 0) # 數據表格高度固定
        main_layout.setStretchFactor(charts_container_widget, 1) # 圖表區域填滿剩餘空間

        # --- 關閉按鈕 ---
        button_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Close)
        button_box.rejected.connect(dialog.reject)
        main_layout.addWidget(button_box)

        dialog.exec()

    def _create_boxplots(self, grouped):
        """創建 SPC 圖和盒鬚圖，將 figure 物件保存在 self.chart_figures 中，不在 UI 上顯示。"""
        try:
            # 使用 Figure() 直接建立，不透過 pyplot 管理，避免 plt.close() 後物件狀態異常
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_agg import FigureCanvasAgg
            from matplotlib import cm
            import numpy as np
            from io import BytesIO
        except ImportError:
            print("[ERROR] Matplotlib is not installed.")
            return

        # 保存圖表與分組鍵的對應關係，用於後續的彈出視窗和 Excel 匯出
        self.chart_figures = {}

        # 為每個 (GroupName, ChartName) 組合創建圖表
        for (gname, cname), subdf in grouped:
            # 依 matching_group 字母順序排序（固定左到右順序）
            unique_groups = sorted(subdf["matching_group"].unique(), key=lambda x: str(x))
            labels = [str(mg) for mg in unique_groups]

            # 檢查是否有數據可供繪圖
            if subdf.empty or not any(len(grp["point_val"]) > 0 for _, grp in subdf.groupby("matching_group")):
                print(f"[WARNING] Skipping chart creation for {gname} - {cname} due to empty data.")
                for mg in unique_groups:
                    self.chart_figures[(gname, cname, str(mg))] = {'scatter': None, 'box': None}
                continue

            # 依排序後 unique_groups 組裝 box_data，確保顏色/label/資料一致
            box_data = [subdf[subdf["matching_group"] == mg]["point_val"].values for mg in unique_groups]
            group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count'])

            # 為不同的組設置顏色（固定顏色對應，不隨 focus 改變）
            colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))

            # 為每個 site 建立一組 per-site focus 圖
            # focus site：實心 + 正常大小；其他 site：同色但高度透明
            for focus_site in unique_groups:
                # 1. SPC 圖（使用獨立 Figure，不透過 pyplot 管理）
                scatter_fig = Figure(figsize=(7, 4.5))
                scatter_ax = scatter_fig.add_subplot(111)

                x_position = 0
                group_positions = []
                x_pos = 0
                for i, mg in enumerate(unique_groups):
                    group_data = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                    group_size = len(subdf[subdf["matching_group"] == mg])
                    group_positions.append(x_pos + group_size / 2 - 0.5)
                    x_pos += group_size

                    if not group_data.empty:
                        x_vals = np.arange(x_position, x_position + len(group_data))
                        y_vals = group_data["point_val"].values

                        is_focus = str(mg) == str(focus_site)
                        pt_alpha = 0.95 if is_focus else 0.18
                        pt_size  = 55  if is_focus else 25
                        ln_alpha = 0.6 if is_focus else 0.1
                        ln_width = 1.5 if is_focus else 0.7

                        scatter_ax.scatter(x_vals, y_vals, color=colors[i],
                                           alpha=pt_alpha, s=pt_size, label=f'{mg}', zorder=3)
                        scatter_ax.plot(x_vals, y_vals, color=colors[i],
                                        alpha=ln_alpha, linewidth=ln_width, zorder=2)

                        if i < len(unique_groups) - 1:
                            separator_x = x_position + len(group_data) - 0.5
                            scatter_ax.axvline(x=separator_x, color='gray',
                                               linestyle='-', alpha=0.3, zorder=1)

                        x_position += len(group_data)

                scatter_ax.set_title(f"SPC Chart: {gname} - {cname}  [Focus: {focus_site}]", fontsize=10)
                scatter_ax.set_xlabel("Sample Sequence (Grouped by Matching Group)")
                scatter_ax.set_ylabel("Point Value")
                scatter_ax.grid(True, linestyle='--', alpha=0.3, zorder=0)
                scatter_ax.set_xticks(group_positions)
                scatter_ax.set_xticklabels(labels, rotation=0, ha='center')
                scatter_ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
                scatter_fig.tight_layout()

                # 2. 盒鬚圖（focus site 實心，其他同色但透明）
                box_fig = Figure(figsize=(7, 4.5))
                box_ax = box_fig.add_subplot(111)
                if box_data:
                    bp = box_ax.boxplot(box_data, labels=labels, patch_artist=True, widths=0.6)
                    for patch, color, mg in zip(bp['boxes'], colors, unique_groups):
                        is_focus = str(mg) == str(focus_site)
                        patch.set_facecolor(color)
                        patch.set_alpha(0.95 if is_focus else 0.18)

                    legend_labels = [
                        f"{label}: μ={group_stats.loc[mg, 'mean']:.2f}, σ={group_stats.loc[mg, 'std']:.2f}, n={int(group_stats.loc[mg, 'count'])}"
                        for label, mg in zip(labels, unique_groups)
                    ]
                    box_ax.legend([bp["boxes"][i] for i in range(len(labels))],
                                  legend_labels, loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')

                box_ax.set_title(f"Boxplot: {gname} - {cname}  [Focus: {focus_site}]", fontsize=10)
                box_ax.set_xlabel("Matching Group")
                box_ax.set_ylabel("Point Value")
                box_ax.grid(True, linestyle='--', alpha=0.6)
                box_fig.subplots_adjust(right=0.7)
                box_fig.tight_layout()

                # 立即渲染為 PNG bytes，避免 Figure 物件跨次顯示時狀態污染
                FigureCanvasAgg(scatter_fig)
                buf_s = BytesIO()
                scatter_fig.savefig(buf_s, format='png', bbox_inches='tight', dpi=100)

                FigureCanvasAgg(box_fig)
                buf_b = BytesIO()
                box_fig.savefig(buf_b, format='png', bbox_inches='tight', dpi=100)

                key = (gname, cname, str(focus_site))
                self.chart_figures[key] = {'scatter': buf_s.getvalue(), 'box': buf_b.getvalue()}

    def _export_to_excel(self, all_results, source_path):
        """將分析結果匯出為 Excel 檔案，並在第一欄嵌入完整的盒鬚圖和散點圖。包含異常類型欄。"""
        try:
            # 檢查是否已安裝 openpyxl
            if openpyxl is None:
                QtWidgets.QMessageBox.warning(
                    self, "Missing Package", 
                    "Please install openpyxl to export Excel files.\nRun in terminal: pip install openpyxl"
                )
                self.status_label.setText(f"Analysis completed. Cannot export Excel: openpyxl package required.")
                return None

            # 嘗試導入所需的模組
            try:
                import matplotlib.pyplot as plt
                import numpy as np
                import io
                from PIL import Image
                import matplotlib.cm as cm
                from openpyxl.drawing.image import Image as XLImage
            except ImportError as e:
                QtWidgets.QMessageBox.warning(
                    self, "Missing Package", 
                    f"Additional packages required for embedding charts: {str(e)}\nPlease install the required packages."
                )
                print(f"[WARNING] Missing packages required for embedding charts: {e}")
                return None

            # 新增異常類型欄位，all_results: [is_abnormal, abnormal_type, ...]
            columns = [
                "Need_matching", "AbnormalType", "GroupName", "ChartName", "matching_group", "mean_matching_index", 
                "sigma_matching_index", "K", "mean", "sigma", "mean_median", "sigma_median", "samplesize", "characteristic"
            ]
            df = pd.DataFrame(all_results, columns=columns)

            # 打印資料框資訊以確認結構
            print(f"DataFrame info: {df.shape}")
            print(f"DataFrame columns: {df.columns.tolist()}")
            print(f"First row: {df.iloc[0].tolist() if len(df) > 0 else 'No data'}")

            # 生成輸出檔案路徑（與輸入檔案相同目錄）
            dir_path = os.path.dirname(source_path)
            file_name = os.path.splitext(os.path.basename(source_path))[0]
            output_path = os.path.join(dir_path, f"{file_name}_matching_results.xlsx")

            # 創建臨時目錄用於保存圖片
            import tempfile
            temp_dir = tempfile.mkdtemp()
            print(f"[INFO] Creating temporary directory: {temp_dir}")

            # 先在 DataFrame 前添加兩個空白欄位，分別用於SPC圖和盒鬚圖
            df.insert(0, "SPC_Chart", "")    # 第一欄：SPC圖
            df.insert(1, "BoxPlot", "")      # 第二欄：盒鬚圖

            # 創建 Excel 文件
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            df.to_excel(writer, sheet_name='Tool Matching Results', index=False)

            # 獲取工作表
            workbook = writer.book
            worksheet = writer.sheets['Tool Matching Results']

            # 設定標題列格式
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="344CB7", end_color="344CB7", fill_type="solid")
            header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

            # 設置標題列格式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 增加圖表欄寬度以容納圖片
            worksheet.column_dimensions['A'].width = 70  # 第一欄：SPC圖
            worksheet.column_dimensions['B'].width = 70  # 第二欄：盒鬚圖

            # 設定異常行的格式
            abnormal_fill = openpyxl.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            # 定義圖表在 Excel 中顯示的尺寸 (單位：像素)
            img_display_width, img_display_height = 450, 250

            # 檢查是否有可用的圖表數據
            has_chart_figures = hasattr(self, 'chart_figures') and self.chart_figures
            if not has_chart_figures:
                print("[WARNING] No chart data available, will use simple status indicator images")

            # 從第二行開始遍歷（跳過標題行）
            for row_idx, row in enumerate(df.iterrows(), start=2):
                _, row_data = row

                # 檢查Need_matching欄位是否為True
                is_abnormal = row_data["Need_matching"]

                if is_abnormal:
                    # 將整行設為淺紅色
                    for cell in worksheet[row_idx]:
                        cell.fill = abnormal_fill

                # 創建並嵌入圖表到第一欄
                try:
                    # 獲取關鍵數據
                    group_name = str(row_data["GroupName"])
                    chart_name = str(row_data["ChartName"])
                    group_id = str(row_data["matching_group"])
                    mean_index = row_data["mean_matching_index"]
                    sigma_index = row_data["sigma_matching_index"]
                    k_value = row_data["K"]

                    # 檢查是否資料不足
                    is_data_insufficient = (mean_index == 'Insufficient Data' or sigma_index == 'Insufficient Data' or k_value == 'No Compare')

                    # 嘗試使用完整的SPC圖和盒鬚圖（以 group_id 為第三維 key 取得 per-site focus 圖）
                    chart_key = (group_name, chart_name, group_id)
                    if has_chart_figures and chart_key in self.chart_figures:
                        # 存在完整的分析圖表，使用實際的SPC圖和盒鬚圖
                        chart_data = self.chart_figures[chart_key]

                        # 1. 處理SPC圖 (放在第一欄)
                        try:
                            scatter_bytes = chart_data['scatter']
                            temp_scatter_path = os.path.join(temp_dir, f"spc_{group_name}_{chart_name}_{row_idx}.png")
                            with open(temp_scatter_path, 'wb') as _f:
                                _f.write(scatter_bytes)
                            try:
                                scatter_img = XLImage(temp_scatter_path)
                                scatter_img.width = img_display_width
                                scatter_img.height = img_display_height
                                scatter_position = f"A{row_idx}"
                                worksheet.add_image(scatter_img, scatter_position)
                                print(f"[INFO] SPC chart added to cell: {scatter_position}")
                            except Exception as img_e:
                                print(f"[ERROR] Failed to add SPC chart to Excel: {img_e}")
                                worksheet.cell(row=row_idx, column=1).value = "SPC chart failed to load"
                        except Exception as scatter_e:
                            print(f"[ERROR] Error occurred while processing SPC chart: {scatter_e}")
                            import traceback
                            traceback.print_exc()
                            worksheet.cell(row=row_idx, column=1).value = "SPC chart generation failed"

                        # 2. 處理盒鬚圖 (放在第二欄)
                        try:
                            box_bytes = chart_data['box']
                            temp_box_path = os.path.join(temp_dir, f"box_{group_name}_{chart_name}_{row_idx}.png")
                            with open(temp_box_path, 'wb') as _f:
                                _f.write(box_bytes)
                            try:
                                box_img = XLImage(temp_box_path)
                                box_img.width = img_display_width
                                box_img.height = img_display_height
                                box_position = f"B{row_idx}"
                                worksheet.add_image(box_img, box_position)
                                print(f"[INFO] Boxplot added to cell: {box_position}")
                            except Exception as img_e:
                                print(f"[ERROR] Failed to add boxplot to Excel: {img_e}")
                                worksheet.cell(row=row_idx, column=2).value = "Boxplot failed to load"
                        except Exception as box_e:
                            print(f"[ERROR] Error occurred while processing boxplot: {box_e}")
                            import traceback
                            traceback.print_exc()
                            worksheet.cell(row=row_idx, column=2).value = "Boxplot generation failed"

                    else:
                        # 沒有找到匹配的圖表，使用狀態指示器
                        print(f"[INFO] No analysis chart found for {group_name}/{chart_name}, using status indicator")
                        fig, ax = plt.subplots(figsize=(6, 4), dpi=100)
                        title = f"{group_name}\n{chart_name}\nGroup: {group_id}"
                        ax.set_title(title, fontsize=12)
                        if is_data_insufficient:
                            circle = plt.Circle((0.5, 0.5), 0.3, color='yellow', alpha=0.6, edgecolor='goldenrod', linewidth=2)
                            ax.add_patch(circle)
                            ax.text(0.5, 0.5, "Insufficient Data", ha='center', va='center', fontsize=14, color='black')
                            status_text = "Insufficient data for analysis"
                        elif is_abnormal:
                            circle = plt.Circle((0.5, 0.5), 0.3, color='red', alpha=0.6, edgecolor='darkred', linewidth=2)
                            ax.add_patch(circle)
                            ax.text(0.5, 0.5, "Need Matching", ha='center', va='center', fontsize=14, color='white', fontweight='bold')
                            status_text = f"Mean Index: {mean_index}, Sigma Index: {sigma_index}, K: {k_value}"
                        else:
                            circle = plt.Circle((0.5, 0.5), 0.3, color='green', alpha=0.6, edgecolor='darkgreen', linewidth=2)
                            ax.add_patch(circle)
                            ax.text(0.5, 0.5, "Normal", ha='center', va='center', fontsize=14, color='white', fontweight='bold')
                            status_text = f"Mean Index: {mean_index}, Sigma Index: {sigma_index}, K: {k_value}"
                        ax.text(0.5, 0.2, status_text, ha='center', va='center', fontsize=10, 
                               bbox=dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.8))
                        ax.set_xticks([])
                        ax.set_yticks([])
                        ax.set_xlim(0, 1)
                        ax.set_ylim(0, 1)
                        ax.set_aspect('equal')
                        temp_img_path = os.path.join(temp_dir, f"status_chart_{row_idx}.png")
                        plt.savefig(temp_img_path, format='png', bbox_inches='tight', transparent=True, dpi=300)
                        plt.close(fig)
                        try:
                            # 使用 xlsxwriter 寫法 (insert_image) 取代 openpyxl 的 add_image
                            # 需先取得 xlsxwriter 的 worksheet 物件
                            # 但目前本程式是用 openpyxl，無法直接用 insert_image
                            # 所以這裡僅說明：如果你要用 insert_image，必須用 xlsxwriter 建立 writer
                            # 下面是 xlsxwriter 寫法範例：
                            # worksheet.insert_image(row_idx-1, 0, temp_img_path, {'x_scale': 1, 'y_scale': 1, 'x_offset': 0, 'y_offset': 0, 'object_position': 1})
                            # worksheet.insert_image(row_idx-1, 1, temp_img_path, {'x_scale': 1, 'y_scale': 1, 'x_offset': 0, 'y_offset': 0, 'object_position': 1})
                            # 但 openpyxl 不支援 insert_image，僅支援 add_image
                            # 若要完全改用 xlsxwriter，需重構整個 Excel 輸出流程。
                            # 這裡保留原本 openpyxl add_image 寫法，僅註明差異。
                            img1 = XLImage(temp_img_path)
                            img1.width = img_display_width
                            img1.height = img_display_height
                            cell_position_1 = f"A{row_idx}"
                            worksheet.add_image(img1, cell_position_1)
                            img2 = XLImage(temp_img_path)
                            img2.width = img_display_width
                            img2.height = img_display_height
                            cell_position_2 = f"B{row_idx}"
                            worksheet.add_image(img2, cell_position_2)
                            print(f"[INFO] Status chart added to cells: {cell_position_1} and {cell_position_2}")
                        except Exception as img_e:
                            print(f"[ERROR] Failed to add image to Excel: {img_e}")
                            worksheet.cell(row=row_idx, column=1).value = "Image failed to load"
                            worksheet.cell(row=row_idx, column=2).value = "Image failed to load"

                except Exception as img_e:
                    print(f"[ERROR] Error occurred while adding chart to row {row_idx}: {img_e}")
                    import traceback
                    traceback.print_exc()
                    worksheet.cell(row=row_idx, column=1).value = "Image generation failed"

            # 調整行高以適應圖表
            for i in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[i].height = 190

            # 調整其他列寬
            for col_idx, column in enumerate(worksheet.columns, start=1):
                if col_idx <= 2:  # 跳過圖表列 A 和 B，已手動設置寬度
                    continue
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 4)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 儲存 Excel 檔案
            try:
                writer.close()
                print(f"[INFO] Excel file saved to: {output_path}")
            except Exception as save_e:
                print(f"[ERROR] Failed to save Excel file: {save_e}")
                import traceback
                traceback.print_exc()
            finally:
                try:
                    import shutil
                    shutil.rmtree(temp_dir)
                    print(f"[INFO] Temporary directory cleaned: {temp_dir}")
                except Exception as e:
                    print(f"[WARNING] Unable to clean temporary directory: {temp_dir}, error: {e}")

            # self.status_label.setText(f"Analysis completed. Results exported to: {output_path}")
            print(f"Analysis completed. Results exported to: {output_path}")
            return output_path
        except Exception as e:
            # self.status_label.setText(f"Excel export failed: {e}")
            print(f"Excel export failed: {e}")
            import traceback
            traceback.print_exc()
            return None


# ==================================================
# 無 UI 版本的核心分析函數，供 FastAPI 使用
# ==================================================

def _legacy_analyze_tool_matching_data_v1(df, config):
    """
    無 UI 版本的 Tool Matching 分析（舊版，已由批次模式取代）
    
    Args:
        df (pd.DataFrame): 輸入的 DataFrame
        config (dict): 分析配置參數，包含：
            - mean_index_threshold: float, Mean Index 閾值
            - sigma_index_threshold: float, Sigma Index 閾值
            - use_statistical_test: bool, 是否使用統計檢定
            - statistical_method: str, 統計檢定方法 ("unpaired", "paired")
            - alpha_level: float, 顯著水準
            - fill_sample_size: int, 補滿樣本數
            - filter_mode: str, 資料篩選模式 ("all_data", "specified_date", "latest_data")
            - base_date: str, 基準日期 (YYYY-MM-DD)
    
    Returns:
        dict: 包含分析結果的字典
            {
                "summary": {"total_groups": int, "abnormal_groups": int},
                "results": [list of analysis results],
                "excel_output": str or None
            }
    """
    # 檢查必要欄位
    required_cols = ["GroupName", "ChartName", "matching_group", "point_val", "characteristic", "point_time"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")
    
    # 轉換 point_time 為 datetime
    df = df.copy()
    df["point_time"] = pd.to_datetime(df["point_time"])
    
    filter_mode = config.get("filter_mode", "all_data")
    base_date = config.get("base_date")
    fill_num = config.get("fill_sample_size", 5)
    
    results = []
    
    if filter_mode == "all_data":
        # 全算模式
        grouped = df.groupby(["GroupName", "ChartName"])
        for (gname, cname), subdf in grouped:
            characteristic = subdf["characteristic"].dropna().unique()
            if len(characteristic) != 1:
                continue
            characteristic = characteristic[0]
            
            group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count']).reset_index()
            n_groups = len(group_stats)
            
            # 檢查有效群組數量（樣本數 >= 5）
            valid_groups = group_stats[group_stats['count'] >= 5]
            n_valid_groups = len(valid_groups)
            
            if n_valid_groups == 2 and n_groups > 2:
                # 只有兩個群組有效，使用兩群比較
                _analyze_two_groups_headless(valid_groups, gname, cname, characteristic, results, subdf, config)
            elif n_groups == 2:
                # 總共只有兩群
                _analyze_two_groups_headless(group_stats, gname, cname, characteristic, results, subdf, config)
            else:
                # 多群比較
                _analyze_multiple_groups_headless(subdf, group_stats, gname, cname, characteristic, results, config)
                
    elif filter_mode == "specified_date":
        # 指定日期模式
        if not base_date:
            raise ValueError("Base date is required for specified_date mode")
        base_date = pd.to_datetime(base_date)
        grouped = df.groupby(["GroupName", "ChartName"])
        
        for (gname, cname), subdf in grouped:
            characteristic = subdf["characteristic"].dropna().unique()
            if len(characteristic) != 1:
                continue
            characteristic = characteristic[0]
            
            mean_end = base_date
            mean_start = mean_end - pd.DateOffset(days=30)
            week_start = mean_end - pd.Timedelta(days=7)
            MIN_FILL = 5
            # 先抓 30 天初始區間（mean 與 sigma 共用同一資料集）
            mean_df = subdf[(subdf["point_time"] > mean_start) & (subdf["point_time"] <= mean_end)].copy()
            sigma_df = mean_df.copy()
            # 三段式補點邏輯
            for mg in subdf["matching_group"].unique():
                mg_week = subdf[
                    (subdf["matching_group"] == mg) &
                    (subdf["point_time"] > week_start) &
                    (subdf["point_time"] <= mean_end)
                ]
                if len(mg_week) == 0:
                    mean_df = mean_df[mean_df["matching_group"] != mg]
                    sigma_df = sigma_df[sigma_df["matching_group"] != mg]
                    continue
                mg_30d = mean_df[mean_df["matching_group"] == mg]
                if len(mg_30d) < MIN_FILL:
                    mg_fill = (
                        subdf[subdf["matching_group"] == mg]
                        .sort_values("point_time", ascending=False)
                        .head(MIN_FILL)
                    )
                    mean_df = pd.concat([mean_df[mean_df["matching_group"] != mg], mg_fill]).drop_duplicates()
                    sigma_df = pd.concat([sigma_df[sigma_df["matching_group"] != mg], mg_fill]).drop_duplicates()
            
            mean_stats = mean_df.groupby("matching_group")["point_val"].agg(['mean', 'count']).reset_index()
            sigma_stats = sigma_df.groupby("matching_group")["point_val"].agg(['std']).reset_index()
            group_stats = pd.merge(mean_stats, sigma_stats, on="matching_group", how="outer")
            group_stats = group_stats.fillna({"mean": 0, "std": 0, "count": 0})
            
            n_groups = len(group_stats)
            valid_groups = group_stats[group_stats['count'] >= 5]
            n_valid_groups = len(valid_groups)
            
            if n_valid_groups == 2 and n_groups > 2:
                _analyze_two_groups_time_headless(mean_df, sigma_df, valid_groups, gname, cname, characteristic, results, config)
            elif n_groups == 2:
                _analyze_two_groups_time_headless(mean_df, sigma_df, group_stats, gname, cname, characteristic, results, config)
            else:
                _analyze_multiple_groups_time_headless(mean_df, sigma_df, group_stats, gname, cname, characteristic, results, config)
                
    elif filter_mode == "latest_data":
        # 最新進點模式
        grouped = df.groupby(["GroupName", "ChartName"])
        
        for (gname, cname), subdf in grouped:
            characteristic = subdf["characteristic"].dropna().unique()
            if len(characteristic) != 1:
                continue
            characteristic = characteristic[0]
            
            latest_time = subdf["point_time"].max()
            mean_end = latest_time
            mean_start = mean_end - pd.DateOffset(days=30)
            week_start = mean_end - pd.Timedelta(days=7)
            MIN_FILL = 5
            # 先抓 30 天初始區間（mean 與 sigma 共用同一資料集）
            mean_df = subdf[(subdf["point_time"] > mean_start) & (subdf["point_time"] <= mean_end)].copy()
            sigma_df = mean_df.copy()
            # 三段式補點邏輯
            for mg in subdf["matching_group"].unique():
                mg_week = subdf[
                    (subdf["matching_group"] == mg) &
                    (subdf["point_time"] > week_start) &
                    (subdf["point_time"] <= mean_end)
                ]
                if len(mg_week) == 0:
                    mean_df = mean_df[mean_df["matching_group"] != mg]
                    sigma_df = sigma_df[sigma_df["matching_group"] != mg]
                    continue
                mg_30d = mean_df[mean_df["matching_group"] == mg]
                if len(mg_30d) < MIN_FILL:
                    mg_fill = (
                        subdf[subdf["matching_group"] == mg]
                        .sort_values("point_time", ascending=False)
                        .head(MIN_FILL)
                    )
                    mean_df = pd.concat([mean_df[mean_df["matching_group"] != mg], mg_fill]).drop_duplicates()
                    sigma_df = pd.concat([sigma_df[sigma_df["matching_group"] != mg], mg_fill]).drop_duplicates()
            
            mean_stats = mean_df.groupby("matching_group")["point_val"].agg(['mean', 'count']).reset_index()
            sigma_stats = sigma_df.groupby("matching_group")["point_val"].agg(['std']).reset_index()
            group_stats = pd.merge(mean_stats, sigma_stats, on="matching_group", how="outer")
            group_stats = group_stats.fillna({"mean": 0, "std": 0, "count": 0})
            
            n_groups = len(group_stats)
            valid_groups = group_stats[group_stats['count'] >= 5]
            n_valid_groups = len(valid_groups)
            
            if n_valid_groups == 2 and n_groups > 2:
                _analyze_two_groups_time_headless(mean_df, sigma_df, valid_groups, gname, cname, characteristic, results, config)
            elif n_groups == 2:
                _analyze_two_groups_time_headless(mean_df, sigma_df, group_stats, gname, cname, characteristic, results, config)
            else:
                _analyze_multiple_groups_time_headless(mean_df, sigma_df, group_stats, gname, cname, characteristic, results, config)
    
    # 統計異常項目
    abnormal_count = sum(1 for result in results if result[0])  # 第一欄是 is_abnormal
    
    return {
        "summary": {
            "total_groups": len(results),
            "abnormal_groups": abnormal_count
        },
        "results": results,
        "excel_output": None  # 可以後續實作 Excel 匯出
    }


def _analyze_two_groups_headless(group_stats, gname, cname, characteristic, results, raw_data, config):
    """無 UI 版本的兩群分析（與 PyQt 版本完全一致）"""
    row1 = group_stats.iloc[0]
    row2 = group_stats.iloc[1]

    group1 = row1["matching_group"]
    group2 = row2["matching_group"]
    mean1, std1, n1 = row1["mean"], row1["std"], row1["count"]
    mean2, std2, n2 = row2["mean"], row2["std"], row2["count"]

    min_sigma = min(std1, std2)
    use_stats = config.get("use_statistical_test", False)
    
    if use_stats and raw_data is not None and SCIPY_AVAILABLE:
        # 統計檢定
        group1_data = raw_data[raw_data["matching_group"] == group1]["point_val"].tolist()
        group2_data = raw_data[raw_data["matching_group"] == group2]["point_val"].tolist()
        
        if len(group1_data) >= 2 and len(group2_data) >= 2:
            method = config.get("statistical_method", "unpaired")
            alpha = config.get("alpha_level", 0.05)
            
            # 統一使用 perform_statistical_test 函數
            data_groups = {group1: group1_data, group2: group2_data}
            stats_result = perform_statistical_test(data_groups, method, alpha)
            
            if "error" in stats_result:
                error_msg = stats_result['error']
                mean_status_1 = f"Statistical test failed: {error_msg}"
                mean_status_2 = f"Statistical test failed: {error_msg}"
            else:
                p_val = stats_result.get("p_value", 1.0)
                if stats_result.get("significant", False):
                    mean_status_1 = f"Significant (p={p_val:.4f})"
                    mean_status_2 = f"Significant (p={p_val:.4f})"
                else:
                    mean_status_1 = f"No Significant (p={p_val:.4f})"
                    mean_status_2 = f"No Significant (p={p_val:.4f})"
        else:
            mean_status_1 = "Statistical test - Insufficient data"
            mean_status_2 = "Statistical test - Insufficient data"
    else:
        # 使用原有的 Mean Index 計算（考慮方向性）
        mean_index_1 = calculate_mean_index_headless(mean1, mean2, min_sigma, characteristic)
        mean_index_2 = calculate_mean_index_headless(mean2, mean1, min_sigma, characteristic)
        mean_status_1 = round(mean_index_1, 2)
        mean_status_2 = round(mean_index_2, 2)

    # 計算 sigma index (保持原邏輯)
    if min_sigma > 0:
        sigma_index_1 = std1 / min_sigma
        sigma_index_2 = std2 / min_sigma
    else:
        sigma_index_1 = 0 if std1 == std2 else float('inf')
        sigma_index_2 = 0 if std1 == std2 else float('inf')

    # 統一格式：第4欄都用 'group_all'，與多群分析一致
    # mean_median, sigma_median 欄位（兩組時用 mean2, min_sigma 或 mean1, min_sigma）

    if n1 < 5 or n2 < 5:
        results.append([
            gname, cname, group1, 'group_all',
            'Insufficient Data', 'Insufficient Data',
            get_k_value_headless(n1), mean1, std1,
            mean2, min_sigma, n1, characteristic
        ])
        results.append([
            gname, cname, group2, 'group_all',
            'Insufficient Data', 'Insufficient Data',
            get_k_value_headless(n2), mean2, std2,
            mean1, min_sigma, n2, characteristic
        ])
        return

    k1 = get_k_value_headless(n1)
    k2 = get_k_value_headless(n2)

    if k1 == "No Compare":
        results.append([
            gname, cname, group1, 'group_all',
            'Insufficient Data', 'Insufficient Data',
            'No Compare', round(mean1, 2), round(std1, 2),
            round(mean2, 2), round(min_sigma, 2), n1, characteristic
        ])
    else:
        results.append([
            gname, cname, group1, 'group_all',
            mean_status_1, round(sigma_index_1, 2),
            round(k1, 2), round(mean1, 2), round(std1, 2),
            round(mean2, 2), round(min_sigma, 2), n1, characteristic
        ])

    if k2 == "No Compare":
        results.append([
            gname, cname, group2, 'group_all',
            'Insufficient Data', 'Insufficient Data',
            'No Compare', round(mean2, 2), round(std2, 2),
            round(mean1, 2), round(min_sigma, 2), n2, characteristic
        ])
    else:
        results.append([
            gname, cname, group2, 'group_all',
            mean_status_2, round(sigma_index_2, 2),
            round(k2, 2), round(mean2, 2), round(std2, 2),
            round(mean1, 2), round(min_sigma, 2), n2, characteristic
        ])


def _analyze_multiple_groups_headless(subdf, group_stats, gname, cname, characteristic, results, config):
    """無 UI 版本的多群分析（與 PyQt 版本完全一致）"""
    # 只納入樣本數 >= 5 的 group 計算 median
    valid_stats = group_stats[group_stats['count'] >= 5]
    if valid_stats.shape[0] <= 1:
        # 只有一個有效群組，全部標記資料不足
        for i, row in group_stats.iterrows():
            group = row["matching_group"]
            mean = row["mean"]
            std = row["std"]
            n = row["count"]
            results.append([
                gname, cname, group, "group_all",
                'Insufficient Data', 'Insufficient Data', 
                get_k_value_headless(n), mean, std, 
                '-', '-', n, characteristic
            ])
        return

    mean_median = valid_stats['mean'].median() if not valid_stats.empty else 0
    median_sigma = valid_stats['std'].median() if not valid_stats.empty else 0

    # 檢查是否使用統計檢定
    use_stats = config.get("use_statistical_test", False)
    stats_result = None
    
    if use_stats and SCIPY_AVAILABLE and len(valid_stats) >= 3:
        # 準備 ANOVA 的數據
        data_groups = {}
        for _, row in valid_stats.iterrows():
            group_name = row["matching_group"]
            group_data = subdf[subdf["matching_group"] == group_name]["point_val"].tolist()
            if len(group_data) >= 2:
                data_groups[group_name] = group_data
        
        if len(data_groups) >= 3:
            alpha = config.get("alpha_level", 0.05)
            stats_result = perform_statistical_test(data_groups, "unpaired", alpha)

    for i, row in group_stats.iterrows():
        group = row["matching_group"]
        mean = row["mean"]
        std = row["std"]
        n = row["count"]

        # 計算 mean matching index（考慮方向性）
        if n < 5:  # 樣本數不足5個，不進行比較
            results.append([
                gname, cname, group, "group_all",
                'Insufficient Data', 'Insufficient Data', 
                get_k_value_headless(n), mean, std, 
                mean_median, median_sigma, n, characteristic
            ])
            continue

        # 決定 mean matching 狀態
        if use_stats and stats_result and "error" not in stats_result:
            # 使用統計檢定結果
            if stats_result.get("significant", False):
                # ANOVA 顯著，檢查事後檢定結果
                mean_status = "ANOVA Significant"
                if stats_result.get("post_hoc") and "significant_pairs" in stats_result["post_hoc"]:
                    # 檢查這個群組是否在顯著配對中
                    significant_pairs = stats_result["post_hoc"]["significant_pairs"]
                    group_in_significant = any(
                        pair["group1"] == group or pair["group2"] == group 
                        for pair in significant_pairs
                    )
                    if group_in_significant:
                        # 找出與此群組有顯著差異的其他群組
                        different_groups = []
                        for pair in significant_pairs:
                            if pair["group1"] == group:
                                different_groups.append(pair["group2"])
                            elif pair["group2"] == group:
                                different_groups.append(pair["group1"])
                        if different_groups:
                            mean_status = f"Significant vs {','.join(different_groups)}"
                        else:
                            mean_status = "Participate in Significant"
                    else:
                        mean_status = "No Significant"
                else:
                    mean_status = f"ANOVA Significant(p={stats_result.get('p_value', 0):.4f})"
            else:
                mean_status = f"No Significant(p={stats_result.get('p_value', 1):.4f})"
        else:
            # 使用原有的 Mean Index 計算
            if median_sigma > 0:
                if characteristic == 'Bigger':
                    mean_index = (mean_median - mean) / median_sigma
                elif characteristic in ['Smaller', 'Sigma']:
                    mean_index = (mean - mean_median) / median_sigma
                else:
                    mean_index = abs(mean - mean_median) / median_sigma
            else:
                # 分母為零時，判斷所有 mean 是否相等
                all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    mean_index = 0
                else:
                    mean_index = float('inf')
            
            mean_status = round(mean_index, 2)

        # 計算 sigma index (保持原邏輯)
        if median_sigma > 0:
            sigma_index = std / median_sigma
        else:
            all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
            if len(set([round(m, 8) for m in all_means])) == 1:
                sigma_index = 0
            else:
                sigma_index = float('inf')

        K = get_k_value_headless(n)

        # 檢查 K 值是否為字串 "No Compare"
        if K == "No Compare":
            # 樣本數不足，使用 "Insufficient Data" 標記
            results.append([
                gname, cname, group, "group_all",
                'Insufficient Data', 'Insufficient Data', 
                'No Compare', round(mean, 2), round(std, 2), 
                round(mean_median, 2), round(median_sigma, 2), n, characteristic
            ])
        else:
            # 正常比較情況
            results.append([
                gname, cname, group, "group_all",
                mean_status, round(sigma_index, 2), 
                round(K, 2), round(mean, 2), round(std, 2), 
                round(mean_median, 2), round(median_sigma, 2), n, characteristic
            ])


def _analyze_multiple_groups_time_headless(mean_df, sigma_df, group_stats, gname, cname, characteristic, results, config):
    """
    無 UI 版本的多組分析（時間模式）
    - mean, std, count: 來自 mean_df（30 天 window，三段式補點）
    - median_sigma: 來自 sigma_df（同 mean_df，30 天 window，三段式補點）
    """
    # 只納入樣本數 >= 5 的 group 計算 median
    valid_mean_df = mean_df.groupby("matching_group").filter(lambda x: len(x) >= 5)
    sigma_by_group = sigma_df.groupby("matching_group")["point_val"].std()
    valid_groups = group_stats[group_stats['count'] >= 5]['matching_group']
    valid_sigma = sigma_by_group[valid_groups] if not valid_groups.empty else pd.Series(dtype=float)
    
    # 防呆：如果有效 group 只有一個，全部標記資料不足
    if len(valid_groups) <= 1:
        for i, row in group_stats.iterrows():
            group = row["matching_group"]
            mean = row["mean"]
            std = row["std"]
            n = row["count"]
            results.append([
                gname, cname, group, "group_all",
                'Insufficient Data', 'Insufficient Data', 
                get_k_value_headless(n), mean, std, 
                '-', '-', n, characteristic
            ])
        return
    
    mean_median = valid_mean_df["point_val"].median() if not valid_mean_df.empty else 0
    median_sigma = valid_sigma.median() if not valid_sigma.empty else 0
    
    for i, row in group_stats.iterrows():
        group = row["matching_group"]
        mean = row["mean"]
        std = row["std"]  # 這是來自 mean_df（一個月 window）
        n = row["count"]
        
        if n < 5:
            results.append([
                gname, cname, group, "group_all",
                'Insufficient Data', 'Insufficient Data', 
                get_k_value_headless(n), mean, std, 
                mean_median, median_sigma, n, characteristic
            ])
            continue
        
        if median_sigma > 0:
            if characteristic == 'Bigger':
                mean_index = (mean_median - mean) / median_sigma
            elif characteristic in ['Smaller', 'Sigma']:
                mean_index = (mean - mean_median) / median_sigma
            else:
                mean_index = abs(mean - mean_median) / median_sigma
            sigma_index = std / median_sigma
        else:
            # 分母為零時，判斷所有 mean 是否相等
            all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
            if len(set([round(m, 8) for m in all_means])) == 1:
                mean_index = 0
                sigma_index = 0
            else:
                mean_index = float('inf')
                sigma_index = float('inf')
        
        K = get_k_value_headless(n)
        if K == "No Compare":
            results.append([
                gname, cname, group, "group_all",
                'Insufficient Data', 'Insufficient Data', 
                'No Compare', round(mean, 2), round(std, 2), 
                round(mean_median, 2), round(median_sigma, 2), n, characteristic
            ])
        else:
            results.append([
                gname, cname, group, "group_all",
                round(mean_index, 2), round(sigma_index, 2), 
                round(K, 2), round(mean, 2), round(std, 2), 
                round(mean_median, 2), round(median_sigma, 2), n, characteristic
            ])


def _analyze_two_groups_time_headless(mean_df, sigma_df, group_stats, gname, cname, characteristic, results, config):
    """無 UI 版本的時間模式兩群分析"""
    _analyze_two_groups_headless(group_stats, gname, cname, characteristic, results, mean_df, config)


def perform_statistical_test(data_groups, method="auto", alpha=0.05):
    """
    執行統計檢定（無 UI 版本）
    
    Parameters:
    - data_groups: dict, key 為群組名稱，value 為數據 list
    - method: str, "unpaired", "paired", "auto"
    - alpha: float, 顯著水準
    
    Returns:
    - dict: 包含檢定結果的字典
    """
    if not SCIPY_AVAILABLE:
        return {"error": "SciPy package not installed, cannot perform statistical tests"}
    
    group_names = list(data_groups.keys())
    group_data = list(data_groups.values())
    
    # 過濾掉空群組
    valid_groups = [(name, data) for name, data in zip(group_names, group_data) if len(data) >= 2]
    
    if len(valid_groups) < 2:
        return {"error": "Insufficient number of valid groups, cannot perform statistical tests"}
    
    valid_names, valid_data = zip(*valid_groups)
    
    result = {
        "method": method,
        "alpha": alpha,
        "groups": valid_names,
        "n_groups": len(valid_groups),
        "significant": False,
        "p_value": None,
        "test_statistic": None,
        "post_hoc": None
    }
    
    try:
        if len(valid_groups) == 2:
            # 兩群比較
            group1_data, group2_data = valid_data[0], valid_data[1]
            
            if method == "paired":
                # 檢查是否能配對（樣本數相等）
                if len(group1_data) == len(group2_data):
                    # 配對 t 檢定
                    stat, p_val = ttest_rel(group1_data, group2_data)
                    result["method"] = "Paired t-test"
                else:
                    return {
                        "error": f"配對 t 檢定需要兩組樣本數相等。目前：{valid_names[0]}={len(group1_data)} vs {valid_names[1]}={len(group2_data)}"
                    }
            elif method == "unpaired":
                # 非配對 t 檢定
                stat, p_val = ttest_ind(group1_data, group2_data)
                result["method"] = "Unpaired t-test"
            else:
                # 預設使用非配對 t 檢定
                stat, p_val = ttest_ind(group1_data, group2_data)
                result["method"] = "Unpaired t-test"
            
            result["test_statistic"] = stat
            result["p_value"] = p_val
            result["significant"] = p_val < alpha
            
        else:
            # 三群以上：使用 ANOVA
            stat, p_val = f_oneway(*valid_data)
            result["method"] = "One-way ANOVA"
            result["test_statistic"] = stat
            result["p_value"] = p_val
            result["significant"] = p_val < alpha
            
            # 如果 ANOVA 顯著且有 Tukey HSD 可用，進行事後檢定
            if result["significant"] and TUKEY_AVAILABLE:
                # 準備事後檢定數據（省略具體實作）
                result["post_hoc"] = "Tukey HSD available but not implemented"
                
    except Exception as e:
        result["error"] = f"Statistical test failed: {str(e)}"
    
    return result


def calculate_mean_index_headless(mean1, mean2, min_sigma, characteristic):
    """計算 mean matching index，考慮方向性（與 PyQt 版本完全一致）"""
    if min_sigma <= 0:
        # 分母為零時的處理
        if mean1 == mean2:
            return 0
        else:
            return float('inf')
    
    if characteristic == 'Bigger':  # Bigger is better
        return (mean2 - mean1) / min_sigma
    elif characteristic in ['Smaller', 'Sigma']:  # Smaller is better, Sigma 與 Smaller 邏輯相同
        return (mean1 - mean2) / min_sigma
    else:  # Nominal
        return abs(mean1 - mean2) / min_sigma


def get_k_value_headless(n):
    """根據樣本數量 n 返回 K 值（與 PyQt 版本完全一致）"""
    if n <= 4:  # 樣本數量太少，不進行比較
        return "No Compare"  # 返回特殊標記，表示不進行比較
    elif 5 <= n <= 10:
        return 1.73
    elif 11 <= n <= 120:
        return 1.414
    else:
        return 1.15


def _legacy_analyze_tool_matching_data_v2(df, config=None):
    """分析 Tool Matching 資料（舊版，已由批次模式取代）"""
    try:
        config = config or {}
        
        # 標準化欄位名稱
        if 'GroupName' in df.columns and 'gname' not in df.columns:
            df = df.rename(columns={'GroupName': 'gname'})
        if 'ChartName' in df.columns and 'cname' not in df.columns:
            df = df.rename(columns={'ChartName': 'cname'})
        
        # 檢查必要欄位
        required_columns = ['gname', 'cname', 'matching_group', 'point_val']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            # 嘗試其他可能的欄位名稱
            alt_mapping = {
                'gname': ['GroupName', 'Group Name', 'group_name'],
                'cname': ['ChartName', 'Chart Name', 'chart_name', 'ITEM NAME'],
                'matching_group': ['MatchingGroup', 'Matching Group', 'matching_grp'],
                'point_val': ['PointVal', 'Point Val', 'value', 'val']
            }
            
            for missing_col in missing_columns:
                found = False
                if missing_col in alt_mapping:
                    for alt_name in alt_mapping[missing_col]:
                        if alt_name in df.columns:
                            df = df.rename(columns={alt_name: missing_col})
                            found = True
                            break
                if not found:
                    raise ValueError(f"Missing required column: {missing_col}. Available columns: {list(df.columns)}")
        
        # 檢查 'time_diff_days' 欄位並計算
        if 'time_diff_days' not in df.columns:
            df['time_diff_days'] = 0
        
        results = []
        processed_count = 0
        
        # 取得 chart info 如果有的話
        chart_info_file = config.get('excel_file', '')
        chart_info_df = pd.DataFrame()
        
        if chart_info_file and os.path.exists(chart_info_file):
            try:
                chart_info_df = pd.read_excel(chart_info_file)
            except Exception as e:
                print(f"Warning: Failed to read chart info file: {e}")
        
        if not chart_info_df.empty:
            # 使用 chart info 模式
            valid_charts = chart_info_df.dropna(subset=['ITEM NAME'])
            total_count = len(valid_charts)
            
            for _, chart_row in valid_charts.iterrows():
                gname = str(chart_row.get('Group Name', ''))
                cname = str(chart_row.get('ITEM NAME', ''))
                characteristic = str(chart_row.get('Characteristic', 'Nominal'))
                
                # 篩選對應資料
                subdf = df[(df['gname'].astype(str) == gname) & 
                          (df['cname'].astype(str) == cname)]
                
                if subdf.empty:
                    continue
                
                # 群組統計
                group_stats = subdf.groupby("matching_group").agg({
                    "point_val": ["mean", "std", "count"]
                }).round(2)
                group_stats.columns = ["mean", "std", "count"]
                group_stats = group_stats.reset_index()
                group_stats = group_stats[group_stats["count"] >= 1]
                
                # 選擇分析方法
                num_groups = len(group_stats)
                analysis_mode = config.get('analysis_mode', 'default')
                
                if analysis_mode == 'time_analysis':
                    # 時間分析模式
                    mean_df = subdf[(subdf['time_diff_days'] >= -30) & 
                                   (subdf['time_diff_days'] <= 30)]
                    sigma_df = subdf[(subdf['time_diff_days'] >= -180) & 
                                    (subdf['time_diff_days'] <= 180)]
                    
                    if num_groups >= 3:
                        _analyze_multiple_groups_time_headless(
                            mean_df, sigma_df, group_stats, gname, cname, 
                            characteristic, results, config
                        )
                    elif num_groups == 2:
                        _analyze_two_groups_headless(
                            group_stats, gname, cname, characteristic, 
                            results, mean_df, config
                        )
                else:
                    # 標準分析模式
                    if num_groups >= 3:
                        _analyze_multiple_groups_headless(
                            subdf, group_stats, gname, cname, 
                            characteristic, results, config
                        )
                    elif num_groups == 2:
                        _analyze_two_groups_headless(
                            group_stats, gname, cname, characteristic, 
                            results, subdf, config
                        )
                
                processed_count += 1
        
        else:
            # 無 chart info 模式，分析所有 gname/cname 組合
            unique_combinations = df[['gname', 'cname']].drop_duplicates()
            total_count = len(unique_combinations)
            
            for _, row in unique_combinations.iterrows():
                gname, cname = str(row['gname']), str(row['cname'])
                characteristic = 'Nominal'  # 預設值
                
                subdf = df[(df['gname'].astype(str) == gname) & 
                          (df['cname'].astype(str) == cname)]
                
                if subdf.empty:
                    continue
                
                group_stats = subdf.groupby("matching_group").agg({
                    "point_val": ["mean", "std", "count"]
                }).round(2)
                group_stats.columns = ["mean", "std", "count"]
                group_stats = group_stats.reset_index()
                group_stats = group_stats[group_stats["count"] >= 1]
                
                num_groups = len(group_stats)
                analysis_mode = config.get('analysis_mode', 'default')
                
                if analysis_mode == 'time_analysis':
                    mean_df = subdf[(subdf['time_diff_days'] >= -30) & 
                                   (subdf['time_diff_days'] <= 30)]
                    sigma_df = subdf[(subdf['time_diff_days'] >= -180) & 
                                    (subdf['time_diff_days'] <= 180)]
                    
                    if num_groups >= 3:
                        _analyze_multiple_groups_time_headless(
                            mean_df, sigma_df, group_stats, gname, cname, 
                            characteristic, results, config
                        )
                    elif num_groups == 2:
                        _analyze_two_groups_headless(
                            group_stats, gname, cname, characteristic, 
                            results, mean_df, config
                        )
                else:
                    if num_groups >= 3:
                        _analyze_multiple_groups_headless(
                            subdf, group_stats, gname, cname, 
                            characteristic, results, config
                        )
                    elif num_groups == 2:
                        _analyze_two_groups_headless(
                            group_stats, gname, cname, characteristic, 
                            results, subdf, config
                        )
                
                processed_count += 1
        
        # 建立結果 DataFrame（與 PyQt 版本完全一致的欄位）
        if results:
            result_df = pd.DataFrame(results, columns=[
                'gname', 'cname', 'group', 'group_all',
                'mean_index', 'sigma_index', 'k_value', 
                'mean', 'std', 'mean_median', 'sigma_median', 
                'n', 'characteristic'
            ])
        else:
            result_df = pd.DataFrame(columns=[
                'gname', 'cname', 'group', 'group_all',
                'mean_index', 'sigma_index', 'k_value', 
                'mean', 'std', 'mean_median', 'sigma_median', 
                'n', 'characteristic'
            ])
        
        return {
            "summary": f"處理了 {processed_count} 個項目，生成 {len(results)} 筆結果",
            "results": result_df
        }
        
    except Exception as e:
        print(f"分析過程中發生錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "summary": f"分析失敗: {str(e)}",
            "results": pd.DataFrame()
        }


    if config.get("use_statistical_test", False):
        # 使用統計檢定時不檢查 Mean Index
        pass
    else:
        try:
            mean_threshold = config.get("mean_index_threshold", 1.0)
            if isinstance(mean_status, (int, float)) and mean_status >= mean_threshold:
                return True
        except:
            pass
    
    # 檢查 Sigma Index
    try:
        sigma_threshold = config.get("sigma_index_threshold", 2.0)
        if isinstance(sigma_index, (int, float)) and isinstance(k_value, (int, float)):
            if sigma_index >= k_value * sigma_threshold:
                return True
    except:
        pass
    
    return False


def _get_abnormal_type_headless(mean_status, sigma_index, k_value, config):
    """取得異常類型（無 UI 版本）"""
    abnormal_types = []
    
    # 檢查資料不足
    if (mean_status == 'Insufficient Data' or 
        sigma_index == 'Insufficient Data' or 
        k_value == 'No Compare'):
        return 'Insufficient Data'
    
    # 檢查統計檢定顯著
    if isinstance(mean_status, str) and "Significant" in mean_status:
        if "No Significant" not in mean_status:
            if "ANOVA" in mean_status:
                abnormal_types.append('ANOVA Significant')
            else:
                abnormal_types.append('Statistical Test Significant')
    
    # 檢查 Mean Index（非統計檢定模式）
    if not config.get("use_statistical_test", False):
        try:
            mean_threshold = config.get("mean_index_threshold", 1.0)
            if isinstance(mean_status, (int, float)) and mean_status >= mean_threshold:
                abnormal_types.append('Mean Not Matched')
        except:
            pass
    
    # 檢查 Sigma Index
    try:
        sigma_threshold = config.get("sigma_index_threshold", 2.0)
        if isinstance(sigma_index, (int, float)) and isinstance(k_value, (int, float)):
            if sigma_index >= k_value * sigma_threshold:
                abnormal_types.append('Sigma Not Matched')
    except:
        pass
    
    if not abnormal_types:
        return 'Normal'
    
    return ', '.join(abnormal_types)


# ==================================================
# 批次任務模式核心分析函數（OOB Excel + 資料夾）
# ==================================================

def analyze_tool_matching_data(all_charts_info: "pd.DataFrame", raw_data_directory: str, config: dict) -> dict:
    """
    批次任務模式：遍歷 all_charts_info 中每個 (GroupName, ChartName)，
    在 raw_data_directory 找對應 CSV，執行 30 天 window 三段式補點分析（Dormant/Force Fill/Normal）。

    Args:
        all_charts_info: DataFrame with [GroupName, ChartName, Characteristics]
        raw_data_directory: str, 含原始 CSV 的資料夾路徑
        config: dict，可包含 base_date, fill_sample_size, mean_index_threshold,
                sigma_index_threshold, use_statistical_test, statistical_method, alpha_level

    Returns:
        dict: {
            "summary": {"total_groups": int, "abnormal_groups": int},
            "results": pd.DataFrame[gname, cname, group, group_all,
                        mean_index, sigma_index, k_value, mean, std,
                        mean_median, sigma_median, n, characteristic]
        }
    """
    try:
        from oob_eng import find_matching_file as _find_file
    except ImportError:
        _find_file = None

    base_date_raw = config.get('base_date')
    fill_num = int(config.get('fill_sample_size', 5))
    base_date = pd.Timestamp(base_date_raw) if base_date_raw else pd.Timestamp.now().normalize()

    results = []
    filtered_data = {}  # (gname, cname) -> mean_df（三段式過濾後，供圖表使用）

    for _, chart_row in all_charts_info.iterrows():
        gname = str(chart_row.get('GroupName', '')).strip()
        cname = str(chart_row.get('ChartName', '')).strip()
        characteristic = str(
            chart_row.get('Characteristics') or chart_row.get('Characteristic') or 'Nominal'
        ).strip()

        if not gname or not cname:
            continue

        # 尋找對應 CSV
        csv_path = None
        if _find_file is not None:
            try:
                csv_path = _find_file(raw_data_directory, gname, cname)
            except Exception:
                pass

        if not csv_path or not os.path.isfile(csv_path):
            continue

        # 讀取 CSV
        try:
            subdf = pd.read_csv(csv_path)
        except Exception as e:
            print(f"[WARNING] Failed to read {csv_path}: {e}")
            continue

        # 移除 point_time 無效行並轉換為 datetime
        if 'point_time' not in subdf.columns:
            print(f"[WARNING] {gname}/{cname}: No point_time column")
            continue
        subdf = subdf.dropna(subset=['point_time']).copy()
        try:
            subdf['point_time'] = pd.to_datetime(subdf['point_time'])
        except Exception:
            continue
        subdf = subdf[subdf['point_time'].notna()].copy()

        # 自動偵測群組欄位：ByTool > EQP_id > Matching > matching_group
        if 'matching_group' not in subdf.columns:
            for possible_col in ['ByTool', 'EQP_id', 'Tool', 'tool_id', 'Matching']:
                if possible_col in subdf.columns:
                    subdf = subdf.rename(columns={possible_col: 'matching_group'})
                    break
                    
        if 'matching_group' not in subdf.columns:
            print(f"[WARNING] {gname}/{cname}: No group column found")
            continue
        if 'point_val' not in subdf.columns:
            print(f"[WARNING] {gname}/{cname}: No point_val column")
            continue

        all_groups = subdf['matching_group'].unique()

        # 安全開關：總資料量 < 5 筆則全部標記 Insufficient Data
        if len(subdf) < 5:
            for mg in all_groups:
                results.append([
                    gname, cname, str(mg), 'group_all',
                    'Insufficient Data', 'Insufficient Data', 'No Compare',
                    0.0, 0.0, '-', '-',
                    int(len(subdf[subdf['matching_group'] == mg])), characteristic
                ])
            continue

        # 三段式補點（1M window）
        mean_end = base_date
        mean_start = mean_end - pd.DateOffset(days=30)
        week_start = mean_end - pd.Timedelta(days=7)
        MIN_FILL = 5
        # 先抓 30 天初始區間（mean 與 sigma 共用同一資料集）
        mean_df = subdf[(subdf['point_time'] > mean_start) & (subdf['point_time'] <= mean_end)].copy()
        sigma_df = mean_df.copy()
        for mg in all_groups:
            mg_week = subdf[
                (subdf['matching_group'] == mg) &
                (subdf['point_time'] > week_start) &
                (subdf['point_time'] <= mean_end)
            ]
            if len(mg_week) == 0:
                # Dormant：當週無資料，完全排除此 group
                mean_df = mean_df[mean_df['matching_group'] != mg]
                sigma_df = sigma_df[sigma_df['matching_group'] != mg]
                continue
            mg_30d = mean_df[mean_df['matching_group'] == mg]
            if len(mg_30d) < MIN_FILL:
                # 強制補足：取歷史最新 MIN_FILL 筆
                mg_fill = (
                    subdf[subdf['matching_group'] == mg]
                    .sort_values('point_time', ascending=False)
                    .head(MIN_FILL)
                )
                mean_df = pd.concat([mean_df[mean_df['matching_group'] != mg], mg_fill]).drop_duplicates()
                sigma_df = pd.concat([sigma_df[sigma_df['matching_group'] != mg], mg_fill]).drop_duplicates()
            # else: 正常使用 30 天內所有資料

        # 全 Dormant 偵測：若所有 group 近 7 天都無資料，mean_df 為空
        if mean_df.empty:
            group_all_str = '/'.join(str(g) for g in all_groups)
            for mg in all_groups:
                results.append([
                    gname, cname, str(mg), group_all_str,
                    'Insufficient Data', 'Insufficient Data', 'No Compare',
                    0.0, 0.0, '-', '-', 0, characteristic
                ])
            print(f"[INFO] {gname}-{cname}: all groups have no data in last 7 days, skipped.")
            continue

        # 儲存三段式過濾後的資料（供圖表使用，資料範圍與統計計算一致）
        filtered_data[(gname, cname)] = mean_df.copy()

        # 計算統計量
        mean_stats = mean_df.groupby('matching_group')['point_val'].agg(['mean', 'count']).reset_index()
        sigma_stats = sigma_df.groupby('matching_group')['point_val'].agg(['std']).reset_index()
        group_stats = pd.merge(mean_stats, sigma_stats, on='matching_group', how='outer')
        group_stats = group_stats.fillna({'mean': 0.0, 'std': 0.0, 'count': 0})

        n_groups = len(group_stats)
        valid_groups_df = group_stats[group_stats['count'] >= 5]
        n_valid = len(valid_groups_df)

        row_results = []
        if n_valid == 2 and n_groups > 2:
            _analyze_two_groups_time_headless(
                mean_df, sigma_df, valid_groups_df, gname, cname, characteristic, row_results, config)
        elif n_groups == 2:
            _analyze_two_groups_time_headless(
                mean_df, sigma_df, group_stats, gname, cname, characteristic, row_results, config)
        else:
            _analyze_multiple_groups_time_headless(
                mean_df, sigma_df, group_stats, gname, cname, characteristic, row_results, config)
        results.extend(row_results)

    _COLS = ['gname', 'cname', 'group', 'group_all', 'mean_index', 'sigma_index',
             'k_value', 'mean', 'std', 'mean_median', 'sigma_median', 'n', 'characteristic']
    result_df = pd.DataFrame(results, columns=_COLS) if results else pd.DataFrame(columns=_COLS)

    mean_th = config.get('mean_index_threshold', 1.0)
    sigma_th_cfg = config.get('sigma_index_threshold', 2.0)
    use_stats = config.get('use_statistical_test', False)

    def _is_abnormal_row(mi, si, kv):
        if mi == 'Insufficient Data' or si == 'Insufficient Data' or kv == 'No Compare':
            return False
        if use_stats:
            return isinstance(mi, str) and 'Significant' in mi and 'No Significant' not in mi
        try:
            sigma_th = float(kv) if isinstance(kv, (int, float)) else sigma_th_cfg
            return (isinstance(mi, (int, float)) and float(mi) >= mean_th) or \
                   (isinstance(si, (int, float)) and float(si) >= sigma_th)
        except Exception:
            return False

    result_df['is_abnormal'] = result_df.apply(
        lambda r: _is_abnormal_row(r.get('mean_index'), r.get('sigma_index'), r.get('k_value')), axis=1
    )
    abnormal_count = int(result_df['is_abnormal'].sum())

    return {
        "summary": {"total_groups": len(result_df), "abnormal_groups": abnormal_count},
        "results": result_df,
        "filtered_data": filtered_data,
    }


def _create_spc_chart(
    subdf: "pd.DataFrame", gname: str, cname: str,
    output_dir: str = "output", return_bytes: bool = False
):
    """生成 SPC 散佈圖。
    - return_bytes=False：儲存到 output_dir，回傳檔案路徑 (str)。
    - return_bytes=True：不儲檔，直接回傳 bytes。
    兩者都會在 finally 執行 plt.close(fig) 防止記憶體溢出。
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib import cm
    from io import BytesIO

    unique_groups = sorted(subdf["matching_group"].unique(), key=lambda x: str(x))
    n_g = max(len(unique_groups), 1)
    colors = cm.get_cmap('tab20')(np.linspace(0, 1, n_g))

    fig, ax = plt.subplots(figsize=(9, 4.5))
    try:
        import pandas as pd
        import matplotlib.dates as mdates

        has_time = "point_time" in subdf.columns
        if has_time:
            subdf = subdf.copy()
            subdf["point_time"] = pd.to_datetime(subdf["point_time"], errors="coerce")
            subdf = subdf.sort_values("point_time")

        for i, mg in enumerate(unique_groups):
            grp = subdf[subdf["matching_group"] == mg]
            if grp.empty:
                continue
            if has_time:
                x_vals = grp["point_time"]
            else:
                x_vals = grp.index
            y_vals = grp["point_val"].values
            ax.scatter(x_vals, y_vals, color=colors[i], alpha=0.8, s=40, label=str(mg), zorder=3)
            ax.plot(x_vals, y_vals, color=colors[i], alpha=0.5, linewidth=1, zorder=2)

        ax.set_title(f"SPC Chart: {gname} - {cname}", fontsize=10)
        ax.set_xlabel("Time" if has_time else "Sample Index")
        ax.set_ylabel("Point Value")
        ax.grid(True, linestyle='--', alpha=0.3, zorder=0)
        if has_time:
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
            fig.autofmt_xdate(rotation=45)
        ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
        fig.tight_layout()
        if return_bytes:
            buf = BytesIO()
            fig.savefig(buf, format='png', bbox_inches='tight', dpi=100)
            return buf.getvalue()
        else:
            os.makedirs(output_dir, exist_ok=True)
            safe = lambda s: str(s).replace('/', '_').replace(' ', '_')
            out_path = os.path.join(output_dir, f"{safe(gname)}_{safe(cname)}_spc.png")
            fig.savefig(out_path, format='png', bbox_inches='tight', dpi=100)
            return out_path
    finally:
        plt.close(fig)


def _create_boxplot_chart(
    subdf: "pd.DataFrame", gname: str, cname: str,
    output_dir: str = "output", return_bytes: bool = False
):
    """生成盒鬚圖。
    - return_bytes=False：儲存到 output_dir，回傳檔案路徑 (str)。
    - return_bytes=True：不儲檔，直接回傳 bytes。
    兩者都會在 finally 執行 plt.close(fig) 防止記憶體溢出。
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib import cm
    from io import BytesIO

    unique_groups = sorted(subdf["matching_group"].unique(), key=lambda x: str(x))
    labels = [str(mg) for mg in unique_groups]
    n_g = max(len(unique_groups), 1)
    colors = cm.get_cmap('tab20')(np.linspace(0, 1, n_g))
    box_data = [subdf[subdf["matching_group"] == mg]["point_val"].values for mg in unique_groups]
    group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count'])

    fig, ax = plt.subplots(figsize=(7, 4.5))
    try:
        if box_data and any(len(d) > 0 for d in box_data):
            bp = ax.boxplot(box_data, labels=labels, patch_artist=True, widths=0.6)
            for patch, color in zip(bp['boxes'], colors):
                patch.set_facecolor(color)
            legend_labels = []
            for label, mg in zip(labels, unique_groups):
                if mg in group_stats.index:
                    s = group_stats.loc[mg]
                    legend_labels.append(
                        f"{label}: \u03bc={s['mean']:.2f}, \u03c3={s['std']:.2f}, n={int(s['count'])}"
                    )
                else:
                    legend_labels.append(label)
            ax.legend([bp["boxes"][i] for i in range(len(labels))], legend_labels,
                      loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
        ax.set_title(f"Boxplot: {gname} - {cname}", fontsize=10)
        ax.set_xlabel("Matching Group")
        ax.set_ylabel("Point Value")
        ax.grid(True, linestyle='--', alpha=0.6)
        fig.tight_layout()
        if return_bytes:
            buf = BytesIO()
            fig.savefig(buf, format='png', bbox_inches='tight', dpi=100)
            return buf.getvalue()
        else:
            os.makedirs(output_dir, exist_ok=True)
            safe = lambda s: str(s).replace('/', '_').replace(' ', '_')
            out_path = os.path.join(output_dir, f"{safe(gname)}_{safe(cname)}_box.png")
            fig.savefig(out_path, format='png', bbox_inches='tight', dpi=100)
            return out_path
    finally:
        plt.close(fig)


# 只在有 UI 且直接執行時才啟動 GUI
if __name__ == "__main__" and UI_AVAILABLE:
    import sys
    app = QtWidgets.QApplication(sys.argv)
    widget = ToolMatchingWidget()
    widget.show()
    sys.exit(app.exec())